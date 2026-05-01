"""Pricing agent tool executors — currency conversion and config read."""

from __future__ import annotations

import io
import time
import uuid
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook

from app.core.encryption import decrypt_credentials
from app.mcp.tools.sheets_tools import _get_sheets_connector, _get_user_email
from app.models.pricing_conversion_log import PricingConversionLog
from app.schemas.pricing import (
    CurrencyConfig,
    PricingInput,
    TenantPricingConfig,
)
from app.services.chat.result_cache import get_latest_result_by_type
from app.services.pricing_config_service import get_config, upsert_config
from app.services.pricing_engine import PricingEngine
from app.services.sheets_service import (
    create_spreadsheet,
    share_spreadsheet,
    write_range,
)
from app.services.task_file_service import TaskFileService
from app.services.template_filler import TemplateFiller

_engine = PricingEngine()
_filler = TemplateFiller()
_file_svc = TaskFileService()

_CACHE_MISS_MESSAGE = (
    "No prior pricing state in this conversation (cache expired or no pricing run yet). "
    "Re-run pricing_convert with the upload, or pricing_export with inline items."
)


def _build_preview(results, max_rows: int = 10) -> list[dict]:
    """Frontend-friendly preview: flat dict per row with sorted currency keys."""
    preview: list[dict] = []
    for r in results[:max_rows]:
        row = {"SKU": r.sku, "USD": float(r.usd_price)}
        for code, cr in sorted(r.results.items()):
            row[code] = float(cr.final_price)
        preview.append(row)
    return preview


def _seed_pricing_state(
    *,
    items: list[PricingInput],
    pricing_config: TenantPricingConfig,
    excel_file_id: str,
    netsuite_csv_file_id: str,
    row_count: int,
) -> dict:
    """Build the initial pricing_state payload that will be cached for follow-ups.

    Holds only what pricing_revise / pricing_to_sheets need: the seed inputs,
    effective inputs/currencies/overrides (all empty on the seed run), and
    pointers to the saved Excel + NetSuite CSV files. Row data is NOT cached
    — pricing_to_sheets re-parses the Excel file on demand.
    """
    seed_items = [
        {"sku": it.sku, "usd_price": str(it.usd_price), "item_name": it.item_name}
        for it in items
    ]
    currency_codes = list(pricing_config.currencies.keys())
    header_columns = ["SKU", "Item Name", "USD", *sorted(currency_codes)]
    return {
        "seed_items": seed_items,
        "effective_items": [dict(s) for s in seed_items],
        "effective_currencies": currency_codes,
        "effective_fx_overrides": {},
        "effective_vat_overrides": {},
        "effective_rounding_overrides": {},
        "effective_uplift_by_currency": {},
        "applied_overrides_log": [],
        "excel_file_id": excel_file_id,
        "netsuite_csv_file_id": netsuite_csv_file_id,
        "header_columns": header_columns,
        "row_count": row_count,
    }


async def pricing_convert_execute(params: dict, context: dict, **kwargs) -> dict:
    """Convert prices in uploaded Excel using tenant FX rates."""
    if not context or not context.get("db") or not context.get("tenant_id"):
        return {"error": True, "message": "Missing context — tenant_id and db are required."}
    db = context["db"]
    tenant_id = context["tenant_id"]
    user_id = context.get("actor_id") or context.get("user_id")

    # 1. Load config
    config_row = await get_config(db, tenant_id)
    if not config_row:
        return {"error": True, "message": "No pricing configuration found. Set up FX rates in Settings."}
    pricing_config = TenantPricingConfig(**config_row.config)

    # 2. Load file
    file_id = params.get("file_id")
    if not file_id:
        return {"error": True, "message": "file_id is required"}
    try:
        task_file, content = await _file_svc.get_file(db, tenant_id, uuid.UUID(file_id))
    except ValueError:
        return {"error": True, "message": f"File not found: {file_id}"}

    # 3. Parse Excel
    wb = load_workbook(io.BytesIO(content))
    mapping = _filler.detect_columns(wb)
    ws = wb.active
    items = []
    for row in range(2, ws.max_row + 1):
        sku_val = ws.cell(row=row, column=mapping.sku_col + 1).value
        price_val = ws.cell(row=row, column=mapping.price_col + 1).value
        if not sku_val or not price_val:
            continue
        items.append(PricingInput(sku=str(sku_val).strip(), usd_price=Decimal(str(price_val))))
    if not items:
        return {"error": True, "message": "No valid rows found. Need SKU and USD Price columns."}

    # 4. Convert
    results = _engine.convert_batch(items, pricing_config)

    # 5. Generate output
    output_files = {}
    if mapping.currency_cols:
        _filler.fill(wb, results, mapping)
        buf = io.BytesIO()
        wb.save(buf)
        output_content = buf.getvalue()
    else:
        out_wb = _filler.generate_default_output(results)
        buf = io.BytesIO()
        out_wb.save(buf)
        output_content = buf.getvalue()

    excel_file = await _file_svc.save_output(
        db=db,
        tenant_id=tenant_id,
        user_id=user_id,
        filename=f"pricing-{task_file.filename}",
        content=output_content,
    )
    output_files["excel"] = str(excel_file.id)

    # 6. NetSuite CSV (always generated)
    csv_str = _filler.generate_netsuite_csv(results)
    csv_file = await _file_svc.save_output(
        db=db,
        tenant_id=tenant_id,
        user_id=user_id,
        filename=f"netsuite-import-{task_file.filename.rsplit('.', 1)[0]}.csv",
        content=csv_str.encode("utf-8"),
    )
    output_files["netsuite_csv"] = str(csv_file.id)

    # 7. Audit log

    db.add(
        PricingConversionLog(
            tenant_id=tenant_id,
            user_id=user_id,
            input_file_id=uuid.UUID(file_id),
            output_file_id=excel_file.id,
            sku_count=len(items),
            currency_count=len(pricing_config.currencies),
            config_snapshot=pricing_config.model_dump(mode="json"),
        )
    )

    pricing_state = _seed_pricing_state(
        items=items,
        pricing_config=pricing_config,
        excel_file_id=str(excel_file.id),
        netsuite_csv_file_id=str(csv_file.id),
        row_count=len(items),
    )

    return {
        "success": True,
        "sku_count": len(items),
        "currency_count": len(pricing_config.currencies),
        "output_files": output_files,
        "preview": _build_preview(results),
        "pricing_state": pricing_state,
        "template_mode": bool(mapping.currency_cols),
    }


async def pricing_export_execute(params: dict, context: dict, **kwargs) -> dict:
    """Export computed prices to downloadable Excel from inline data (no upload required)."""
    if not context or not context.get("db") or not context.get("tenant_id"):
        return {"error": True, "message": "Missing context — tenant_id and db are required."}
    db = context["db"]
    tenant_id = context["tenant_id"]
    user_id = context.get("actor_id") or context.get("user_id")

    # 1. Load config
    config_row = await get_config(db, tenant_id)
    if not config_row:
        return {"error": True, "message": "No pricing configuration found. Set up FX rates in Settings."}
    pricing_config = TenantPricingConfig(**config_row.config)

    # 2. Parse inline items
    raw_items = params.get("items")
    if not raw_items or not isinstance(raw_items, list):
        return {"error": True, "message": "items is required (list of {sku, usd_price, item_name?})"}
    items = []
    for raw in raw_items:
        sku = raw.get("sku")
        price = raw.get("usd_price")
        if not sku or price is None:
            continue
        items.append(
            PricingInput(
                sku=str(sku).strip(),
                item_name=raw.get("item_name"),
                usd_price=Decimal(str(price)),
            )
        )
    if not items:
        return {"error": True, "message": "No valid items. Each needs at least sku and usd_price."}

    # 3. Convert
    results = _engine.convert_batch(items, pricing_config)

    # 4. Generate output Excel (default 3-sheet workbook)
    out_wb = _filler.generate_default_output(results)
    buf = io.BytesIO()
    out_wb.save(buf)
    output_content = buf.getvalue()

    output_files = {}
    excel_file = await _file_svc.save_output(
        db=db,
        tenant_id=tenant_id,
        user_id=user_id,
        filename=f"pricing-export-{len(items)}-skus.xlsx",
        content=output_content,
    )
    output_files["excel"] = str(excel_file.id)

    # 5. NetSuite CSV (always generated)
    csv_str = _filler.generate_netsuite_csv(results)
    csv_file = await _file_svc.save_output(
        db=db,
        tenant_id=tenant_id,
        user_id=user_id,
        filename=f"netsuite-import-{len(items)}-skus.csv",
        content=csv_str.encode("utf-8"),
    )
    output_files["netsuite_csv"] = str(csv_file.id)

    # 6. Audit log

    db.add(
        PricingConversionLog(
            tenant_id=tenant_id,
            user_id=user_id,
            input_file_id=None,
            output_file_id=excel_file.id,
            sku_count=len(items),
            currency_count=len(pricing_config.currencies),
            config_snapshot=pricing_config.model_dump(mode="json"),
        )
    )

    pricing_state = _seed_pricing_state(
        items=items,
        pricing_config=pricing_config,
        excel_file_id=str(excel_file.id),
        netsuite_csv_file_id=str(csv_file.id),
        row_count=len(items),
    )

    return {
        "success": True,
        "sku_count": len(items),
        "currency_count": len(pricing_config.currencies),
        "output_files": output_files,
        "preview": _build_preview(results),
        "pricing_state": pricing_state,
        "template_mode": False,
    }


async def pricing_config_read_execute(params: dict, context: dict, **kwargs) -> dict:
    """Read current tenant pricing configuration."""
    if not context or not context.get("db") or not context.get("tenant_id"):
        return {"error": True, "message": "Missing context — tenant_id and db are required."}
    db = context["db"]
    tenant_id = context["tenant_id"]
    config_row = await get_config(db, tenant_id)
    if not config_row:
        return {"error": True, "message": "No pricing configuration set."}
    config = TenantPricingConfig(**config_row.config)
    currencies = {}
    for code, cc in config.currencies.items():
        currencies[code] = {
            "fx_rate": float(cc.fx_rate),
            "tier": cc.tier,
            "vat_rate": float(cc.vat_rate) if cc.vat_rate else None,
            "rounding_rule": cc.rounding_rule,
        }
    return {
        "success": True,
        "base_currency": config.base_currency,
        "eur_fx_rate": float(config.eur_fx_rate),
        "currency_count": len(currencies),
        "currencies": currencies,
    }


async def pricing_config_update_execute(params: dict, context: dict, **kwargs) -> dict:
    """Update tenant pricing configuration — FX rates, VAT, rounding rules.

    Accepts partial updates: only the fields provided will be changed.
    """
    if not context or not context.get("db") or not context.get("tenant_id"):
        return {"error": True, "message": "Missing context — tenant_id and db are required."}

    db = context["db"]
    tenant_id = context["tenant_id"]
    user_id = context.get("user_id")

    # Read current config
    config_row = await get_config(db, tenant_id)
    if not config_row:
        return {"error": True, "message": "No pricing configuration exists. Create one in Settings first."}

    current_config = dict(config_row.config)

    # Apply updates
    updates = params.get("updates", {})
    if not updates:
        return {"error": True, "message": "No updates provided. Pass 'updates' with fields to change."}

    # Update EUR FX rate
    if "eur_fx_rate" in updates:
        current_config["eur_fx_rate"] = float(updates["eur_fx_rate"])

    # Update individual currency configs
    if "currencies" in updates:
        existing_currencies = current_config.get("currencies", {})
        for code, changes in updates["currencies"].items():
            code_upper = code.upper()
            if code_upper not in existing_currencies:
                existing_currencies[code_upper] = {}
            for field in ("fx_rate", "vat_rate", "rounding_rule", "tier"):
                if field in changes:
                    existing_currencies[code_upper][field] = changes[field]
        current_config["currencies"] = existing_currencies

    # Save
    await upsert_config(db, tenant_id, current_config, user_id)
    await db.commit()

    return {
        "success": True,
        "message": f"Pricing config updated: {list(updates.keys())}",
        "updated_fields": list(updates.keys()),
    }


# ---------------------------------------------------------------------------
# pricing_revise — follow-up edits to a prior pricing run
# ---------------------------------------------------------------------------


def _coerce_bool(v: Any) -> bool:
    """Strict bool coercion that won't be fooled by ``"false"`` strings.

    Plain ``bool(v)`` is dangerous here: ``bool("false")`` is ``True``. If a
    chat tool call ever surfaces ``reset`` as the string ``"false"`` (some
    JSON-Schema-to-tool-input adapters round-trip booleans through strings),
    ``bool(v)`` would silently discard the user's overrides and regenerate
    from the seed. This helper treats only the values the LLM/JSON spec
    actually means by "true" as truthy.
    """
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        return v.strip().lower() in {"true", "1", "yes"}
    if isinstance(v, (int, float)):
        return bool(v)
    return False


def _to_decimal_str(v: Any) -> str:
    """Coerce a JSON-supplied number to a stringified Decimal for cache storage.

    Decimal(str(v)) is the spec's precision rule — float(0.1+0.2) becomes
    Decimal("0.30000000000000004"), not Decimal("0.3"). We store as str so the
    payload survives JSON round-trips without losing precision.
    """
    return str(Decimal(str(v)))


def _items_from_state(items: list[dict]) -> list[PricingInput]:
    out: list[PricingInput] = []
    for raw in items:
        try:
            price = Decimal(str(raw.get("usd_price", 0)))
        except (InvalidOperation, ValueError, TypeError):
            continue
        out.append(
            PricingInput(
                sku=str(raw["sku"]),
                item_name=raw.get("item_name"),
                usd_price=price,
            )
        )
    return out


def _apply_overrides(state: dict, overrides: dict) -> dict | None:
    """Mutate effective_* in place per spec (Sec. "Cumulative semantics" + D3).

    Order: skus_to_remove → skus_to_add → sku_price_changes → currency ops →
    uplift / FX / VAT / rounding (replace per key).

    Returns an error dict if validation fails, else None.
    """
    log_entry: dict[str, Any] = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "ignored": [],
    }

    # 1. skus_to_remove
    to_remove = {str(s) for s in (overrides.get("skus_to_remove") or [])}
    if to_remove:
        kept: list[dict] = []
        for it in state["effective_items"]:
            if it["sku"] in to_remove:
                continue
            kept.append(it)
        state["effective_items"] = kept
        log_entry["skus_to_remove"] = sorted(to_remove)

    # 2. skus_to_add (silent no-op for already present)
    to_add = overrides.get("skus_to_add") or []
    if to_add:
        existing = {it["sku"] for it in state["effective_items"]}
        added_log: list[str] = []
        for raw in to_add:
            sku = str(raw["sku"])
            if sku in existing:
                log_entry["ignored"].append({"action": "skus_to_add", "sku": sku, "reason": "already present"})
                continue
            state["effective_items"].append(
                {
                    "sku": sku,
                    "usd_price": _to_decimal_str(raw["usd_price"]),
                    "item_name": raw.get("item_name"),
                }
            )
            existing.add(sku)
            added_log.append(sku)
        if added_log:
            log_entry["skus_to_add"] = added_log

    # 3. sku_price_changes (latest-wins, silent no-op for unknown SKU)
    price_changes = overrides.get("sku_price_changes") or []
    if price_changes:
        by_sku: dict[str, str] = {}
        # Latest-wins within a single revise call.
        for raw in price_changes:
            by_sku[str(raw["sku"])] = _to_decimal_str(raw["usd_price"])
        applied_log: list[dict] = []
        for it in state["effective_items"]:
            if it["sku"] in by_sku:
                it["usd_price"] = by_sku[it["sku"]]
                applied_log.append({"sku": it["sku"], "usd_price": it["usd_price"]})
        # Silent no-op for SKUs that didn't match.
        unmatched = [s for s in by_sku if not any(it["sku"] == s for it in state["effective_items"])]
        for s in unmatched:
            log_entry["ignored"].append({"action": "sku_price_changes", "sku": s, "reason": "sku not in effective set"})
        if applied_log:
            log_entry["sku_price_changes"] = applied_log

    if not state["effective_items"]:
        return {
            "error": True,
            "message": "No SKUs left after removals. Use reset=true to start over, or add SKUs back.",
        }

    # 4. Currency ops
    currencies_to_remove = overrides.get("currencies_to_remove") or []
    if currencies_to_remove:
        state["effective_currencies"] = [
            c for c in state["effective_currencies"] if c not in set(currencies_to_remove)
        ]
        log_entry["currencies_to_remove"] = list(currencies_to_remove)

    # currencies_to_add is validated + appended by the caller after this returns
    # (needs fresh tenant config to check, and must run after _apply_overrides).

    # 5. Replace-per-key dicts
    for key in ("percent_uplift", "fx_rate_overrides", "vat_rate_overrides"):
        incoming = overrides.get(key) or {}
        if not incoming:
            continue
        target_key = {
            "percent_uplift": "effective_uplift_by_currency",
            "fx_rate_overrides": "effective_fx_overrides",
            "vat_rate_overrides": "effective_vat_overrides",
        }[key]
        merged = dict(state.get(target_key) or {})
        for cur, val in incoming.items():
            merged[cur] = _to_decimal_str(val)
        state[target_key] = merged
        log_entry[key] = {k: _to_decimal_str(v) for k, v in incoming.items()}

    rounding = overrides.get("rounding_overrides") or {}
    if rounding:
        merged_r = dict(state.get("effective_rounding_overrides") or {})
        for cur, rule in rounding.items():
            merged_r[cur] = str(rule)
        state["effective_rounding_overrides"] = merged_r
        log_entry["rounding_overrides"] = dict(rounding)

    # Strip empty ignored if nothing was logged.
    if not log_entry["ignored"]:
        log_entry.pop("ignored")

    state["applied_overrides_log"].append(log_entry)
    return None


def _compose_effective_config(
    base_config: TenantPricingConfig,
    *,
    fx_overrides: dict[str, str],
    vat_overrides: dict[str, str],
    rounding_overrides: dict[str, str],
    effective_currencies: list[str],
) -> TenantPricingConfig:
    """Return a new TenantPricingConfig with overrides applied + currencies filtered."""
    base_dump = base_config.model_dump(mode="json")
    filtered: dict[str, dict] = {}
    for code, cc in base_dump["currencies"].items():
        if code not in effective_currencies:
            continue
        cc_copy = dict(cc)
        if code in fx_overrides:
            cc_copy["fx_rate"] = fx_overrides[code]
        if code in vat_overrides:
            cc_copy["vat_rate"] = vat_overrides[code]
        if code in rounding_overrides:
            cc_copy["rounding_rule"] = rounding_overrides[code]
        filtered[code] = cc_copy
    base_dump["currencies"] = filtered
    return TenantPricingConfig(**base_dump)


def _state_uplift_to_decimals(uplift: dict[str, str]) -> dict[str, Decimal]:
    out: dict[str, Decimal] = {}
    for cur, val in (uplift or {}).items():
        try:
            out[cur] = Decimal(str(val))
        except (InvalidOperation, ValueError, TypeError):
            continue
    return out


async def pricing_revise_execute(params: dict, context: dict, **kwargs) -> dict:
    """Apply override edits to the prior pricing run and regenerate outputs.

    Reads the latest cached pricing payload, mutates effective_* state per the
    spec's cumulative + replace-per-key rules, fetches a FRESH tenant config
    every call (no stale FX from cache), runs the engine with the merged
    uplift_by_currency, writes new Excel + NetSuite CSV outputs, and returns
    a fresh pricing_state dict the orchestrator interceptor pipes back into
    the cache.
    """
    if not context or not context.get("db") or not context.get("tenant_id"):
        return {"error": True, "message": "Missing context — tenant_id and db are required."}
    db = context["db"]
    tenant_id = context["tenant_id"]
    user_id = context.get("actor_id") or context.get("user_id")
    conversation_id = context.get("conversation_id")
    if not conversation_id:
        return {"error": True, "message": "Missing conversation_id — pricing_revise needs a session-scoped cache."}

    # 1. Read latest pricing payload from the typed cache.
    cached = await get_latest_result_by_type(str(conversation_id), "pricing")
    if not cached or not cached.payload:
        return {"error": True, "message": _CACHE_MISS_MESSAGE}
    state = dict(cached.payload)
    state.setdefault("seed_items", [])
    state.setdefault("effective_items", list(state["seed_items"]))
    state.setdefault("effective_currencies", [])
    state.setdefault("effective_fx_overrides", {})
    state.setdefault("effective_vat_overrides", {})
    state.setdefault("effective_rounding_overrides", {})
    state.setdefault("effective_uplift_by_currency", {})
    state.setdefault("applied_overrides_log", [])

    # 2. Always fetch a FRESH tenant config (Spec D4 — never cache base config).
    config_row = await get_config(db, tenant_id)
    if not config_row:
        return {"error": True, "message": "No pricing configuration found. Set up FX rates in Settings."}
    current_config = TenantPricingConfig(**config_row.config)

    overrides = params.get("overrides") or {}
    reset = _coerce_bool(params.get("reset"))

    # 3. Reset path — drop accumulated overrides, restore effective_items from seed.
    if reset:
        state["effective_items"] = [dict(it) for it in state["seed_items"]]
        state["effective_currencies"] = list(current_config.currencies.keys())
        state["effective_fx_overrides"] = {}
        state["effective_vat_overrides"] = {}
        state["effective_rounding_overrides"] = {}
        state["effective_uplift_by_currency"] = {}
        state["applied_overrides_log"].append(
            {
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "reset": True,
            }
        )
    else:
        # 4a. Validate currencies_to_add against the FRESH config.
        for cur in overrides.get("currencies_to_add") or []:
            if cur not in current_config.currencies:
                return {
                    "error": True,
                    "message": (
                        f"Currency {cur} is not configured for this tenant. "
                        "Add it in Settings → Pricing first, then retry."
                    ),
                }
        # 4b. Apply overrides (mutates state, may return error dict).
        err = _apply_overrides(state, overrides)
        if err is not None:
            return err
        # 4c. Append currencies_to_add now (after validation passed).
        for cur in overrides.get("currencies_to_add") or []:
            if cur not in state["effective_currencies"]:
                state["effective_currencies"].append(cur)

    # 5. Compose the effective config for the engine call.
    effective_config = _compose_effective_config(
        current_config,
        fx_overrides=state["effective_fx_overrides"],
        vat_overrides=state["effective_vat_overrides"],
        rounding_overrides=state["effective_rounding_overrides"],
        effective_currencies=state["effective_currencies"],
    )

    # 6. Convert.
    items = _items_from_state(state["effective_items"])
    if not items:
        return {
            "error": True,
            "message": "No SKUs left after removals. Use reset=true to start over, or add SKUs back.",
        }
    uplift_by_currency = _state_uplift_to_decimals(state["effective_uplift_by_currency"])
    results = _engine.convert_batch(items, effective_config, uplift_by_currency=uplift_by_currency)

    # 7. Generate output Excel (default 3-sheet workbook).
    out_wb = _filler.generate_default_output(results)
    buf = io.BytesIO()
    out_wb.save(buf)
    output_content = buf.getvalue()
    suffix = int(time.time())

    output_files: dict[str, str] = {}
    excel_file = await _file_svc.save_output(
        db=db,
        tenant_id=tenant_id,
        user_id=user_id,
        filename=f"pricing-revised-{len(items)}-skus-{suffix}.xlsx",
        content=output_content,
    )
    output_files["excel"] = str(excel_file.id)

    # 8. NetSuite CSV.
    csv_str = _filler.generate_netsuite_csv(results)
    csv_file = await _file_svc.save_output(
        db=db,
        tenant_id=tenant_id,
        user_id=user_id,
        filename=f"netsuite-import-revised-{len(items)}-skus-{suffix}.csv",
        content=csv_str.encode("utf-8"),
    )
    output_files["netsuite_csv"] = str(csv_file.id)

    # 9. Audit log — snapshot the FRESH config + the cumulative override log so
    # the audit trail captures both the original seed run AND every revise's
    # config-at-time-of-run.
    snapshot = effective_config.model_dump(mode="json")
    snapshot["applied_overrides_log"] = state["applied_overrides_log"]
    db.add(
        PricingConversionLog(
            tenant_id=tenant_id,
            user_id=user_id,
            input_file_id=None,
            output_file_id=excel_file.id,
            sku_count=len(items),
            currency_count=len(effective_config.currencies),
            config_snapshot=snapshot,
        )
    )

    # 10. Update the pricing_state to reflect the new outputs.
    state["excel_file_id"] = str(excel_file.id)
    state["netsuite_csv_file_id"] = str(csv_file.id)
    state["row_count"] = len(items)
    state["header_columns"] = ["SKU", "Item Name", "USD", *sorted(state["effective_currencies"])]

    return {
        "success": True,
        "sku_count": len(items),
        "currency_count": len(effective_config.currencies),
        "output_files": output_files,
        "preview": _build_preview(results),
        "pricing_state": state,
        "template_mode": False,
    }


# Suppress unused-import warning — CurrencyConfig is referenced indirectly via
# TenantPricingConfig serialization.
_ = CurrencyConfig


# ---------------------------------------------------------------------------
# pricing_to_sheets — export the latest pricing run to a new Google Sheet
# ---------------------------------------------------------------------------


_SHEETS_CONNECTOR_MISSING = (
    "Google Sheets connector not configured. Set it up in Settings → Connectors first."
)
_PRICING_NOT_RUN_FOR_SHEETS = (
    "No pricing run in this conversation yet. Run pricing_convert (with an upload) "
    "or pricing_export (with inline items) first, then export to Sheets."
)


def _parse_excel_to_sheets_data(content: bytes) -> list[list]:
    """Re-parse a saved pricing Excel into [headers, *rows] for sheets.write_range.

    The default 3-sheet workbook has the row-data on the active sheet (Prices).
    Empty trailing cells are preserved as None to keep the column count stable.
    """
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows: list[list] = []
    for row in ws.iter_rows(values_only=True):
        # Trim purely empty trailing rows.
        if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row):
            continue
        rows.append(list(row))
    return rows


async def pricing_to_sheets_execute(params: dict, context: dict, **kwargs) -> dict:
    """Export the most recent pricing result to a new Google Sheet.

    Read-only consumer — calls get_latest_result_by_type("pricing"), re-parses
    the cached Excel file via TaskFileService, and writes the full row set to
    a new spreadsheet. The LLM never sees the rows. Connector check FIRST so
    the user gets the right setup-vs-session error.
    """
    if not context or not context.get("db") or not context.get("tenant_id"):
        return {"error": True, "message": "Missing context — tenant_id and db are required."}
    db = context["db"]
    tenant_id = context["tenant_id"]
    conversation_id = context.get("conversation_id")
    if not conversation_id:
        return {"error": True, "message": "Missing conversation_id — pricing_to_sheets needs a session-scoped cache."}

    # 1. Connector first.
    connector = await _get_sheets_connector(context)
    if not connector:
        return {"error": True, "message": _SHEETS_CONNECTOR_MISSING}

    # 2. Cache lookup.
    cached = await get_latest_result_by_type(str(conversation_id), "pricing")
    if not cached or not cached.payload:
        return {"error": True, "message": _PRICING_NOT_RUN_FOR_SHEETS}
    payload = cached.payload
    excel_file_id = payload.get("excel_file_id")
    if not excel_file_id:
        return {"error": True, "message": "Cached pricing payload is missing excel_file_id."}

    # 3. Re-parse the saved Excel into a 2D array.
    try:
        _task_file, content = await _file_svc.get_file(db, tenant_id, uuid.UUID(excel_file_id))
    except (ValueError, FileNotFoundError) as e:
        return {
            "error": True,
            "message": (
                "The Excel file from the prior pricing run is no longer available "
                f"({e}). Re-run pricing_convert / pricing_export and try again."
            ),
        }
    data_rows = _parse_excel_to_sheets_data(content)
    if not data_rows:
        return {"error": True, "message": "Excel file is empty — nothing to export."}

    # 4. Create + write + (optionally) share.
    credentials_envelope = decrypt_credentials(connector.encrypted_credentials)
    credentials = credentials_envelope.get("service_account_json", credentials_envelope)
    shared_drive_id = (connector.metadata_json or {}).get("shared_drive_id")

    title = params.get("title") or f"Pricing Export — {datetime.now(timezone.utc).strftime('%Y-%m-%d')}"

    create_result = await create_spreadsheet(
        credentials=credentials,
        title=title,
        shared_drive_id=shared_drive_id,
    )
    spreadsheet_id = create_result["spreadsheet_id"]
    url = create_result["url"]

    await write_range(
        credentials=credentials,
        spreadsheet_id=spreadsheet_id,
        data=data_rows,
        range_str="Sheet1!A1",
    )

    if not shared_drive_id:
        user_email = await _get_user_email(context)
        if user_email:
            try:
                await share_spreadsheet(
                    credentials=credentials,
                    spreadsheet_id=spreadsheet_id,
                    email=user_email,
                )
            except Exception:
                # Don't fail the export if share fails — user still has the URL.
                pass

    return {
        "success": True,
        "spreadsheet_id": spreadsheet_id,
        "url": url,
        "title": title,
        "sku_count": payload.get("row_count", max(0, len(data_rows) - 1)),
    }
