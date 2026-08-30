"""Read an uploaded task file's full contents, as structured rows.

THE GAP THIS CLOSES. Chat has accepted .xlsx/.csv/.xls/.json attachments for a
while, but the agent only ever saw a PREVIEW — the first 20 rows x 12 columns
(``_preview_xlsx_attachment`` in orchestrator.py). Beyond that it was guessing,
with nothing in the payload to say anything had been cut.

And the preview's own ``.xls`` branch already instructs the model to "use
file-aware tools with the file_id below" — a tool that did not exist. Only
``pricing.convert`` read a task file, and only for pricing. Directing a model at
an affordance that isn't there is the same defect class as the
``ns_selector_app`` dead end (see selector_app_redirect.py); this makes the
sentence true rather than deleting it.

WHAT IT DELIBERATELY DOES NOT DO. It does not interpret, summarise, or compute.
It hands back cells. Whatever reads those values is subject to the usual rule
that the model must not be the source of numbers that reach a record — for a
single-record write the human checks every field on the confirmation card, and
before any BATCH write lands, extraction has to become deterministic
server-side. That decision is recorded in the multimodal design; this tool is
the read half and is safe under either.
"""

from __future__ import annotations

import csv
import io
import json
import uuid
from typing import Any

from app.services.task_file_service import TaskFileService

__all__ = ["read_execute"]

_file_svc = TaskFileService()

# Hard ceiling on rows returned per call. Enforced HERE rather than described in
# the tool's prompt text: a cap the model is asked to respect is a request, a
# cap applied at the choke point is a guarantee (.claude/rules/agent-graph.md
# #4). Sized so a full page still fits comfortably in context alongside the
# rest of a turn; the model pages with `offset` when it needs more.
_MAX_ROWS = 200
_DEFAULT_ROWS = 100

# Column ceiling mirrors the row cap for the same reason — one pathological
# spreadsheet should not be able to blow the turn's context on its own.
_MAX_COLS = 50

# Legacy binary .xls is deliberately NOT here. openpyxl reads .xlsx (a zip)
# and raises BadZipFile on .xls; xlrd is not a dependency. Upload DOES accept
# .xls, so a user can reach this — they get a named refusal with a remedy
# rather than a parse error. Claiming a format we cannot read would be the
# same asserted-affordance defect this tool exists to remove.
_SUPPORTED = {"csv", "xlsx", "json"}
_LEGACY_XLS = "xls"


def _coerce_int(value: Any, default: int) -> int:
    """Models send numbers as strings often enough that refusing them is just
    a worse tool. Anything unreadable falls back to the default rather than
    raising — a bad page size must not fail a read."""
    try:
        return int(float(str(value)))
    except (TypeError, ValueError):
        return default


def _cell(value: Any) -> str:
    return "" if value is None else str(value)


def _rows_from_csv(content: bytes) -> tuple[list[str], list[list[str]]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    all_rows = [r for r in reader if any((c or "").strip() for c in r)]
    if not all_rows:
        return [], []
    return [_cell(c) for c in all_rows[0][:_MAX_COLS]], [[_cell(c) for c in r[:_MAX_COLS]] for r in all_rows[1:]]


def _rows_from_xlsx(content: bytes, sheet: str | None) -> tuple[list[str], list[list[str]], list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet_names = list(wb.sheetnames)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    rows = [
        [_cell(c) for c in row[:_MAX_COLS]]
        for row in ws.iter_rows(values_only=True)
        if any(c is not None and str(c).strip() for c in row)
    ]
    if not rows:
        return [], [], sheet_names
    return rows[0], rows[1:], sheet_names


def _rows_from_json(content: bytes) -> tuple[list[str], list[list[str]]]:
    """A JSON array of flat objects reads as a table; anything else is returned
    as a single value column rather than reshaped into a lie."""
    parsed = json.loads(content.decode("utf-8", errors="replace") or "[]")
    if isinstance(parsed, dict):
        parsed = [parsed]
    if not isinstance(parsed, list) or not parsed:
        return [], []
    if all(isinstance(item, dict) for item in parsed):
        columns: list[str] = []
        for item in parsed:
            for key in item:
                if key not in columns:
                    columns.append(str(key))
        columns = columns[:_MAX_COLS]
        return columns, [[_cell(item.get(c)) for c in columns] for item in parsed]
    return ["value"], [[_cell(item)] for item in parsed]


async def read_execute(params: dict, context: dict, **kwargs) -> dict:
    """Return a bounded page of an uploaded file's rows.

    Signature is fixed by the dispatcher: ``governed_execute`` calls every tool
    as ``execute_fn(validated_params, context=context)`` (governance.py:518),
    so ``db`` arrives INSIDE context — it is not a positional argument. The
    first version of this function took ``db`` positionally; every real
    invocation therefore died with "missing 1 required positional argument",
    while 17 unit tests passed because they called it the way I had imagined
    rather than the way the dispatcher does. Match pricing_convert_execute,
    which is the working precedent.

    Always reports ``total_rows`` and ``has_more`` so a partial read cannot be
    mistaken for a complete one — a silent cap reads as "I saw everything" when
    it didn't.
    """
    context = context or {}
    tenant_id = context.get("tenant_id")
    db = context.get("db")
    if db is None or not tenant_id:
        return {"error": True, "message": "Missing context — tenant_id and db are required."}
    raw_file_id = params.get("file_id")
    if not raw_file_id:
        return {"error": True, "message": "file_id is required"}

    try:
        file_uuid = uuid.UUID(str(raw_file_id))
    except (ValueError, AttributeError):
        return {"error": True, "message": f"Not a valid file_id: {raw_file_id}"}

    try:
        # Tenant scoping lives here and is not re-implemented — get_file raises
        # for a file this tenant does not own.
        task_file, content = await _file_svc.get_file(db, uuid.UUID(str(tenant_id)), file_uuid)
    except ValueError as exc:
        return {"error": True, "message": f"File not available: {exc}"}

    file_type = (getattr(task_file, "file_type", "") or "").lower()
    if file_type == _LEGACY_XLS:
        return {
            "error": True,
            "message": (
                "This is a legacy .xls workbook, which cannot be read here — only the newer "
                ".xlsx format is supported. Ask the user to re-save it as .xlsx (Excel: "
                "File -> Save As -> Excel Workbook) and attach it again."
            ),
        }
    if file_type not in _SUPPORTED:
        # Named explicitly rather than attempted. A PDF or an image read as text
        # produces plausible garbage, and plausible garbage on a path that feeds
        # record creation is worse than a refusal.
        return {
            "error": True,
            "message": (
                f"Cannot read '{file_type or 'unknown'}' files — this tool reads "
                f"{', '.join(sorted(_SUPPORTED))}. The file is attached but its contents "
                "are not machine-readable here."
            ),
        }

    sheets: list[str] = []
    try:
        if file_type == "xlsx":
            columns, data_rows, sheets = _rows_from_xlsx(content, params.get("sheet"))
        elif file_type == "json":
            columns, data_rows = _rows_from_json(content)
        else:
            columns, data_rows = _rows_from_csv(content)
    except Exception as exc:
        return {"error": True, "message": f"Could not parse the file ({type(exc).__name__}: {exc})"}

    offset = max(0, _coerce_int(params.get("offset"), 0))
    limit = _coerce_int(params.get("limit"), _DEFAULT_ROWS)
    limit = max(1, min(limit, _MAX_ROWS))

    total = len(data_rows)
    page = data_rows[offset : offset + limit]
    end = offset + len(page)

    result: dict[str, Any] = {
        "filename": getattr(task_file, "filename", None),
        "file_type": file_type,
        "columns": columns,
        "rows": page,
        "total_rows": total,
        "offset": offset,
        "returned": len(page),
        "has_more": end < total,
    }
    if end < total:
        result["next_offset"] = end
        result["note"] = (
            f"Showing rows {offset + 1}-{end} of {total}. Call again with offset={end} for more. "
            "Do NOT describe rows you have not read."
        )
    if sheets:
        result["sheets"] = sheets
    return result
