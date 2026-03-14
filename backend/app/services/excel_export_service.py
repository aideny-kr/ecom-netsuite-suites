"""Professional Excel export with financial-grade formatting."""

import io
import re
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


class ExcelExportConfig:
    """Branding & style configuration per tenant."""
    brand_name: str = "SuiteStudio"
    brand_color_hex: str = "1a73e8"
    header_bg_hex: str = "1a73e8"
    header_font_color: str = "FFFFFF"
    stripe_bg_hex: str = "f8f9fa"
    font_name: str = "Arial"
    font_size: int = 10
    header_font_size: int = 10
    title_font_size: int = 14


def generate_excel(
    columns: list[str],
    rows: list[list[Any]],
    *,
    title: str = "Query Results",
    sheet_name: str = "Data",
    config: ExcelExportConfig | None = None,
    metadata: dict[str, str] | None = None,
    column_types: dict[str, str] | None = None,
) -> io.BytesIO:
    """Generate a professionally formatted Excel workbook.

    Args:
        columns: Column headers
        rows: Data rows (list of lists, aligned to columns)
        title: Report title shown in the title row
        sheet_name: Excel sheet tab name
        config: Branding/style config (defaults to SuiteStudio theme)
        metadata: Key-value pairs for the info block
        column_types: Override auto-detected column types

    Returns:
        BytesIO buffer containing the .xlsx file
    """
    cfg = config or ExcelExportConfig()
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.sheet_properties.tabColor = cfg.brand_color_hex

    # --- Styles ---
    thin_border = Border(
        bottom=Side(style="thin", color="E0E0E0"),
    )
    header_fill = PatternFill("solid", fgColor=cfg.header_bg_hex)
    header_font = Font(
        name=cfg.font_name, size=cfg.header_font_size,
        bold=True, color=cfg.header_font_color,
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    stripe_fill = PatternFill("solid", fgColor=cfg.stripe_bg_hex)
    body_font = Font(name=cfg.font_name, size=cfg.font_size)
    body_alignment = Alignment(vertical="center")
    number_alignment = Alignment(horizontal="right", vertical="center")

    title_font = Font(
        name=cfg.font_name, size=cfg.title_font_size,
        bold=True, color=cfg.brand_color_hex,
    )
    meta_font = Font(name=cfg.font_name, size=9, color="666666")

    # --- Title block ---
    current_row = 1
    ws.merge_cells(
        start_row=current_row, start_column=1,
        end_row=current_row, end_column=min(len(columns), 6) if columns else 1,
    )
    title_cell = ws.cell(row=current_row, column=1, value=title)
    title_cell.font = title_font
    title_cell.alignment = Alignment(vertical="center")
    current_row += 1

    # --- Metadata block ---
    if metadata:
        for key, value in metadata.items():
            ws.cell(row=current_row, column=1, value=f"{key}:").font = Font(
                name=cfg.font_name, size=9, bold=True, color="333333",
            )
            ws.cell(row=current_row, column=2, value=value).font = meta_font
            current_row += 1

    # Timestamp
    ws.cell(row=current_row, column=1, value="Generated:").font = Font(
        name=cfg.font_name, size=9, bold=True, color="333333",
    )
    ws.cell(
        row=current_row, column=2,
        value=datetime.now(timezone.utc).strftime("%b %d, %Y %I:%M %p UTC"),
    ).font = meta_font
    current_row += 1

    # Spacer row
    current_row += 1

    # --- Column headers ---
    header_row = current_row
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=_humanize_header(col_name))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = Border(
            bottom=Side(style="medium", color=cfg.brand_color_hex),
        )
    current_row += 1

    # --- Auto-detect column types ---
    detected_types = _detect_column_types(columns, rows)
    if column_types:
        detected_types.update(column_types)

    # --- Data rows ---
    for row_idx, row_data in enumerate(rows):
        excel_row = current_row + row_idx
        is_striped = row_idx % 2 == 1

        for col_idx, value in enumerate(row_data):
            cell = ws.cell(row=excel_row, column=col_idx + 1)
            col_type = detected_types.get(columns[col_idx], "text")

            # Write value with type-appropriate formatting
            _write_typed_cell(cell, value, col_type)

            # Style
            cell.font = body_font
            cell.border = thin_border
            if col_type in ("currency", "number", "percent"):
                cell.alignment = number_alignment
            else:
                cell.alignment = body_alignment

            if is_striped:
                cell.fill = stripe_fill

    # --- Auto-fit column widths ---
    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        # Sample header + first 50 data rows for width
        max_len = len(_humanize_header(col_name))
        for row_data in rows[:50]:
            if col_idx - 1 < len(row_data):
                cell_len = len(str(row_data[col_idx - 1] or ""))
                max_len = max(max_len, cell_len)
        # Clamp between 8 and 45 characters
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 8), 45)

    # --- Freeze panes (header row) ---
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # --- Print setup ---
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = f"{header_row}:{header_row}"

    # --- Footer ---
    footer_row = current_row + len(rows) + 1
    ws.cell(row=footer_row, column=1, value=f"{len(rows)} rows").font = Font(
        name=cfg.font_name, size=9, italic=True, color="999999",
    )

    # Write to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _humanize_header(col: str) -> str:
    """Convert snake_case or camelCase to Title Case."""
    # camelCase -> spaced
    col = re.sub(r"([a-z])([A-Z])", r"\1 \2", col)
    # snake_case -> spaced
    col = col.replace("_", " ")
    return col.strip().title()


def _detect_column_types(
    columns: list[str], rows: list[list[Any]]
) -> dict[str, str]:
    """Auto-detect column types by sampling data.

    Returns dict of column_name -> type ("currency", "number", "percent", "date", "text")
    """
    CURRENCY_PATTERNS = {"amount", "balance", "total", "price", "cost", "revenue", "income", "expense", "debit", "credit", "net", "gross", "payment", "refund"}
    PERCENT_PATTERNS = {"rate", "margin", "percent", "pct", "ratio"}
    DATE_PATTERNS = {"date", "created", "modified", "updated", "posted", "period"}
    ID_PATTERNS = {"id", "internalid", "tranid", "acctnumber", "acct", "account", "number", "num", "code", "sku", "ref", "zip", "postal", "phone", "fax"}

    types: dict[str, str] = {}

    for col_idx, col_name in enumerate(columns):
        # Strip underscores, spaces, and non-alphanumeric chars (e.g., "Acct #" → "acct")
        lower = re.sub(r'[^a-z0-9]', '', col_name.lower())

        # Name-based heuristics first
        if any(p in lower for p in ID_PATTERNS):
            types[col_name] = "text"  # IDs should never be formatted as numbers
            continue
        if any(p in lower for p in CURRENCY_PATTERNS):
            types[col_name] = "currency"
            continue
        if any(p in lower for p in PERCENT_PATTERNS):
            types[col_name] = "percent"
            continue
        if any(p in lower for p in DATE_PATTERNS):
            types[col_name] = "date"
            continue

        # Data-based detection: sample first 20 non-null values
        sample_values = []
        for row in rows[:20]:
            if col_idx < len(row) and row[col_idx] is not None:
                sample_values.append(row[col_idx])

        if not sample_values:
            types[col_name] = "text"
            continue

        # Check if numeric
        numeric_count = 0
        for v in sample_values:
            if isinstance(v, (int, float)):
                numeric_count += 1
            elif isinstance(v, str):
                try:
                    float(v.replace(",", "").replace("$", "").replace("(", "-").replace(")", ""))
                    numeric_count += 1
                except ValueError:
                    pass

        if numeric_count / len(sample_values) >= 0.7:
            types[col_name] = "number"
        else:
            types[col_name] = "text"

    return types


def _write_typed_cell(cell, value: Any, col_type: str) -> None:
    """Write a value to a cell with appropriate type and number format."""
    if value is None:
        cell.value = ""
        return

    if col_type == "currency":
        num = _to_number(value)
        if num is not None:
            cell.value = num
            cell.number_format = '#,##0.00;(#,##0.00);"-"'
        else:
            cell.value = str(value)

    elif col_type == "number":
        num = _to_number(value)
        if num is not None:
            cell.value = num
            if isinstance(num, float) and not num.is_integer():
                cell.number_format = "#,##0.00"
            else:
                cell.number_format = "#,##0"
        else:
            cell.value = str(value)

    elif col_type == "percent":
        num = _to_number(value)
        if num is not None:
            # If value > 1, assume it's already a percentage (e.g. 45.2 = 45.2%)
            if abs(num) > 1:
                cell.value = num / 100
            else:
                cell.value = num
            cell.number_format = "0.0%"
        else:
            cell.value = str(value)

    elif col_type == "date":
        cell.value = str(value)  # Keep as-is; NetSuite dates are already formatted

    else:
        cell.value = str(value) if value is not None else ""


def _to_number(value: Any) -> float | int | None:
    """Coerce a value to a number, handling common formats."""
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        cleaned = value.strip().replace(",", "").replace("$", "")
        # Handle accounting parens: (123.45) -> -123.45
        if cleaned.startswith("(") and cleaned.endswith(")"):
            cleaned = "-" + cleaned[1:-1]
        try:
            num = float(cleaned)
            return int(num) if num.is_integer() else num
        except ValueError:
            return None
    return None
