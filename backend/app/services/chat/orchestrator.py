"""Agentic chat orchestrator using a multi-provider LLM adapter layer.

Supports Anthropic, OpenAI, and Gemini via the adapter pattern.
Claude decides which tools to call, sees results (including errors),
and can retry/correct — all within a single turn, up to MAX_STEPS iterations.
"""

import asyncio
import json
import logging
import re
import time
import uuid
from datetime import datetime, timezone
from typing import Any, AsyncGenerator

from sqlalchemy import func
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.services.chat.prompt_cache import split_system_prompt
from app.services.chat.tool_categories import categorize
from app.services.drive_rag.retriever import retrieve_drive_chunks

# Regex to strip leaked Anthropic tool-call XML from assistant text
_TOOL_XML_RE = re.compile(r"</?(?:invoke|parameter|tool_use)[^>]*>", re.DOTALL)
_TOOL_TAG_RE = re.compile(r"\s*\[tool:\s*[^\]]+\]")
_REASONING_RE = re.compile(r"<reasoning>.*?</reasoning>\s*", re.DOTALL)

# Chitchat regex — matches greetings, compliments, affirmations, farewells.
# Short-circuits expensive context assembly (entity resolution, domain knowledge, etc.)
_FINANCIAL_MODE_TAG = "FINANCIAL REPORT MODE"

# Hard cap on optional pre-flight LLM calls (e.g. entity resolution). These
# feed *optional* context — vernacular enrichment — and must never block the
# main chat turn. A single stalled Haiku socket used to burn the full 300s
# chat budget; this cap lets us proceed without vernacular if the call stalls.
_RESOLVE_ENTITIES_TIMEOUT_SECONDS = 15

# Wall-clock cap on per-turn Drive RAG retrieval (embed query + cosine search).
# Embedding API typically completes in 200-500ms. The OpenAI SDK has its own
# 60s read timeout (configured in app/services/chat/embeddings.py — see
# CLAUDE.md Mistake #51), but this 15s outer cap is the user-facing budget:
# a stalled provider must not hang the chat turn. On timeout the gather-result
# branch treats it as failed retrieval (existing Exception path) and the turn
# proceeds with empty drive_knowledge / drive_sources.
_GATHER_DRIVE_TIMEOUT_SECONDS = 15.0


def _build_financial_mode_task(user_message: str) -> str:
    """Build task for financial report queries.

    Always uses the LOCAL netsuite_financial_report tool which uses BUILTIN.CONSOLIDATE
    for correct multi-currency consolidation at posting-time exchange rates.
    MCP ns_runReport uses real-time FX rates which diverge from NetSuite UI numbers
    on multi-currency tenants — not suitable for penny-perfect financial statements.
    """
    return (
        f"{user_message}\n\n"
        f"[{_FINANCIAL_MODE_TAG}] Use the LOCAL netsuite_financial_report tool.\n\n"
        "Parameters:\n"
        '  report_type: "income_statement" | "balance_sheet" | "trial_balance" | "income_statement_trend" | "balance_sheet_trend"\n'
        '  period: "Feb 2026" (single) or "Jan 2026, Feb 2026, Mar 2026" (trend/quarter)\n'
        "  subsidiary_id: number (optional, use <tenant_vernacular> to resolve names to IDs)\n\n"
        "For quarters: use income_statement_trend with all 3 months.\n\n"
        "The financial report tool renders data directly in a visual table component.\n"
        "Do NOT format tables yourself. Do NOT rebuild or reproduce the data as markdown tables.\n"
        "Instead, provide COMMENTARY ONLY:\n"
        "1. A brief summary of the key findings\n"
        "2. Notable trends or anomalies\n"
        "3. Comparisons if the user asked for them\n"
        "Reference the pre-computed summary numbers (total_revenue, gross_profit, net_income, etc.) for your analysis.\n"
        "For trend reports, summary.by_period contains per-period breakdowns — compare across periods.\n\n"
        "FALLBACK: MCP ns_runReport if local tool errors (note: MCP may show slight FX differences)."
    )


_CHITCHAT_FILLER = r"(?:\s+(?:bro|man|dude|mate|buddy|guys?|y'?all|there))?"
_CHITCHAT_SEP = r"[!.\s,]*"
_CHITCHAT_RE = re.compile(
    rf"""(?xi)^[\s]*(?:
        (?:thanks?|thank\s*you|thx|ty|cheers)\b{_CHITCHAT_FILLER}{_CHITCHAT_SEP} |
        (?:great|nice|awesome|perfect|amazing|cool|excellent|wonderful|fantastic|brilliant)\b{_CHITCHAT_FILLER}{_CHITCHAT_SEP} |
        (?:good\s*job|well\s*done|nice\s*work|nailed\s*it)\b{_CHITCHAT_FILLER}{_CHITCHAT_SEP} |
        (?:you(?:'re|\s+are)\s+(?:the\s+)?(?:best|great|awesome|perfect|amazing))\b{_CHITCHAT_FILLER}{_CHITCHAT_SEP} |
        (?:(?:you\s+)?rock|love\s+(?:it|this|you)|bravo)\b{_CHITCHAT_FILLER}{_CHITCHAT_SEP} |
        (?:wow|lol|(?:ha){{2,}}|ok(?:ay)?|sure|yep|nope|got\s*it|understood|i\s*see|makes?\s*sense)\b{_CHITCHAT_FILLER}{_CHITCHAT_SEP} |
        (?:hi|hello|hey|good\s*(?:morning|afternoon|evening|night))\b{_CHITCHAT_FILLER}{_CHITCHAT_SEP} |
        (?:bye|goodbye|see\s*ya|later|gn)\b{_CHITCHAT_FILLER}{_CHITCHAT_SEP}
    ){{1,5}}[\s!.]*$""",
)


# ---------------------------------------------------------------------------
# Simple lookup detection — route to Haiku for 10x faster, 10x cheaper
# ---------------------------------------------------------------------------

HAIKU_MODEL = "claude-haiku-4-5-20251001"

# Patterns that indicate a simple single-entity lookup or count
_SIMPLE_LOOKUP_RE = re.compile(
    r"(?i)^(?:"
    r"(?:show|find|get|look\s*up|pull|what(?:'s|\s+is))\s+(?:me\s+)?(?:the\s+)?"
    r"(?:(?:status|details?)\s+(?:of|for)\s+)?"
    r"(?:(?:item|order|invoice|customer|vendor|po|so|rma|inv|vb)\s+)?"
    r"[\w\-#][\w\-#\s]{2,30}"  # ID/SKU/number or name (up to 30 chars)
    r"|(?:how\s+many|count|total\s+number\s+of)\s+(?:open\s+|active\s+|pending\s+)?"
    r"(?:sales\s+orders?|purchase\s+orders?|invoices?|rmas?|pos?|sos?|items?|customers?|vendors?)"
    r"|(?:SO|PO|RMA|INV|VB|WO|TO|IF|IR)[\s\-#]?\d+"
    r"|#\d+"
    r")\s*\??$",
)

# Patterns that indicate complex analysis — should NOT use Haiku
_COMPLEX_QUERY_RE = re.compile(
    r"(?i)(?:compar|vs\.?|versus|trend|month.over.month|year.over|breakdown|analyz|analysis"
    r"|all\s+\w+\s+with\s+their|why\s+|how\s+do\s+i|explain|create|write|fix|deploy"
    r"|refactor|p\s*&\s*l|income\s+statement|balance\s+sheet)",
)


def _is_simple_lookup(query: str) -> bool:
    """Detect simple single-entity lookups or counts.

    Conservative: only matches clearly simple queries. Returns False
    when uncertain — better to use Sonnet than give a bad Haiku answer.
    """
    query = query.strip()
    if not query or len(query) > 120:
        return False
    if _COMPLEX_QUERY_RE.search(query):
        return False
    return bool(_SIMPLE_LOOKUP_RE.match(query))


# ---------------------------------------------------------------------------
# Connection health check — prevent wasted tool calls on dead connections
# ---------------------------------------------------------------------------

_BROKEN_STATUSES = frozenset({"needs_reauth", "error", "expired"})

# Local tools that require a healthy REST API connection
_REST_TOOLS = frozenset({"netsuite_suiteql", "netsuite_financial_report"})


# ---------------------------------------------------------------------------
# R1: Explicit web search override
# ---------------------------------------------------------------------------
_WEB_SEARCH_RE = re.compile(
    r"""(?ix)
    (?:search\s+(?:the\s+)?web)
    | (?:web\s+search)
    | (?:look\s+(?:this\s+)?up\s+online)
    | (?:search\s+online)
    | (?:google\s+this)
    | (?:use\s+web\s*search)
    """,
)


def _detect_web_search_intent(user_message: str) -> bool:
    """Detect explicit web search requests (R1)."""
    return bool(_WEB_SEARCH_RE.search(user_message))


# ---------------------------------------------------------------------------
# R5: Data reference detection (reduce re-query bias)
# ---------------------------------------------------------------------------
_DATA_REFERENCE_RE = re.compile(
    r"""(?ix)
    (?:(?:the|that|those|same)\s+(?:list|data|result|customers|orders|items|records))
    | (?:(?:from|use|with|based\s+on)\s+(?:the\s+)?(?:previous|earlier|above|before|last))
    | (?:we\s+(?:just\s+)?(?:pulled|looked\s+up|queried|got|fetched))
    | (?:(?:use|take)\s+(?:the\s+)?(?:same|those|that)\s+(?:data|result|list))
    """,
)


def _detect_data_reference(user_message: str) -> bool:
    """Detect when user references previously-returned data (R5)."""
    return bool(_DATA_REFERENCE_RE.search(user_message))


# ---------------------------------------------------------------------------
# R6: NetSuite entity routing
# ---------------------------------------------------------------------------
_NETSUITE_ENTITY_RE = re.compile(
    r"""(?ix)
    (?:customer\s+balance)
    | (?:outstanding\s+balance)
    | (?:accounts?\s+receivable)
    | (?:AR\s+aging)
    | (?:open\s+(?:invoices?|sales\s+orders?|purchase\s+orders?|vendor\s+bills?))
    | (?:(?:purchase|sales)\s+order)
    | (?:journal\s+entr)
    | (?:credit\s+memo)
    | (?:vendor\s+bill)
    | (?:GL\s+balance)
    | (?:general\s+ledger)
    | (?:inventory\s+(?:count|adjustment|transfer))
    | (?:item\s+fulfillment)
    | (?:item\s+receipt)
    """,
)


def _detect_netsuite_entity(user_message: str) -> bool:
    """Detect queries about NetSuite-native entities (R6)."""
    return bool(_NETSUITE_ENTITY_RE.search(user_message))


async def _check_connection_health(db: AsyncSession, tenant_id: uuid.UUID) -> list[str]:
    """Check REST API and MCP connection health for a tenant.

    Returns a list of warning strings for broken connections.
    Fail-open: returns [] on any DB error so chat is never blocked.
    """
    try:
        from sqlalchemy import select as sa_select

        from app.models.connection import Connection
        from app.models.mcp_connector import McpConnector

        warnings: list[str] = []

        # Check REST API connections
        conn_result = await db.execute(
            sa_select(Connection.label, Connection.status, Connection.error_reason)
            .where(Connection.tenant_id == tenant_id)
            .where(Connection.provider == "netsuite")
        )
        for conn in conn_result.all():
            if conn.status in _BROKEN_STATUSES:
                warnings.append(f"REST API ({conn.label}): {conn.status}")

        # Check MCP connections
        mcp_result = await db.execute(
            sa_select(McpConnector.label, McpConnector.status)
            .where(McpConnector.tenant_id == tenant_id)
            .where(McpConnector.is_enabled == True)  # noqa: E712
        )
        for mcp in mcp_result.all():
            if mcp.status in _BROKEN_STATUSES:
                warnings.append(f"MCP ({mcp.label}): {mcp.status}")

        return warnings
    except Exception:
        # Fail-open — don't block chat if health check fails
        return []


def _filter_tools_for_dead_connections(tool_definitions: list[dict], connection_warnings: list[str]) -> list[dict]:
    """Remove tools whose backing connection is broken.

    - REST dead → strip local netsuite_suiteql, netsuite_financial_report
    - MCP dead → strip all ext__ prefixed tools
    """
    if not connection_warnings:
        return tool_definitions

    rest_dead = any("REST API" in w for w in connection_warnings)
    mcp_dead = any("MCP" in w for w in connection_warnings)

    filtered = []
    for t in tool_definitions:
        name = t["name"]
        if rest_dead and name in _REST_TOOLS:
            continue
        if mcp_dead and name.startswith("ext__"):
            continue
        filtered.append(t)
    return filtered


def _build_connection_warning_block(connection_warnings: list[str]) -> str:
    """Build a system prompt block warning the agent about broken connections."""
    if not connection_warnings:
        return ""

    warning_lines = "\n".join(f"  - {w}" for w in connection_warnings)
    return (
        f"\n\n⚠️ CONNECTION STATUS — BROKEN:\n{warning_lines}\n"
        "IMMEDIATELY tell the user which connections are down. "
        "Do NOT attempt queries against broken connections — they will fail. "
        "Direct user to Settings > Connections to re-authorize."
    )


from app.services.chat.tool_inventory import build_mcp_execution_guidance, build_tool_inventory_block


def _assemble_system_prompt(*, template: str, tool_definitions: list[dict]) -> str:
    """Resolve the {{TOOL_INVENTORY}} placeholder with the real tool schema.

    The replacement bundles:
    - the <available_tools> block (build_tool_inventory_block)
    - per-tool MCP guidance + EXECUTION PRIORITY (build_mcp_execution_guidance)
      when external MCP tools are present.

    Both are derived from the same tool_definitions, so the LLM's view of
    what it can call AND how to choose between them stays in sync with the
    real schema.
    """
    inventory = build_tool_inventory_block(tool_definitions)
    guidance = build_mcp_execution_guidance(tool_definitions)
    combined = inventory + guidance
    if "{{TOOL_INVENTORY}}" in template:
        return template.replace("{{TOOL_INVENTORY}}", combined)
    # DB-stored custom templates may lack the placeholder — append at end
    return template + f"\n\n{combined}" if combined else template


# ---------------------------------------------------------------------------
# Knowledge profiles — replace routing with context injection
# ---------------------------------------------------------------------------

from app.services.chat.knowledge_profiles import load_all_profiles

_knowledge_profiles = load_all_profiles()


# ---------------------------------------------------------------------------
# Smart context injection — classify how much context each query needs
# ---------------------------------------------------------------------------


class ContextNeed:
    """How much dynamic context to inject into the agent prompt."""

    FULL = "full"  # Custom fields, complex joins — inject everything
    DATA = "data"  # Standard tables, no custom fields — skip onboarding profile
    DOCS = "docs"  # Documentation question — skip schemas, inject RAG only
    WORKSPACE = "workspace"  # Script question — skip all NetSuite schemas
    FINANCIAL = "financial"  # Financial report — inject only vernacular + onboarding


def _compute_need_patterns(context_need: str, tool_names: set[str]) -> bool:
    """Decide whether to retrieve seeded SuiteQL/BigQuery patterns this turn.

    Pre-2026-04-16 this was `context_need in (ContextNeed.DATA,)` — patterns
    only fired when the classifier said DATA. That stranded admin-seeded
    patterns whenever a query classified as FULL (investigation), which is
    exactly when worked examples are most useful. New rule: patterns retrieve
    whenever a SQL query tool is connected, regardless of context_need.

    The `context_need` parameter is intentionally unused today — kept in the
    signature so call sites read self-documentingly and so future context-aware
    tuning (e.g., suppressing patterns for WORKSPACE queries even when SuiteQL
    tools are present) does not require touching call sites.

    The ext__ match is intentionally broader than `_is_suiteql_tool_call` in
    `query_pattern_service.py` — fetching patterns when a metadata tool
    (`ns_getSuiteQLMetadata`) is present is harmless, while not fetching them
    when a query tool is present is the bug we fixed. Similarly, `bigquery_sql`
    triggers the gate even though `tenant_query_patterns` doesn't yet store
    BigQuery patterns; `retrieve_similar_patterns` returns `[]` cheaply for
    BQ-only sessions, and the gate is ready when Phase 2 adds BigQuery storage.
    """
    if {"netsuite_suiteql", "bigquery_sql"} & tool_names:
        return True
    return any(name.startswith("ext__") and "suiteql" in name.lower() for name in tool_names)


_WORKSPACE_RE = re.compile(
    r"\b(?:scripts?|deploy(?:ment)?s?|triggers?|automation|scheduled|user\s*events?|"
    r"suitelets?|restlets?|map\s*reduce|client\s*scripts?|mass\s*updates?|portlets?|"
    r"bundles?|sdf|customscript\w*|fix\s+(?:the|my|this)\s+\w*script|write\s+a?\s*(?:suite)?script|"
    r"workflow\s*action\s*scripts?)\b",
    re.IGNORECASE,
)

_FINANCIAL_RE = re.compile(
    r"\b(?:p\s*&?\s*l|profit\s*(?:and|&)\s*loss|income\s*statement|balance\s*sheet|"
    r"trial\s*balance|aging|financial\s*(?:report|statement))\b",
    re.IGNORECASE,
)

_DOCS_RE = re.compile(
    r"\b(?:how\s+(?:do|does|can|should|to)|what\s+is|what\s+are|explain|documentation|"
    r"workflow|error\s+(?:message|code)|tutorial)\b",
    re.IGNORECASE,
)

_DATA_KEYWORDS = re.compile(
    r"(?:\b(?:show|list|find|get|pull|fetch|look\s*up|search|query|"
    r"how\s+many|total|count|sum|average|revenue|sales|orders?|"
    r"inventory|customers?|vendors?|invoices?|purchase\s*orders?)\b"
    r"|(?:^|\s)(?:RMA|PO|SO|INV|VB|WO|TO|IF|IR|#)\d)",
    re.IGNORECASE,
)

# Investigation queries — "why" questions that need FULL context (data + workspace tools)
_INVESTIGATION_RE = re.compile(
    r"why\s+(?:is|was|isn't|wasn't|did|didn't|does|doesn't|aren't|weren't)|"
    r"how\s+(?:does|did|is)\s+\w+\s+(?:set|triggered|controlled|determined|calculated|routed|sent)|"
    r"what\s+(?:controls|triggers|sets|determines|calculates|routes)|"
    r"root\s+cause|investigate|dig\s+into|"
    r"(?:give\s+me|show\s+me|get)\s+.*?\bhistory\b(?!\s+(?:for\s+|by\s+))|"
    r"(?:brief|full|complete)\s+history|"
    r"\bhistory\b.*(?:RMA|PO|SO|INV|WO|TO|IF|IR|#|[A-Z]\d{4,})|(?:RMA|PO|SO|INV|WO|TO|IF|IR|#|[A-Z]\d{4,}).*\bhistory\b|"
    r"timeline|audit.?trail|what.?happened|how.?long|when.?was",
    re.IGNORECASE,
)


def _classify_context_need(user_message: str, is_financial: bool = False) -> str:
    """Classify how much dynamic context a query needs.

    Returns one of ContextNeed.FULL/DATA/DOCS/WORKSPACE/FINANCIAL.
    When uncertain, returns FULL — never risk under-injecting.
    """
    msg = user_message.strip()

    # Investigation: "why" questions need FULL context (data tools + workspace tools)
    # Must check BEFORE financial — a "why" follow-up in a financial session still needs investigation
    if _INVESTIGATION_RE.search(msg):
        return ContextNeed.FULL

    if is_financial:
        return ContextNeed.FINANCIAL

    # Workspace: script/deploy/SuiteScript keywords
    if _WORKSPACE_RE.search(msg):
        # But if it also has data keywords, it might be mixed — use FULL
        if _DATA_KEYWORDS.search(msg):
            return ContextNeed.FULL
        return ContextNeed.WORKSPACE

    # Docs: how-to/explain/documentation without data keywords
    if _DOCS_RE.search(msg):
        if _DATA_KEYWORDS.search(msg):
            return ContextNeed.FULL  # Mixed intent — safe fallback
        return ContextNeed.DOCS

    # Data: most common path — orders, inventory, revenue, etc.
    if _DATA_KEYWORDS.search(msg):
        return ContextNeed.DATA

    # Uncertain — always default to FULL
    return ContextNeed.FULL


def _sanitize_assistant_text(text: str) -> str:
    """Remove leaked tool-call XML, tool tags, and reasoning blocks from assistant response text."""
    text = _TOOL_XML_RE.sub("", text)
    text = _TOOL_TAG_RE.sub("", text)
    text = _REASONING_RE.sub("", text)
    return text.strip()


_NO_RESULT_FALLBACK = "I wasn't able to find relevant information for that question."
# Used when the agent called at least one tool but still produced no final
# text. The original "find relevant information" wording reads like the
# agent has no knowledge of the topic — even when prior turns answered the
# question — and consistently registers as memory loss to users. This
# variant acknowledges the tool spiral and points at the prior answer.
_TOOL_SPIRAL_FALLBACK = (
    "My tool searches didn't surface a new answer for that. If an earlier "
    "turn already covered it, refer back; otherwise try rephrasing."
)
_PRICING_TASK_OUTPUT_MESSAGE = "Pricing output is ready. Review the table and download files below."


def _is_pricing_task_output(persisted_output: dict | None) -> bool:
    if not isinstance(persisted_output, dict) or persisted_output.get("type") != "task_output":
        return False
    data = persisted_output.get("data") or {}
    return isinstance(data, dict) and data.get("task_kind") == "pricing"


def _coerce_assistant_content(
    final_text: str | None,
    persisted_output: dict | None,
    *,
    tool_calls: list[dict] | None = None,
) -> str:
    """Decide what string to persist as ``ChatMessage.content``.

    When ``persisted_output`` carries ``type == "clarification"``, the
    clarification card IS the message — any text rendered above it (including
    the empty-result fallback) confuses the user. So in that case we return
    an empty string and let the frontend render the card alone.

    Otherwise, return ``final_text`` if non-empty. Empty text falls back to
    one of two messages depending on whether tools were called:

    - No tools called → the agent had nothing to say and didn't try; the
      original "find relevant information" wording fits.
    - Tools called → the agent went into a tool spiral and produced no
      final text. Use the spiral-specific wording so the user doesn't read
      this as "I don't remember anything we just discussed".
    """
    if isinstance(persisted_output, dict) and persisted_output.get("type") == "clarification":
        return ""
    if _is_pricing_task_output(persisted_output):
        return _PRICING_TASK_OUTPUT_MESSAGE
    if final_text:
        return final_text
    if tool_calls:
        return _TOOL_SPIRAL_FALLBACK
    return _NO_RESULT_FALLBACK


def _build_workspace_context_block(
    *,
    workspace_name: str,
    workspace_id: str,
    file_paths: list[str],
) -> str:
    """Build the workspace context block appended to the system prompt.

    Cap the file listing at 50 entries with a tail summary so very large
    workspaces don't dominate the prompt.

    The block also nudges the agent to prefer prior conversation history
    on follow-up questions rather than re-running workspace searches.
    Without this nudge, follow-up questions like "what does it convert
    to?" trigger fresh searches for terms the prior answer already
    explained — and when those searches turn up empty, the empty-text
    fallback fires and reads to the user as memory loss.
    """
    file_listing = "\n".join(f"- {_sanitize_for_prompt(p)}" for p in file_paths[:50])
    if len(file_paths) > 50:
        file_listing += f"\n... and {len(file_paths) - 50} more files"

    # workspace_name comes from tenant-admin user input (POST /api/v1/workspaces)
    # so strip control chars + cap length the same way file paths are handled.
    # workspace_id is a UUID set by the server, no sanitization needed.
    safe_name = _sanitize_for_prompt(workspace_name)
    return (
        f"\n\nWORKSPACE CONTEXT:\n"
        f"Active workspace: '{safe_name}' (ID: {workspace_id}).\n"
        f"Files in workspace:\n{file_listing}\n\n"
        f"Use workspace tools (workspace_list_files, workspace_read_file, "
        f"workspace_search, workspace_propose_patch) to browse and modify files. "
        f"The workspace_id is '{workspace_id}' — it will be auto-injected."
        "\nWhen the user mentions they are 'viewing' or 'looking at' a specific file, "
        "or when the message includes '[Currently viewing file: ...]', "
        "use the workspace_read_file tool to read that file's content before responding. "
        "This lets you see exactly what the user sees in their editor."
        "\n\n## IDE Chat Behavior\n"
        "You are an IDE assistant with direct file access. "
        "Be concise — lead with the answer, use code blocks, no preambles.\n"
        "For complex reasoning, use <thinking>...</thinking> tags before your answer. "
        "This block is collapsed by default in the UI.\n"
        "\n## Follow-up questions\n"
        "If the user asks a follow-up that the conversation history already "
        "covers (e.g. asking 'what does it convert to?' after you just "
        "explained a conversion), answer from the prior answer instead of "
        "re-running searches. Only reach for tools when you genuinely need "
        "fresh info — a detail you didn't already cover, or a file you "
        "haven't read yet. Re-searching for terms you already cited "
        "wastes a turn and often returns empty.\n"
    )


def _is_financial_tool(tool_name: str) -> bool:
    """True for local financial-report tools and external MCP ns_runReport."""
    return categorize(tool_name) == "financial"


def _is_data_table_tool(tool_name: str) -> bool:
    """True for tools returning tabular data (SuiteQL, BigQuery, pivot)."""
    return categorize(tool_name) in ("data_table", "bigquery")


async def _gather_drive_knowledge(*, db, tenant_id, query_text: str) -> dict:
    """Retrieve Drive chunks and build a name→url source map.

    Returns {"chunks": [...], "sources": {source_name: web_view_link}}.
    When multiple chunks share a source_name, the first URL wins.
    """
    chunks = await retrieve_drive_chunks(db=db, tenant_id=tenant_id, query_text=query_text)
    sources: dict[str, str] = {}
    for c in chunks:
        name = c.get("source_name")
        link = c.get("web_view_link") or ""
        if name and name not in sources:
            sources[name] = link
    return {"chunks": chunks, "sources": sources}


def _build_drive_knowledge_block(chunks: list[dict]) -> str:
    """Format Drive chunks as a <drive_knowledge> XML block for prompt injection."""
    if not chunks:
        return ""
    lines = ["<drive_knowledge>"]
    for c in chunks:
        src = c.get("source_name", "")
        url = c.get("web_view_link", "")
        lines.append(f'  <chunk source="{src}" url="{url}">')
        content = (c.get("content") or "").replace("\n", "\n    ")
        lines.append(f"    {content}")
        lines.append("  </chunk>")
    lines.append("</drive_knowledge>")
    return "\n".join(lines)


# User-inserted Drive mentions: `[Name](https://docs.google.com/...)` or
# `[Name](https://drive.google.com/...)`. The frontend's `#` mention picker
# emits this exact markdown-link form, but the regex also catches any Drive
# URL a user pastes by hand in the same format.
_DRIVE_MENTION_RE = re.compile(r"\[([^\]\n]+)\]\((https://(?:docs|drive)\.google\.com/[^\s)]+)\)")


def _extract_drive_mentions(user_message: str) -> dict[str, str]:
    """Parse markdown links to Drive / Docs URLs out of *user_message*.

    Returns ``{name: url}`` for each `[name](drive_url)` match. Non-Drive URLs
    are ignored. Broken markdown (missing paren, space between `]` and `(`,
    unclosed bracket) produces no match — callers never need to catch errors.
    """
    if not user_message:
        return {}
    return {m.group(1): m.group(2) for m in _DRIVE_MENTION_RE.finditer(user_message)}


def _merge_drive_mentions(context: dict, user_message: str) -> list[tuple[str, str]]:
    """Hoist user-inserted Drive mentions into ``context["drive_sources"]``.

    Retrieval-based entries are authoritative — if ``drive_sources`` already
    contains an entry for the mentioned name, it's not overwritten. Returns the
    list of newly-added ``(name, url)`` pairs so callers can build prompt hints
    that list exactly what the user pointed at this turn.
    """
    mentions = _extract_drive_mentions(user_message)
    if not mentions:
        return []
    existing = context.setdefault("drive_sources", {})
    new_entries: list[tuple[str, str]] = []
    for name, url in mentions.items():
        if name not in existing:
            existing[name] = url
            new_entries.append((name, url))
    return new_entries


def _build_drive_mentions_hint(mentions: list[tuple[str, str]]) -> str:
    """Render user-mentioned Drive files as a system-prompt hint fragment.

    Emits a leading ``\\n\\n`` so the hint can be concatenated directly to an
    existing prompt without worrying about preceding content.
    """
    if not mentions:
        return ""
    lines = ["\n\n## User-mentioned Drive files this turn:"]
    for name, url in mentions:
        lines.append(f"- [{name}]({url})")
    return "\n".join(lines)


def _is_bigquery_tool(tool_name: str) -> bool:
    """True for BigQuery query tools."""
    return categorize(tool_name) == "bigquery"


# Pricing tools that WRITE the typed pricing cache entry (CachedResult.payload).
# pricing_to_sheets is intentionally absent — it's a read-only consumer that
# reads the latest pricing entry via get_latest_result_by_type.
_PRICING_WRITE_TOOLS = frozenset(
    {
        "pricing_convert",
        "pricing.convert",
        "pricing_export",
        "pricing.export",
        "pricing_revise",
        "pricing.revise",
    }
)


_SAVED_SEARCH_TOOLS = frozenset({"netsuite.saved_search", "netsuite_saved_search"})


def _is_saved_search_tool(tool_name: str) -> bool:
    """Match local saved search tools and external MCP saved search tools."""
    if tool_name in _SAVED_SEARCH_TOOLS:
        return True
    if "savedsearch" in tool_name.lower() or "runsavedsearch" in tool_name.lower():
        return True
    return False


def _compute_source_pin_update(tool_calls_log: list[dict]) -> str | None:
    """Decide whether a turn's tool calls should update session.source_pin.

    Accepts log entries with either "tool_name" or "tool" key (build_tool_call_log_entry
    uses "tool"; test fixtures may use "tool_name" — both are supported).

    Returns:
        "bigquery"  — pin to BigQuery (only BigQuery data tools fired)
        "netsuite"  — pin to NetSuite (only NetSuite data tools fired)
        None        — clear the pin (mixed data sources in this turn)
        "leave_pin" — leave pin unchanged (no data tools fired)
    """
    used_bq = False
    used_ns = False
    for call in tool_calls_log:
        # Support both "tool_name" (test fixtures) and "tool" (build_tool_call_log_entry)
        name = call.get("tool_name") or call.get("tool", "")
        cat = categorize(name)
        if cat == "bigquery":
            used_bq = True
        elif cat in {"data_table", "financial"}:
            # data_table covers netsuite_suiteql + pivot; financial covers report.
            # bigquery is its own category so we only land here for NetSuite.
            used_ns = True

    if used_bq and used_ns:
        return None  # mixed — clear
    if used_bq:
        return "bigquery"
    if used_ns:
        return "netsuite"
    return "leave_pin"


def _intercept_tool_result(
    tool_name: str, result_str: str, context_need: str = ContextNeed.DATA
) -> tuple[str | None, dict | None, str]:
    """Intercept data-producing tool results for frontend DataFrame rendering.

    Returns ``(event_type, sse_event_data, result_str_for_llm)``.
    - Financial reports → ``("financial_report", {...}, condensed)``
    - SuiteQL queries  → ``("data_table", {...}, condensed)``
    - Everything else  → ``(None, None, original_result_str)``

    When context_need == FULL (investigation queries), the LLM receives full
    row data so it can reason over system notes, field changes, etc.
    """

    # --- Financial report path ---
    if _is_financial_tool(tool_name):
        try:
            parsed = json.loads(result_str)
        except (json.JSONDecodeError, TypeError):
            return None, None, result_str
        if not parsed.get("success"):
            return None, None, result_str

        rows = parsed.get("items", [])
        sse_event_data = {
            "report_type": parsed.get("report_type"),
            "period": parsed.get("period"),
            "columns": parsed.get("columns", []),
            "rows": rows,
            "summary": parsed.get("summary"),
        }
        condensed = json.dumps(
            {
                "success": True,
                "report_type": parsed.get("report_type"),
                "period": parsed.get("period"),
                "total_rows": len(rows),
                "summary": parsed.get("summary"),
                "note": (
                    "The full table has been sent to the frontend for rendering. "
                    "Do NOT rebuild or reproduce the table in your response. "
                    "Provide commentary and analysis only."
                ),
            },
            default=str,
        )
        return "financial_report", sse_event_data, condensed

    # --- SuiteQL query path ---
    if _is_data_table_tool(tool_name):
        try:
            parsed = json.loads(result_str)
        except (json.JSONDecodeError, TypeError):
            return None, None, result_str

        # Error results pass through
        if isinstance(parsed, dict) and (parsed.get("error") is True or isinstance(parsed.get("error"), str)):
            return None, None, result_str

        columns = parsed.get("columns")
        rows = parsed.get("rows")

        # Handle external MCP format: {"data"|"items": [{col: val, ...}, ...]}
        if not isinstance(columns, list) or not isinstance(rows, list):
            # External MCP uses "data" or "items" key for list-of-dicts
            items = None
            if isinstance(parsed, dict):
                items = parsed.get("data") or parsed.get("items")
            if isinstance(items, list) and len(items) > 0 and isinstance(items[0], dict):
                # Derive columns from union of all item keys (preserving order)
                seen: set[str] = set()
                columns = []
                for item in items:
                    for key in item:
                        if key not in seen:
                            seen.add(key)
                            columns.append(key)
                rows = [[item.get(col) for col in columns] for item in items]
            else:
                return None, None, result_str

        row_count = parsed.get("row_count") or parsed.get("resultCount") or len(rows)
        query = parsed.get("query") or parsed.get("queryExecuted") or ""
        truncated = parsed.get("truncated", False)

        sse_event_data = {
            "columns": columns,
            "rows": rows,
            "row_count": row_count,
            "query": query,
            "truncated": truncated,
        }
        # Investigation queries (FULL context): send all rows so LLM can reason
        # over system notes, field changes, timelines, etc.
        # Standard queries: 5-row preview to save tokens.
        if context_need == ContextNeed.FULL:
            condensed = json.dumps(
                {
                    "columns": columns,
                    "row_count": row_count,
                    "rows": rows,
                    "truncated": truncated,
                    "note": (
                        "The data table is also rendered in the frontend. "
                        "Do NOT reproduce the table. Analyze the data and explain findings."
                    ),
                },
                default=str,
            )
        else:
            row_preview = rows[:30]
            condensed = json.dumps(
                {
                    "columns": columns,
                    "row_count": row_count,
                    "rows_preview": row_preview,
                    "truncated": truncated,
                    "note": (
                        "The full data table has been sent to the frontend for rendering. "
                        "Do NOT rebuild or reproduce the table in your response. "
                        "Provide commentary, insights, and analysis only. "
                        "Use rows_preview for charting and follow-up analysis."
                    ),
                },
                default=str,
            )
        return "data_table", sse_event_data, condensed

    # --- Saved search path ---
    if _is_saved_search_tool(tool_name):
        try:
            parsed = json.loads(result_str)
        except (json.JSONDecodeError, TypeError):
            return None, None, result_str

        # Error results pass through
        if isinstance(parsed, dict) and (parsed.get("error") is True or isinstance(parsed.get("error"), str)):
            return None, None, result_str

        # Extract list-of-dicts from data, items, or results keys
        items = None
        if isinstance(parsed, dict):
            items = parsed.get("data") or parsed.get("items") or parsed.get("results")
        if not isinstance(items, list) or len(items) == 0 or not isinstance(items[0], dict):
            return None, None, result_str

        # Derive columns from union of all item keys (preserving order)
        seen: set[str] = set()
        columns = []
        for item in items:
            for key in item:
                if key not in seen:
                    seen.add(key)
                    columns.append(key)
        rows = [[item.get(col) for col in columns] for item in items]

        search_id = parsed.get("searchId", "")
        query = f"Saved Search: {search_id}" if search_id else "Saved Search"
        row_count = parsed.get("resultCount") or len(rows)
        truncated = parsed.get("truncated", False)

        sse_event_data = {
            "columns": columns,
            "rows": rows,
            "row_count": row_count,
            "query": query,
            "truncated": truncated,
        }

        row_preview = rows[:5]
        condensed = json.dumps(
            {
                "columns": columns,
                "row_count": row_count,
                "rows_preview": row_preview,
                "truncated": truncated,
                "note": (
                    "The full saved search results have been sent to the frontend for rendering. "
                    "Do NOT rebuild or reproduce the table in your response. "
                    "Provide commentary, insights, and analysis only."
                ),
            },
            default=str,
        )
        return "data_table", sse_event_data, condensed

    # --- Task output path (pricing conversion + revise) ---
    if tool_name in (
        "pricing_convert",
        "pricing.convert",
        "pricing_export",
        "pricing.export",
        "pricing_revise",
        "pricing.revise",
    ):
        try:
            parsed = json.loads(result_str)
        except (json.JSONDecodeError, TypeError):
            return None, None, result_str
        if not parsed.get("success"):
            return None, None, result_str

        sse_event_data = {
            "type": "task_output",
            "task_kind": "pricing",
            "sku_count": parsed.get("sku_count", 0),
            "currency_count": parsed.get("currency_count", 0),
            "output_files": parsed.get("output_files", {}),
            "preview": parsed.get("preview", []),
            "template_mode": parsed.get("template_mode", False),
            # pricing_state is consumed by the cache callback (orchestrator
            # writes it as the typed CachedResult.payload). The frontend
            # ignores unknown keys.
            "pricing_state": parsed.get("pricing_state"),
        }
        condensed = json.dumps(
            {
                "success": True,
                "sku_count": parsed.get("sku_count"),
                "currency_count": parsed.get("currency_count"),
                "output_files": parsed.get("output_files"),
                "template_mode": parsed.get("template_mode"),
                "note": (
                    "Conversion complete. The pricing table is displayed to the user automatically. "
                    "Do NOT list individual prices — the user can see them in the table and download below. "
                    "Just confirm: how many SKUs were converted and how many currencies."
                ),
            },
            default=str,
        )
        return "task_output", sse_event_data, condensed

    # --- Sheets link path (sheets_create + pricing_to_sheets) ---
    if tool_name in (
        "sheets_create",
        "sheets.create",
        "pricing_to_sheets",
        "pricing.to_sheets",
    ):
        try:
            parsed = json.loads(result_str)
        except (json.JSONDecodeError, TypeError):
            return None, None, result_str
        if not isinstance(parsed, dict) or parsed.get("error") is True:
            return None, None, result_str
        url = parsed.get("url")
        if not url:
            return None, None, result_str
        sse_event_data = {
            "url": url,
            "spreadsheet_id": parsed.get("spreadsheet_id", ""),
            "title": parsed.get("title", "Spreadsheet"),
            "shared_with": parsed.get("shared_with"),
        }
        condensed = json.dumps(
            {
                "success": True,
                "spreadsheet_id": parsed.get("spreadsheet_id", ""),
                "title": parsed.get("title", ""),
                "note": (
                    "The Sheet link is shown to the user as a clickable card. "
                    "Confirm what was exported and any follow-ups — do NOT paste the URL in your reply."
                ),
            },
            default=str,
        )
        return "sheets_link", sse_event_data, condensed

    # --- Docs link path ---
    if tool_name in ("docs_create", "docs.create"):
        try:
            parsed = json.loads(result_str)
        except (json.JSONDecodeError, TypeError):
            return None, None, result_str
        if not isinstance(parsed, dict) or parsed.get("error") is True:
            return None, None, result_str
        url = parsed.get("url")
        if not url:
            return None, None, result_str
        sse_event_data = {
            "url": url,
            "doc_id": parsed.get("doc_id", ""),
            "title": parsed.get("title", "Document"),
            "shared_with": parsed.get("shared_with"),
        }
        condensed = json.dumps(
            {
                "success": True,
                "doc_id": parsed.get("doc_id", ""),
                "title": parsed.get("title", ""),
                "note": (
                    "The Doc link is shown to the user as a clickable card. "
                    "Confirm what was saved in one short line — do NOT paste the URL in your reply."
                ),
            },
            default=str,
        )
        return "docs_link", sse_event_data, condensed

    # --- Not a data tool ---
    return None, None, result_str


_NON_DATA_EVENTS = frozenset({"sheets_link", "docs_link"})


def _build_intercept_cache_entry(
    *,
    tool_name: str,
    event_type_str: str,
    event_data: dict,
    conversation_id: str,
):
    """Build the CachedResult for an intercepted tool event, or return None
    when the event should not write to the cache (e.g., sheets_link).

    Lazy-imports CachedResult to avoid the orchestrator-level circular import
    chain. Used by both the unified-agent intercept callback and the legacy
    single-agent intercept site so pricing follow-ups (pricing_revise /
    pricing_to_sheets) work in either path.
    """
    if event_type_str in _NON_DATA_EVENTS:
        return None
    from app.services.chat.result_cache import CachedResult

    if tool_name in _PRICING_WRITE_TOOLS:
        result_type = "pricing"
        payload = event_data.get("pricing_state")
    else:
        result_type = (
            "financial_report"
            if event_type_str == "financial_report"
            else "bigquery"
            if _is_bigquery_tool(tool_name)
            else "suiteql"
        )
        payload = None
    synthetic_id = f"pending-{uuid.uuid4().hex[:12]}"
    return CachedResult(
        message_id=synthetic_id,
        conversation_id=conversation_id,
        result_type=result_type,
        columns=event_data.get("columns", []),
        rows=event_data.get("rows", []),
        row_count=event_data.get("row_count", 0),
        summary=event_data.get("summary"),
        query_text=event_data.get("query", ""),
        payload=payload,
    )


def _strip_cache_only_fields_from_sse(event_data: dict) -> dict:
    """Remove cache-only fields (e.g., pricing_state) before yielding to SSE.

    pricing_state can carry the full seed_items / effective_items list for a
    5K-SKU catalog (~150KB JSON). The frontend only renders the preview; the
    cache callback already has the full payload.
    """
    if "pricing_state" in event_data:
        return {k: v for k, v in event_data.items() if k != "pricing_state"}
    return event_data


def _intercept_with_cache(
    tool_name: str,
    result_str: str,
    *,
    context_need: str,
    session_id: str | None,
) -> tuple[str | None, dict | None, str]:
    """Intercept a tool result, write to the cache (when applicable), and
    return the SSE-safe event data with cache-only fields stripped. Both
    the legacy single-agent path and the unified-agent path use this so
    pricing_revise / pricing_to_sheets see the prior pricing_state in
    either flow."""
    from app.services.chat.result_cache import _cache_result_sync

    event_type, event_data, new_result_str = _intercept_tool_result(tool_name, result_str, context_need=context_need)
    if event_type is None or event_data is None:
        return event_type, event_data, new_result_str

    if session_id:
        cr = _build_intercept_cache_entry(
            tool_name=tool_name,
            event_type_str=event_type,
            event_data=event_data,
            conversation_id=session_id,
        )
        if cr is not None:
            _cache_result_sync(session_id, cr.message_id, cr)

    return event_type, _strip_cache_only_fields_from_sse(event_data), new_result_str


def _make_tool_interceptor(context_need: str = ContextNeed.DATA, cache_callback=None):
    """Create a tool interceptor closure that captures context_need.

    The unified-agent path passes a ``cache_callback`` so it can also
    accumulate ``_pending_caches`` for the assistant-message alias write
    after the agent loop. The shared cache write lives in
    ``_build_intercept_cache_entry`` so the legacy path stays in sync.
    """

    def interceptor(tool_name: str, result_str: str) -> tuple[tuple[str, dict] | None, str]:
        event_type, event_data, new_result_str = _intercept_tool_result(
            tool_name, result_str, context_need=context_need
        )
        if event_type is not None and event_data is not None:
            if cache_callback:
                cache_callback(tool_name, event_type, event_data)
            return (event_type, _strip_cache_only_fields_from_sse(event_data)), new_result_str
        return None, new_result_str

    return interceptor


def _get_confidence_explanation(score: float) -> str:
    """Return a human-readable explanation for a confidence score."""
    if score >= 4.5:
        return "Very high confidence — used proven patterns and all tools succeeded"
    if score >= 3.5:
        return "High confidence — data looks correct"
    if score >= 2.5:
        return "Moderate confidence — results may need verification"
    return "Low confidence — please verify this data before acting on it"


from app.models.chat import ChatMessage, ChatSession
from app.services.audit_service import log_event
from app.services.chat.billing import deduct_chat_credits
from app.services.chat.llm_adapter import get_adapter
from app.services.chat.nodes import (
    OrchestratorState,
    get_tenant_ai_config,
    retriever_node,
    sanitize_user_input,
)
from app.services.chat.onboarding_tools import (
    ONBOARDING_TOOL_DEFINITIONS,
    execute_onboarding_tool,
)
from app.services.chat.prompts import INPUT_SANITIZATION_PREFIX, ONBOARDING_SYSTEM_PROMPT
from app.services.chat.tool_call_results import build_tool_call_log_entry, tool_call_had_error, tool_call_row_count
from app.services.chat.tools import build_all_tool_definitions, execute_tool_call
from app.services.prompt_template_service import get_active_template

logger = logging.getLogger(__name__)

MAX_STEPS = settings.CHAT_MAX_TOOL_CALLS_PER_TURN  # default 5


def _extract_file_paths(nodes: list[dict]) -> list[str]:
    """Flatten a nested file tree into a sorted list of file paths."""
    paths: list[str] = []
    for node in nodes:
        if not node.get("is_directory"):
            paths.append(node.get("path", ""))
        children = node.get("children")
        if children:
            paths.extend(_extract_file_paths(children))
    return sorted(paths)


def _sanitize_for_prompt(text: str) -> str:
    """Strip control characters and limit length for prompt injection safety."""
    cleaned = re.sub(r"[\x00-\x1f\x7f]", "", text)
    return cleaned[:500]


def _truncate_attachment_preview(text: str, limit: int = 12000) -> str:
    if len(text) <= limit:
        return text
    return text[:limit] + f"\n...[truncated {len(text) - limit} chars]"


def _decode_text_attachment(content: bytes) -> str:
    return content.decode("utf-8-sig", errors="replace")


def _encode_attachment_preview_for_prompt(text: str) -> str:
    """Encode file text as JSON so uploaded content cannot close prompt tags."""
    encoded = json.dumps(text, ensure_ascii=False)
    return encoded.replace("<", "\\u003c").replace(">", "\\u003e").replace("&", "\\u0026")


def _preview_xlsx_attachment(content: bytes) -> str:
    import io

    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows: list[str] = []
    for row_index, row in enumerate(sheet.iter_rows(max_row=20, max_col=12, values_only=True), start=1):
        values = ["" if value is None else str(value) for value in row]
        rows.append(f"{row_index}: " + " | ".join(values))
    sheet_names = ", ".join(workbook.sheetnames)
    return f"Workbook sheets: {sheet_names}\nActive sheet preview:\n" + "\n".join(rows)


async def _build_attached_file_context(
    db: AsyncSession,
    tenant_id: uuid.UUID,
    attached_file_id: str | None,
) -> str:
    """Build a bounded prompt preview for an uploaded chat attachment."""
    if not attached_file_id:
        return ""

    from app.services.task_file_service import TaskFileService

    try:
        task_file, content = await TaskFileService().get_file(db, tenant_id, uuid.UUID(attached_file_id))
    except (ValueError, OSError):
        logger.warning("attached_file.load_failed", exc_info=True)
        return (
            "<attached_file>\n"
            f"file_id: {attached_file_id}\n"
            "error: The attached file could not be loaded.\n"
            "</attached_file>"
        )

    file_type = (task_file.file_type or "").lower()
    preview = ""
    try:
        if file_type in {"csv", "json"}:
            text = _decode_text_attachment(content)
            if file_type == "json":
                try:
                    preview = json.dumps(json.loads(text), indent=2, ensure_ascii=False)
                except json.JSONDecodeError:
                    preview = text
            else:
                preview = text
        elif file_type == "xlsx":
            preview = _preview_xlsx_attachment(content)
        else:
            preview = (
                "Binary Excel file attached. Use file-aware tools with the file_id below when the user's task "
                "requires processing the full workbook."
            )
    except Exception:
        logger.warning("attached_file.preview_failed", exc_info=True)
        preview = "Preview unavailable. Use file-aware tools with the file_id below."

    preview = _truncate_attachment_preview(preview)
    safe_preview = _encode_attachment_preview_for_prompt(preview)
    return (
        "<attached_file>\n"
        f"file_id: {attached_file_id}\n"
        f"filename: {_sanitize_for_prompt(task_file.filename)}\n"
        f"file_type: {file_type}\n"
        f"size_bytes: {task_file.file_size}\n"
        "Instructions: This is user-provided input for the current turn. "
        "If you call a file-aware tool such as pricing.convert, pass this exact file_id. "
        "For CSV, JSON, and XLSX questions, parse the JSON-encoded preview string below when it is sufficient.\n"
        "<preview>\n"
        f"{safe_preview}\n"
        "</preview>\n"
        "</attached_file>"
    )


def _is_valid_uuid(val: str) -> bool:
    """Check if a string is a valid UUID."""
    try:
        uuid.UUID(str(val))
        return True
    except (ValueError, AttributeError):
        return False


async def _resolve_default_workspace(
    db: AsyncSession,
    tenant_id: uuid.UUID,
) -> str | None:
    """Find the most recent active workspace for a tenant (cached per request)."""
    from sqlalchemy import select

    from app.models.workspace import Workspace

    result = await db.execute(
        select(Workspace.id)
        .where(Workspace.tenant_id == tenant_id, Workspace.status == "active")
        .order_by(Workspace.created_at.desc())
        .limit(1)
    )
    ws = result.scalar_one_or_none()
    return str(ws) if ws else None


async def _dispatch_memory_update(
    tenant_id: uuid.UUID,
    user_id: uuid.UUID,
    db: AsyncSession,
    user_message: str,
    assistant_message: str,
) -> None:
    """Helper to safely run the regex-gated memory updater in the background."""
    from app.services.chat.llm_adapter import get_adapter
    from app.services.chat.memory_updater import maybe_extract_correction

    provider = settings.MULTI_AGENT_SPECIALIST_PROVIDER
    api_key = settings.ANTHROPIC_API_KEY
    model = settings.MULTI_AGENT_SPECIALIST_MODEL

    try:
        adapter = get_adapter(provider, api_key)
        await maybe_extract_correction(
            db=db,
            tenant_id=tenant_id,
            user_id=user_id,
            user_message=user_message,
            assistant_message=assistant_message,
            adapter=adapter,
            model=model,
        )
    except Exception as e:
        logger.error(f"background.memory_update_failed: {e}")


async def _dispatch_content_summary(
    db: AsyncSession,
    message_id: uuid.UUID,
    user_message: str,
    assistant_message: str,
) -> None:
    """Fire-and-forget: generate and persist a factual summary for history."""
    from app.services.chat.summariser import dispatch_content_summary

    try:
        await dispatch_content_summary(
            db=db,
            message_id=message_id,
            user_message=user_message,
            assistant_message=assistant_message,
        )
    except Exception as e:
        logger.error(f"background.content_summary_failed: {e}")


async def run_chat_turn(
    db: AsyncSession,
    session: ChatSession,
    user_message: str,
    user_id: uuid.UUID,
    tenant_id: uuid.UUID,
    user_msg: ChatMessage | None = None,
    wizard_step: str | None = None,
    user_timezone: str | None = None,
    agent_id: str | None = None,
    run_id: str | None = None,
    write_confirm: dict | None = None,
    attached_file_id: str | None = None,
    plan_mode_choice: dict | None = None,
) -> AsyncGenerator[dict, None]:
    """Execute an agentic chat turn with Claude's native tool use.

    Signature and return type match the previous linear pipeline —
    chat.py needs zero changes.
    """
    correlation_id = str(uuid.uuid4())

    # Round 7 Bug 1 — set after `handle_plan_mode_choice` succeeds and the
    # CAS-flipped row's id is known. The wrapping try/except below uses this
    # to revert the row back to 'pending' if any downstream step in the
    # resumed turn raises. None on non-resume turns → the except block is a
    # no-op revert path (still re-raises so callers see the failure).
    _revert_message_id_on_failure: uuid.UUID | None = None

    # ── Write confirmation short-circuit (HITL) — runs before any expensive
    # context assembly (history, RAG, entity resolution). The approve/reject
    # path only needs a single DB lookup by message ID.
    if write_confirm and isinstance(write_confirm, dict):
        _wc_action = write_confirm.get("action")
        _wc_confirmation_id = write_confirm.get("confirmation_id")

        if _wc_action in ("approve", "reject") and _wc_confirmation_id:
            from sqlalchemy import select as _wc_select
            from sqlalchemy.orm.attributes import flag_modified as _wc_flag_modified

            from app.services.chat.write_confirmation_service import validate_and_extract_confirmation

            _confirm_result = await db.execute(
                _wc_select(ChatMessage).where(
                    ChatMessage.id == uuid.UUID(_wc_confirmation_id),
                    ChatMessage.session_id == session.id,
                )
            )
            _confirm_msg = _confirm_result.scalar_one_or_none()

            if _confirm_msg is None:
                yield {"type": "error", "error": "Confirmation message not found."}
                return

            _so = _confirm_msg.structured_output
            if not isinstance(_so, dict) or _so.get("type") != "write_confirmation" or _so.get("status") != "pending":
                yield {"type": "error", "error": "Confirmation is not in a pending state."}
                return

            if _wc_action == "approve":
                is_valid, tool_name, tool_input = validate_and_extract_confirmation(_so, str(session.id))
                if not is_valid:
                    yield {"type": "error", "error": "Confirmation token is invalid or tampered."}
                    return

                _exec_result_str = await execute_tool_call(
                    tool_name=tool_name,
                    tool_input=tool_input,
                    tenant_id=tenant_id,
                    actor_id=user_id,
                    correlation_id=correlation_id,
                    db=db,
                    session_id=str(session.id),
                )

                _mutation_type = _so.get("mutation_type", "write")
                _record_type = _so.get("record_type", "record")

                _exec_succeeded = False
                try:
                    _exec_result = json.loads(_exec_result_str)
                    if isinstance(_exec_result, dict) and _exec_result.get("error"):
                        _confirm_content = f"The operation failed: {_exec_result['error']}"
                    else:
                        _exec_succeeded = True
                        _confirm_content = f"Done — the {_record_type} {_mutation_type} has been executed successfully."
                except (json.JSONDecodeError, TypeError):
                    _exec_succeeded = True
                    _confirm_content = f"The {_mutation_type} operation has been executed."

                _updated_so = dict(_so)
                _updated_so["status"] = "approved" if _exec_succeeded else "pending"
                _confirm_msg.structured_output = _updated_so
                _wc_flag_modified(_confirm_msg, "structured_output")

                await log_event(
                    db=db,
                    tenant_id=tenant_id,
                    category="chat",
                    action=f"record.{_mutation_type}.{'approved' if _exec_succeeded else 'failed'}",
                    actor_id=user_id,
                    resource_type="chat_session",
                    resource_id=str(session.id),
                    payload={"tool_name": tool_name, "tool_input": tool_input, "result": _exec_result_str[:1000]},
                )

                _assistant_msg = ChatMessage(
                    tenant_id=tenant_id,
                    session_id=session.id,
                    role="assistant",
                    content=_confirm_content,
                    created_at=datetime.now(timezone.utc),
                )
                db.add(_assistant_msg)
                await db.commit()
                await db.refresh(_assistant_msg)

                yield {
                    "type": "message",
                    "message": {
                        "id": str(_assistant_msg.id),
                        "role": "assistant",
                        "content": _confirm_content,
                        "tool_calls": None,
                        "citations": None,
                        "created_at": _assistant_msg.created_at.isoformat(),
                    },
                }
                print(f"[WRITE-CONFIRM] approved {_mutation_type} on {_so.get('record_type')}", flush=True)
                return

            elif _wc_action == "reject":
                _updated_so = dict(_so)
                _updated_so["status"] = "rejected"
                _confirm_msg.structured_output = _updated_so
                _wc_flag_modified(_confirm_msg, "structured_output")

                _reject_content = "No changes were made. The proposed write operation was cancelled."
                _assistant_msg = ChatMessage(
                    tenant_id=tenant_id,
                    session_id=session.id,
                    role="assistant",
                    content=_reject_content,
                    created_at=datetime.now(timezone.utc),
                )
                db.add(_assistant_msg)
                await db.commit()

                yield {
                    "type": "message",
                    "message": {
                        "id": str(_assistant_msg.id),
                        "role": "assistant",
                        "content": _reject_content,
                        "tool_calls": None,
                        "citations": None,
                        "created_at": _assistant_msg.created_at.isoformat(),
                    },
                }
                print(f"[WRITE-CONFIRM] rejected {_so.get('mutation_type')} on {_so.get('record_type')}", flush=True)
                return

    # ── Plan Mode resume short-circuit (HITL clarify) ──
    # Two variants:
    #   - Source pick (option A/B/C): transition pending → chosen, fall
    #     through with chosen_source filter active.
    #   - Manual clarification (typed text): transition pending →
    #     manually_clarified, fall through with NO source filter and the
    #     typed text appended to the user's task.
    plan_mode_resume_directive: str | None = None
    plan_mode_resume_source: str | None = None
    plan_mode_manual_text: str | None = None
    # True when EITHER resume variant fired (source pick or manual clarify).
    # Both variants mean "the user just disambiguated this turn — do not
    # re-fire the clarify-gate / augmentation". Set after the short-circuit
    # below; consumed by the augmentation guard and the gate-arm guard.
    _plan_mode_resume_active = False
    if plan_mode_choice and isinstance(plan_mode_choice, dict):
        from app.services.chat.plan_mode.short_circuit import (
            PlanModeChoiceError,
            PlanModeManualResult,
            handle_plan_mode_choice,
        )

        _pmc_result = await handle_plan_mode_choice(
            plan_mode_choice=plan_mode_choice,
            session_id=str(session.id),
            tenant_id=tenant_id,
            db=db,
        )
        if isinstance(_pmc_result, PlanModeChoiceError):
            yield {
                "type": "error",
                "error": _pmc_result.error,
                "status_code": _pmc_result.status_code,
            }
            return

        plan_mode_resume_directive = _pmc_result.system_directive
        _plan_mode_resume_active = True
        if isinstance(_pmc_result, PlanModeManualResult):
            plan_mode_manual_text = _pmc_result.manual_text
            plan_mode_resume_source = None  # no source filter for manual
            logger.info(
                "[PLAN_MODE] resume turn: manual_clarify (len=%d)",
                len(plan_mode_manual_text),
            )
        else:
            plan_mode_resume_source = _pmc_result.chosen_source
            logger.info(
                "[PLAN_MODE] resume turn: chosen_source=%s",
                plan_mode_resume_source,
            )

        # CFO-grade audit trail (Task 6.4). FATAL by design — if this audit
        # write fails, we error the turn rather than silently approving an
        # untraceable choice. Mirrors the write_confirm audit pattern at
        # ``record.{create,update}.{approved,failed}`` above.
        #
        # codex round 6 Bug 1 — to keep the choice + audit effectively
        # atomic, on any audit failure we MUST revert the CAS that
        # ``handle_plan_mode_choice`` just committed (pending → chosen).
        # Without the revert, the user retries and gets HTTP 409 forever
        # because the row is stuck at ``status='chosen'`` with no
        # corresponding audit row — card consumed, no answer.
        try:
            # Action verb + payload differ by variant. Keeping the literal
            # action strings inline (instead of through an intermediate
            # variable) so static analysis / grep can pin them at the call
            # site (e.g., test_orchestrator_audit checks for "plan_mode.chose").
            if plan_mode_manual_text is not None:
                await log_event(
                    db=db,
                    tenant_id=tenant_id,
                    category="chat",
                    action="plan_mode.manual_clarify",
                    actor_id=user_id,
                    resource_type="chat_session",
                    resource_id=str(session.id),
                    payload={
                        "manual_text_chars": len(plan_mode_manual_text),
                        "confirmation_id": plan_mode_choice.get("confirmation_id"),
                    },
                )
            else:
                await log_event(
                    db=db,
                    tenant_id=tenant_id,
                    category="chat",
                    action="plan_mode.chose",
                    actor_id=user_id,
                    resource_type="chat_session",
                    resource_id=str(session.id),
                    payload={
                        "chosen_id": plan_mode_choice.get("option_id"),
                        "chosen_source": plan_mode_resume_source,
                        "confirmation_id": plan_mode_choice.get("confirmation_id"),
                    },
                )
        except Exception:
            from app.services.chat.plan_mode.short_circuit import (
                revert_clarification_to_pending,
            )

            _msg_id_to_revert = _pmc_result.chat_message_id
            # Round 7 Bug 2 — when log_event raises during its flush, the
            # SQLAlchemy session is left in a failed-transaction state. The
            # subsequent revert query would raise PendingRollbackError and
            # the revert never executes, leaving the row stranded at
            # 'chosen'. Roll back the failed transaction first so the
            # revert can run on a clean session.
            try:
                await db.rollback()
            except Exception:
                logger.exception(
                    "[PLAN_MODE] db.rollback() failed during audit-failure handling — clarification %s may be stranded",
                    _msg_id_to_revert,
                )
            if _msg_id_to_revert is not None:
                try:
                    await revert_clarification_to_pending(
                        message_id=_msg_id_to_revert,
                        tenant_id=tenant_id,
                        db=db,
                    )
                except Exception:
                    logger.exception(
                        "[PLAN_MODE] revert_clarification_to_pending failed "
                        "after audit emission failure — clarification %s "
                        "may be stranded at status='chosen'",
                        _msg_id_to_revert,
                    )
            logger.exception("[PLAN_MODE] plan_mode.chose audit emission failed; reverting CAS so user can retry")
            # Re-raise so the turn fails clearly rather than silently
            # proceeding without an audit row.
            raise

        # Audit succeeded — arm the revert-on-failure guard for the
        # remaining turn body (round 7 Bug 1). The wrapping try/except
        # below will revert the CAS if anything else raises.
        _revert_message_id_on_failure = _pmc_result.chat_message_id
    # NOTE: do NOT return — fall through into the regular flow.
    # plan_mode_resume_* variables are consumed by Task 4.3 (tool filter)
    # and the system-prompt assembly path (just append the directive).

    # ── Round 7 Bug 1: revert-on-resume-failure guard ─────────────
    # If `handle_plan_mode_choice` succeeded above, the structured_output
    # row was just CAS-flipped from 'pending' → 'chosen'. ANY failure
    # in the rest of this turn (LLM error, tool failure, persist commit
    # failure, anything) MUST revert that row back to 'pending' so the
    # user can retry. Without this revert, retries hit the
    # already-resolved CAS guard in handle_plan_mode_choice and the
    # endpoint returns 409 forever — card consumed, no answer.
    #
    # `_revert_message_id_on_failure` is set ONLY on the audit-success
    # branch above; it stays None for normal (non-resume) turns, in
    # which case the except block below is a no-op revert path.
    try:
        # ── Plan Mode: supersede any stale pending clarification ──
        # If this turn is NOT a plan_mode_choice resume, transition any pending
        # clarification on this session to 'superseded'. The user typed instead
        # of clicking the card — their typed message is the de-facto answer.
        if plan_mode_choice is None:
            try:
                from app.services.chat.plan_mode.short_circuit import (
                    supersede_pending_clarifications,
                )

                superseded_ids = await supersede_pending_clarifications(
                    session_id=session.id,
                    tenant_id=tenant_id,
                    db=db,
                )
                # Emit one ``chat.plan_mode.superseded`` audit per row (Task 6.4).
                # NON-FATAL on purpose — supersede is best-effort already, and the
                # ChatDisclosureEvent row is the primary telemetry.
                for _msg_id in superseded_ids:
                    await log_event(
                        db=db,
                        tenant_id=tenant_id,
                        category="chat",
                        action="plan_mode.superseded",
                        actor_id=user_id,
                        resource_type="chat_message",
                        resource_id=str(_msg_id),
                        payload={
                            "reason": "free_text_reply",
                            "session_id": str(session.id),
                        },
                    )
            except Exception:
                logger.warning(
                    "[PLAN_MODE] supersede_pending_clarifications failed",
                    exc_info=True,
                )
                # Non-fatal — telemetry only; chat continues normally.

        # ── Load conversation history (summary-based windowing) ──
        from app.services.chat.history_compactor import KEEP_RECENT
        from app.services.chat.history_tool_trace import build_history_dicts

        max_turns = settings.CHAT_MAX_HISTORY_TURNS
        all_messages: list[dict] = []
        summarised = 0
        if session.messages:
            # Convert ORM → dicts so the history builder can be unit-tested.
            # We include `tool_calls` so build_history_dicts can replay a compact
            # tool-call trace for the next turn — without this, the agent loses
            # the SQL/tool pattern that worked in the previous turn and
            # rediscovers it from scratch (see Olivia 2026-04-09 tangent).
            msg_dicts = [
                {
                    "role": m.role,
                    "content": m.content,
                    "content_summary": m.content_summary,
                    "tool_calls": m.tool_calls,
                    # Codex round 10 P2 Bug 1: surface clarification options
                    # into LLM-facing content. Without this, the resume
                    # directive ("Picked option B (source: netsuite)") refers
                    # to an empty prior assistant message and the agent has
                    # no way to know what option B's definition was.
                    "structured_output": m.structured_output,
                }
                for m in session.messages
                if m.role in ("user", "assistant")
            ]
            all_messages, summarised = build_history_dicts(msg_dicts, keep_recent=KEEP_RECENT)

        # Hard cap at max_turns * 2 messages
        history_messages = all_messages[-(max_turns * 2) :]

        # Condense large tool results in older messages to reduce token bloat
        from app.services.chat.history_compactor import build_condensed_history

        original_chars = sum(len(m.get("content", "")) for m in history_messages)
        history_messages = build_condensed_history(history_messages, keep_recent=4)
        condensed_chars = sum(len(m.get("content", "")) for m in history_messages)
        if original_chars != condensed_chars:
            print(f"[ORCHESTRATOR] history condensed: {original_chars} -> {condensed_chars} chars", flush=True)
        print(f"[ORCHESTRATOR] history loaded: {len(history_messages)} messages ({summarised} summarised)", flush=True)

        # ── Save user message (if not already saved by caller) ──
        if user_msg is None:
            user_msg = ChatMessage(
                tenant_id=tenant_id,
                session_id=session.id,
                role="user",
                content=user_message,
            )
            db.add(user_msg)
            await db.flush()

        is_onboarding = getattr(session, "session_type", "chat") == "onboarding"

        # ── Pre-loop: RAG retrieval for doc context (skip for onboarding) ──
        sanitized_input = sanitize_user_input(user_message)

        # Manual clarify variant: append the user's typed clarification to
        # the original query so the agent has both pieces in context. The
        # short_circuit handler already validated len <= 500 and stripped
        # whitespace; sanitize defensively here too.
        if plan_mode_manual_text:
            _manual_sanitized = sanitize_user_input(plan_mode_manual_text)
            if _manual_sanitized:
                sanitized_input = f"{sanitized_input}\n\nClarification: {_manual_sanitized}"
        attached_file_context = await _build_attached_file_context(db, tenant_id, attached_file_id)
        rag_context = ""
        citations: list[dict] = []

        if not is_onboarding:
            state = OrchestratorState(
                user_message=sanitized_input,
                tenant_id=tenant_id,
                actor_id=user_id,
                session_id=session.id,
                conversation_history=history_messages,
                route={"needs_docs": True},  # always attempt RAG
            )
            try:
                await retriever_node(state, db)
            except Exception:
                logger.warning("Pre-loop RAG retrieval failed, continuing without docs")
                state.doc_chunks = []

            # Build RAG context block
            if state.doc_chunks:
                rag_parts = []
                for chunk in state.doc_chunks:
                    rag_parts.append(f"[Documentation: {chunk['title']}]\n{chunk['content']}")
                    citations.append(
                        {
                            "type": "doc",
                            "title": chunk["title"],
                            "snippet": chunk["content"][:200],
                        }
                    )
                rag_context = "\n\n".join(rag_parts)

        # ── Build messages for Claude ──
        messages: list[dict] = list(history_messages)

        # Compose user message with sanitization prefix and RAG context
        user_content = f"{INPUT_SANITIZATION_PREFIX}\n\n"
        if rag_context:
            user_content += f"<context>\n{rag_context}\n</context>\n\n"
        if attached_file_context:
            user_content += f"{attached_file_context}\n\n"
        user_content += f"User question: {sanitized_input}"
        messages.append({"role": "user", "content": user_content})

        # ── Build tool definitions (with policy-based filtering) ──
        connection_warnings: list[str] = []
        # Initialize before branch (Mistake #47 — variables used after branches must
        # be initialized first; the chitchat path skips the else block below).
        plan_mode_enabled: bool = False
        if is_onboarding:
            tool_definitions = list(ONBOARDING_TOOL_DEFINITIONS)
        else:
            # Plan Mode: register clarify in the tool inventory when the flag is on.
            # The hard gate that ACTIVATES clarify (filters to clarify-only +
            # forces tool_choice) lives further down — this just makes the tool
            # available to the LLM.
            from app.services import feature_flag_service as _ffs_for_inventory

            plan_mode_enabled = await _ffs_for_inventory.is_enabled(db, tenant_id, "plan_mode_enabled")
            tool_definitions = await build_all_tool_definitions(db, tenant_id, plan_mode_enabled=plan_mode_enabled)

            # Pre-flight connection health check — strip tools for dead connections
            connection_warnings = await _check_connection_health(db, tenant_id)
            if connection_warnings:
                tool_definitions = _filter_tools_for_dead_connections(tool_definitions, connection_warnings)

        # ── Compute active knowledge profile partitions for RAG scoping ──
        _profile_partitions: list[str] = []
        if not is_onboarding:
            from app.services.chat.prompt_assembler import collect_rag_partitions, get_active_profiles

            _tool_names = {t["name"] for t in tool_definitions}
            _active_profiles = get_active_profiles(_knowledge_profiles, _tool_names)
            _profile_partitions = collect_rag_partitions(_active_profiles)

        # ── Resolve tenant-specific system prompt ──
        if is_onboarding:
            from app.services.chat.prompts import ONBOARDING_STEP_CONTEXTS

            system_prompt = ONBOARDING_SYSTEM_PROMPT
            if wizard_step and wizard_step in ONBOARDING_STEP_CONTEXTS:
                system_prompt = f"{ONBOARDING_SYSTEM_PROMPT}\n\n## Current Step: {wizard_step}\n{ONBOARDING_STEP_CONTEXTS[wizard_step]}"
        else:
            system_prompt = await get_active_template(db, tenant_id)

        # ── Inject brand identity so the AI knows its name ──
        brand_name = ""
        _tenant_config_row = None
        if not is_onboarding:
            from sqlalchemy import select as sa_select

            from app.models.tenant import Tenant, TenantConfig

            # Single query for all TenantConfig fields used in this function
            try:
                _tc_result = await db.execute(sa_select(TenantConfig).where(TenantConfig.tenant_id == tenant_id))
                _tenant_config_row = _tc_result.scalar_one_or_none()
            except Exception:
                _tenant_config_row = None
            brand_name = (getattr(_tenant_config_row, "brand_name", None) if _tenant_config_row else None) or ""
            if not brand_name:
                tenant_result = await db.execute(sa_select(Tenant.name).where(Tenant.id == tenant_id))
                brand_name = tenant_result.scalar_one_or_none() or "Suite Studio AI"
            system_prompt += (
                f'\n\nYour name is "{brand_name}". When asked to introduce yourself '
                f"or asked your name, say you are {brand_name}.\n"
            )

        # ── Inject AI Soul (Tone & Quirks) ──
        if not is_onboarding:
            from app.services.soul_service import get_soul_config

            soul_config = await get_soul_config(tenant_id)
            if soul_config.exists:
                soul_parts = ["\n\n## Tenant-Specific AI Configuration & Logic\n"]
                if soul_config.bot_tone:
                    soul_parts.append(f"TONE & MANNER:\n{soul_config.bot_tone}\n")
                if soul_config.netsuite_quirks:
                    soul_parts.append(f"NETSUITE QUIRKS & LOGIC:\n{soul_config.netsuite_quirks}\n")
                system_prompt += "\n".join(soul_parts)

        # ── Resolve {{TOOL_INVENTORY}} placeholder in the system prompt ──
        # using the real tool schema (single source of truth).
        if not is_onboarding:
            system_prompt = _assemble_system_prompt(
                template=system_prompt,
                tool_definitions=tool_definitions,
            )

        # ── Connection health warning (appended after tool inventory) ──
        if connection_warnings:
            system_prompt += _build_connection_warning_block(connection_warnings)

        # ── Workspace context injection ──
        workspace_context: dict | None = None
        if getattr(session, "workspace_id", None):
            from app.services import workspace_service as ws_svc

            ws = await ws_svc.get_workspace(db, session.workspace_id, tenant_id)
            if ws:
                workspace_context = {
                    "workspace_id": str(session.workspace_id),
                    "name": ws.name,
                }
                files = await ws_svc.list_files(db, session.workspace_id, tenant_id)
                file_paths = _extract_file_paths(files)
                system_prompt += _build_workspace_context_block(
                    workspace_name=ws.name,
                    workspace_id=str(session.workspace_id),
                    file_paths=file_paths,
                )

        # ── Load active policy for tool gating + output redaction ──
        from app.services.policy_service import evaluate_tool_call as policy_evaluate
        from app.services.policy_service import get_active_policy, redact_output

        active_policy = await get_active_policy(db, tenant_id)

        # ── Resolve tenant AI config ──
        provider, model, api_key, is_byok = await get_tenant_ai_config(db, tenant_id)
        adapter = get_adapter(provider, api_key)

        # ── Importance tier default (overridden in unified/legacy paths) ──
        from app.services.importance_classifier import ImportanceTier, classify_importance

        importance_tier = classify_importance(sanitized_input)

        # ── Multi-agent routing (opt-in per tenant or globally) ──
        if not is_onboarding and not workspace_context:
            from sqlalchemy import select as sa_select

            from app.models.tenant import TenantConfig

            use_multi_agent = settings.MULTI_AGENT_ENABLED
            if _tenant_config_row and isinstance(getattr(_tenant_config_row, "multi_agent_enabled", None), bool):
                use_multi_agent = _tenant_config_row.multi_agent_enabled

            if use_multi_agent:
                from app.services.chat.llm_adapter import get_adapter as get_specialist_adapter
                from app.services.netsuite_metadata_service import get_active_metadata

                if is_byok:
                    # BYOK: use tenant's own provider + API key
                    specialist_adapter = get_specialist_adapter(provider, api_key)
                else:
                    specialist_adapter = get_specialist_adapter(
                        settings.MULTI_AGENT_SPECIALIST_PROVIDER,
                        settings.ANTHROPIC_API_KEY,
                    )
                metadata = await get_active_metadata(db, tenant_id)

                # ── Check for unified agent flag (Phase 2) ──
                use_unified = settings.UNIFIED_AGENT_ENABLED
                if _tenant_config_row and isinstance(getattr(_tenant_config_row, "unified_agent_enabled", None), bool):
                    use_unified = _tenant_config_row.unified_agent_enabled

                if use_unified:
                    # ── Unified agent path: context assembly → single agent → stream ──
                    from app.services.chat.agents import UnifiedAgent

                    # ── Chitchat short-circuit: skip expensive context for conversational messages ──
                    _is_chitchat = bool(_CHITCHAT_RE.match(sanitized_input))
                    _selected_agent_id = None  # no routing fork; kept for audit trail
                    is_financial = False
                    is_web_search = False
                    is_netsuite_entity = False  # noqa: F841  — defensive init (see test_orchestrator_paths.py)
                    _has_data_reference = False

                    from app.services.importance_classifier import ImportanceTier, classify_importance

                    importance_tier = ImportanceTier.CASUAL  # default for chitchat/MCP

                    if _is_chitchat:
                        print("[UNIFIED] Chitchat detected — skipping context assembly", flush=True)
                        context_need = ContextNeed.DOCS  # chitchat needs no data context
                        context: dict[str, Any] = {"user_timezone": user_timezone}
                        unified_agent = UnifiedAgent(
                            tenant_id=tenant_id,
                            user_id=user_id,
                            correlation_id=correlation_id,
                            metadata=None,
                            policy=active_policy,
                        )
                    else:
                        # Detect financial intent for task augmentation + domain knowledge boost
                        is_financial = bool(_FINANCIAL_RE.search(sanitized_input))
                        is_web_search = _detect_web_search_intent(sanitized_input)
                        is_netsuite_entity = _detect_netsuite_entity(sanitized_input)  # noqa: F841
                        _has_data_reference = _detect_data_reference(sanitized_input)

                        # Gate financial reports by permission
                        if is_financial:
                            from app.core.dependencies import has_permission

                            can_access_financial = await has_permission(db, user_id, "chat.financial_reports")
                            if not can_access_financial:
                                is_financial = False
                                sanitized_input = (
                                    sanitized_input + "\n\n[SYSTEM: Financial reports are restricted for your role. "
                                    "Respond to the user that financial reports require Admin or User role access. "
                                    "Do NOT attempt to call netsuite_financial_report.]"
                                )
                                print(
                                    "[ORCHESTRATOR] Financial report blocked — user lacks chat.financial_reports permission",
                                    flush=True,
                                )

                        importance_tier = classify_importance(
                            sanitized_input,
                            intent_hint="financial_report" if is_financial else None,
                        )
                        print(
                            f"[ORCHESTRATOR] Importance tier: {importance_tier.label} ({importance_tier.value})",
                            flush=True,
                        )

                        # Follow-up detection: if the previous turn used financial report mode
                        if not is_financial and history_messages:
                            prev_assistant = next(
                                (m["content"] for m in reversed(history_messages) if m["role"] == "assistant"),
                                "",
                            )
                            _financial_history_markers = (
                                "transactionaccountingline",
                                "accttype",
                                "Net Income",
                                "Income Statement",
                                "Balance Sheet",
                                _FINANCIAL_MODE_TAG,
                            )
                            if any(marker in prev_assistant for marker in _financial_history_markers):
                                is_financial = True
                                print("[UNIFIED] Financial follow-up detected from conversation history", flush=True)

                        dk_top_k = 5 if is_financial else None  # default from settings

                        # ── Smart context injection: classify how much context this query needs ──
                        context_need = _classify_context_need(sanitized_input, is_financial=is_financial)
                        print(f"[ORCHESTRATOR] Context need: {context_need}", flush=True)

                        # Injection matrix:
                        #   Block              FULL  DATA  DOCS  WORKSPACE  FINANCIAL
                        #   tenant_schema       ❌    ✅    ❌      ❌        ❌
                        #   table_schemas       ✅    ✅    ❌      ❌        ❌
                        #   tenant_vernacular   ❌    ✅    ❌      ❌        ✅
                        #   domain_knowledge    ❌    ✅    ✅      ❌        ❌
                        #   onboarding_profile  ❌    ❌    ❌      ❌        ✅
                        #   proven_patterns    (gate by tool presence — see _compute_need_patterns)
                        #   learned_rules       ✅    ✅    ✅      ✅        ✅    (always)
                        # FULL = investigation ("why") — minimal context so agent reasons freely.
                        # proven_patterns: gated by tool presence (any SuiteQL/BQ tool in the toolset),
                        # NOT by context_need. Admin-seeded patterns are high-quality and shouldn't be
                        # stranded by the FULL classification. See _compute_need_patterns docstring.

                        _need_vernacular = context_need in (ContextNeed.DATA, ContextNeed.FINANCIAL)
                        _need_domain_knowledge = context_need in (ContextNeed.DATA, ContextNeed.DOCS)
                        _need_patterns = _compute_need_patterns(context_need, _tool_names)
                        _need_schemas = context_need in (ContextNeed.FULL, ContextNeed.DATA)
                        _need_onboarding = context_need in (ContextNeed.FINANCIAL,)
                        _drive_active = any(p.profile_id == "google_drive" for p in _active_profiles)

                        # Assemble context concurrently (only fetch what we need)
                        from app.services.chat.domain_knowledge import retrieve_domain_knowledge
                        from app.services.chat.tenant_resolver import TenantEntityResolver
                        from app.services.learned_rules_service import retrieve_learned_rules
                        from app.services.query_pattern_service import retrieve_similar_patterns

                        # Build the list of concurrent tasks dynamically
                        _gather_tasks = []
                        _gather_keys = []

                        if _need_vernacular:
                            _gather_tasks.append(
                                asyncio.wait_for(
                                    TenantEntityResolver.resolve_entities(
                                        user_message=sanitized_input,
                                        tenant_id=tenant_id,
                                        db=db,
                                        adapter=specialist_adapter,
                                        model=settings.MULTI_AGENT_SPECIALIST_MODEL,
                                    ),
                                    timeout=_RESOLVE_ENTITIES_TIMEOUT_SECONDS,
                                )
                            )
                            _gather_keys.append("vernacular")
                        if _need_domain_knowledge:
                            _gather_tasks.append(
                                retrieve_domain_knowledge(
                                    db=db,
                                    query_text=sanitized_input,
                                    top_k=dk_top_k,
                                    partition_ids=_profile_partitions or None,
                                )
                            )
                            _gather_keys.append("dk")
                        if _drive_active and context_need in (ContextNeed.DATA, ContextNeed.DOCS):
                            _gather_tasks.append(
                                asyncio.wait_for(
                                    _gather_drive_knowledge(db=db, tenant_id=tenant_id, query_text=sanitized_input),
                                    timeout=_GATHER_DRIVE_TIMEOUT_SECONDS,
                                )
                            )
                            _gather_keys.append("drive")
                        if _need_patterns:
                            _gather_tasks.append(retrieve_similar_patterns(db, tenant_id, sanitized_input))
                            _gather_keys.append("patterns")

                        # Learned rules — query-aware: only inject rules relevant to this query
                        _gather_tasks.append(
                            retrieve_learned_rules(db=db, tenant_id=tenant_id, query_text=sanitized_input)
                        )
                        _gather_keys.append("learned_rules")

                        _gather_t0 = time.time()
                        print(f"[ORCHESTRATOR] context_gather start | tasks={_gather_keys}", flush=True)
                        _gather_results = await asyncio.gather(*_gather_tasks, return_exceptions=True)
                        print(
                            f"[ORCHESTRATOR] context_gather complete in {time.time() - _gather_t0:.2f}s",
                            flush=True,
                        )
                        _results = dict(zip(_gather_keys, _gather_results))

                        vernacular_result = _results.get("vernacular")
                        dk_result = _results.get("dk")
                        patterns_result = _results.get("patterns")
                        learned_rules_result = _results.get("learned_rules")

                        context: dict[str, Any] = {}

                        # tenant_vernacular (FULL, DATA, FINANCIAL)
                        if vernacular_result is not None:
                            if isinstance(vernacular_result, asyncio.TimeoutError):
                                print(
                                    f"[ORCHESTRATOR] entity_resolution timed out after "
                                    f"{_RESOLVE_ENTITIES_TIMEOUT_SECONDS}s — proceeding without vernacular",
                                    flush=True,
                                )
                                logger.warning(
                                    "unified_agent.entity_resolution_timeout",
                                    extra={"timeout_seconds": _RESOLVE_ENTITIES_TIMEOUT_SECONDS},
                                )
                            elif isinstance(vernacular_result, Exception):
                                logger.warning("unified_agent.entity_resolution_failed", exc_info=vernacular_result)
                            elif vernacular_result:
                                context["tenant_vernacular"] = vernacular_result
                                print(f"[UNIFIED] Vernacular injected ({len(vernacular_result)} chars)", flush=True)
                                import re as _re

                                conf_scores = [
                                    float(s)
                                    for s in _re.findall(
                                        r"<confidence_score>([\d.]+)</confidence_score>",
                                        vernacular_result,
                                    )
                                ]
                                if conf_scores:
                                    context["entity_resolution_confidence"] = max(conf_scores)

                        # domain_knowledge (FULL, DATA, DOCS — skip for FINANCIAL & WORKSPACE)
                        if dk_result is not None:
                            if is_financial:
                                context["domain_knowledge"] = []
                            elif isinstance(dk_result, Exception):
                                logger.warning("unified_agent.domain_knowledge_failed", exc_info=dk_result)
                            elif dk_result:
                                context["domain_knowledge"] = [r["raw_text"] for r in dk_result]
                                print(f"[UNIFIED] Domain knowledge injected ({len(dk_result)} chunks)", flush=True)
                                sims = [r["similarity"] for r in dk_result if r.get("similarity")]
                                if sims:
                                    context["domain_knowledge_similarity"] = sum(sims) / len(sims)

                        # drive_knowledge (DATA, DOCS when google_drive profile is active)
                        drive_result = _results.get("drive")
                        if drive_result is not None:
                            if isinstance(drive_result, asyncio.TimeoutError):
                                print(
                                    f"[ORCHESTRATOR] drive_knowledge timed out after "
                                    f"{_GATHER_DRIVE_TIMEOUT_SECONDS}s — proceeding without drive context",
                                    flush=True,
                                )
                                logger.warning(
                                    "unified_agent.drive_knowledge_timeout",
                                    extra={"timeout_seconds": _GATHER_DRIVE_TIMEOUT_SECONDS},
                                )
                            elif isinstance(drive_result, Exception):
                                logger.warning("unified_agent.drive_knowledge_failed", exc_info=drive_result)
                            elif drive_result.get("chunks"):
                                context["drive_knowledge"] = _build_drive_knowledge_block(drive_result["chunks"])
                                context["drive_sources"] = drive_result["sources"]
                                print(
                                    f"[UNIFIED] Drive knowledge injected ({len(drive_result['chunks'])} chunks, "
                                    f"{len(drive_result['sources'])} sources)",
                                    flush=True,
                                )

                        # Hoist user-inserted Drive mentions ([Name](drive_url)) into
                        # drive_sources so citations resolve even when RAG returned 0
                        # chunks for that file. Always runs when google_drive profile
                        # is active, regardless of whether retrieval produced chunks.
                        if _drive_active:
                            context["drive_mentions"] = _merge_drive_mentions(context, sanitized_input)

                        # proven_patterns (gate by tool presence — injects when _need_patterns is True)
                        if patterns_result is not None:
                            if isinstance(patterns_result, Exception):
                                logger.warning("unified_agent.proven_patterns_failed", exc_info=patterns_result)
                            elif patterns_result:
                                context["proven_patterns"] = patterns_result
                                print(
                                    f"[UNIFIED] Proven patterns injected ({len(patterns_result)} patterns)", flush=True
                                )
                                context["matched_pattern_similarity"] = patterns_result[0].get("similarity", 0.0)
                                context["matched_pattern_success_count"] = patterns_result[0].get("success_count", 0)

                        # learned_rules (ALL context needs — always injected)
                        if learned_rules_result is not None:
                            if isinstance(learned_rules_result, Exception):
                                logger.warning("unified_agent.learned_rules_failed", exc_info=learned_rules_result)
                            elif learned_rules_result:
                                context["learned_rules"] = learned_rules_result
                                print(
                                    f"[ORCHESTRATOR] Learned rules injected ({len(learned_rules_result)} rules)",
                                    flush=True,
                                )

                        context["user_timezone"] = user_timezone
                        context["importance_tier"] = importance_tier.value

                        # Fiscal calendar — inject into context so agents interpret
                        # Q1/Q2/Q3/Q4 and "fiscal year" queries using the tenant's
                        # actual FY start month instead of defaulting to calendar.
                        _fy_start = 1
                        if _tenant_config_row:
                            _fy_start = getattr(_tenant_config_row, "fiscal_year_start_month", 1) or 1
                        context["fiscal_year_start_month"] = _fy_start

                        # table_schemas (FULL, DATA — skip for DOCS, WORKSPACE, FINANCIAL)
                        if not _need_schemas:
                            print(f"[ORCHESTRATOR] Skipping schema injection for {context_need} query", flush=True)
                        else:
                            try:
                                from app.services.prompt_template_service import _build_table_schema_section
                                from app.services.schema_context_selector import select_relevant_schemas

                                entity_types: list[str] = []
                                if isinstance(vernacular_result, str):
                                    import re as _re_schema

                                    entity_types = _re_schema.findall(
                                        r"<entity_type>(.*?)</entity_type>", vernacular_result
                                    )

                                relevant_tables = select_relevant_schemas(
                                    sanitized_input,
                                    entity_types=entity_types,
                                )
                                print(f"[ORCHESTRATOR] Schema tables selected: {relevant_tables}", flush=True)

                                schema_xml = _build_table_schema_section(
                                    metadata=metadata,
                                    relevant_tables=relevant_tables,
                                )
                                if schema_xml:
                                    context["table_schemas"] = schema_xml
                                    print(
                                        f"[ORCHESTRATOR] Schema injected ({len(schema_xml)} chars, "
                                        f"{len(relevant_tables)} tables)",
                                        flush=True,
                                    )
                            except Exception:
                                logger.warning("orchestrator.schema_injection_failed", exc_info=True)

                        # onboarding_profile (FULL, FINANCIAL — skip for DATA, DOCS, WORKSPACE)
                        if not _need_onboarding:
                            print(f"[ORCHESTRATOR] Skipping onboarding profile for {context_need} query", flush=True)
                        else:
                            try:
                                _onboarding_profile = (
                                    getattr(_tenant_config_row, "onboarding_profile", None)
                                    if _tenant_config_row
                                    else None
                                )
                                if _onboarding_profile and isinstance(_onboarding_profile, dict):
                                    profile_parts: list[str] = []

                                    # Transaction landscape
                                    txn_types = _onboarding_profile.get("transaction_types", [])
                                    if txn_types:
                                        txn_summary = "\n".join(
                                            f"  - {t['type']}: {t['count']} records ({t.get('earliest', '?')} to {t.get('latest', '?')})"
                                            for t in txn_types[:15]
                                        )
                                        profile_parts.append(
                                            f"<tenant_transaction_landscape>\nThis tenant uses these transaction types:\n{txn_summary}\n</tenant_transaction_landscape>"
                                        )

                                    # Relationship map
                                    rels = _onboarding_profile.get("transaction_relationships", [])
                                    if rels:
                                        rel_summary = "\n".join(
                                            f"  - {r['source']} -> {r['target']} ({r['count']} links)"
                                            for r in rels[:20]
                                        )
                                        profile_parts.append(
                                            f"<transaction_relationships>\n{rel_summary}\n</transaction_relationships>"
                                        )

                                    # Status codes
                                    status_map = _onboarding_profile.get("status_codes", {})
                                    if status_map:
                                        status_lines: list[str] = []
                                        for _txn_type, codes in list(status_map.items())[:10]:
                                            for c in codes[:8]:
                                                status_lines.append(
                                                    f"  - {_txn_type} status '{c['code']}' = {c['display']} ({c['count']})"
                                                )
                                        if status_lines:
                                            profile_parts.append(
                                                "<tenant_status_codes>\n"
                                                + "\n".join(status_lines)
                                                + "\n</tenant_status_codes>"
                                            )

                                    if profile_parts:
                                        profile_xml = "\n".join(profile_parts)
                                        context["onboarding_profile"] = profile_xml
                                        print(
                                            f"[ORCHESTRATOR] Onboarding profile injected ({len(profile_xml)} chars)",
                                            flush=True,
                                        )
                            except Exception:
                                logger.warning("orchestrator.onboarding_profile_injection_failed", exc_info=True)

                        # Three-tier routing replaced by knowledge profiles — inject
                        # domain context based on available tools instead of routing
                        # to specialized agents.  UnifiedAgent handles all queries.
                        _selected_agent_id = None
                        if agent_id and not is_financial:
                            # Client-side agent pin — noted for audit, but UnifiedAgent handles
                            _selected_agent_id = agent_id
                            print(f"[KNOWLEDGE] Client pinned agent_id={agent_id} (audit only)", flush=True)

                        # Knowledge profile matching — inject domain context
                        from app.services.chat.prompt_assembler import (
                            assemble_knowledge_context,
                            build_disambiguation_instruction,
                            build_source_pin_hint,
                        )

                        if _active_profiles:
                            print(
                                f"[KNOWLEDGE] Active profiles: {[p.profile_id for p in _active_profiles]}",
                                flush=True,
                            )

                        knowledge_context = assemble_knowledge_context(_active_profiles)
                        if knowledge_context:
                            system_prompt += f"\n\n{knowledge_context}"
                        if context.get("drive_knowledge"):
                            system_prompt += "\n\n" + context["drive_knowledge"]
                        drive_mentions = context.get("drive_mentions") or []
                        if drive_mentions:
                            system_prompt += _build_drive_mentions_hint(drive_mentions)
                        disambiguation = build_disambiguation_instruction(_active_profiles)
                        if disambiguation:
                            system_prompt += disambiguation
                        pin_hint = build_source_pin_hint(getattr(session, "source_pin", None))
                        if pin_hint:
                            system_prompt += pin_hint

                        # Plan Mode augmentation — `plan_mode_enabled` is computed
                        # earlier (with the tool inventory) so we don't refetch the
                        # flag here. Imported here to keep the import close to use.
                        from app.services.chat.plan_mode.ambiguity_signal import (
                            is_financial_ambiguous,
                            maybe_augment_for_plan_mode,
                            try_force_tool_choice,
                        )

                        # On resume turns (`plan_mode_resume_source` is set) we
                        # MUST skip the clarify augmentation. `sanitized_input`
                        # is still the original financial query, so
                        # `is_financial_ambiguous` would re-fire and the
                        # augmentation block would tell the model "your only
                        # first action is `clarify`" — but the resume tool
                        # filter has stripped `clarify` from the inventory.
                        # Contradictory instructions ⇒ undefined behavior.
                        _plan_mode_connected_sources: list[str] = []
                        if (
                            plan_mode_enabled
                            and not _plan_mode_resume_active
                            and is_financial_ambiguous(sanitized_input)
                        ):
                            # Resolve connected sources so the augmentation can
                            # require options to span distinct sources when ≥2
                            # are connected. Mirrors the post-call resolution
                            # in base_agent.py for clarify_intercept.
                            from app.services.chat.plan_mode.source_resolver import (
                                canonicalize_connector_providers,
                            )
                            from app.services.connection_service import list_connections
                            from app.services.mcp_connector_service import (
                                get_active_connectors_for_tenant,
                            )

                            try:
                                _mcp = await get_active_connectors_for_tenant(db, tenant_id)
                            except Exception:
                                _mcp = []
                            try:
                                _rest = await list_connections(db, tenant_id)
                            except Exception:
                                _rest = []
                            _raw_providers = [getattr(c, "provider", "") for c in _mcp] + [
                                getattr(c, "provider", "") for c in _rest if getattr(c, "status", "active") == "active"
                            ]
                            _plan_mode_connected_sources = sorted(canonicalize_connector_providers(_raw_providers))

                        plan_mode_augmentation = (
                            maybe_augment_for_plan_mode(
                                query=sanitized_input,
                                plan_mode_enabled=plan_mode_enabled,
                                connected_sources=_plan_mode_connected_sources,
                            )
                            if not _plan_mode_resume_active
                            else None
                        )

                        # Always use UnifiedAgent — no routing fork
                        unified_agent = UnifiedAgent(
                            tenant_id=tenant_id,
                            user_id=user_id,
                            correlation_id=correlation_id,
                            metadata=metadata if _need_schemas else None,
                            policy=active_policy,
                            context_need=context_need,
                        )

                        # Pass Plan Mode injections to the agent via instance
                        # attributes — UnifiedAgent.system_prompt is a property
                        # that builds its own prompt from _SYSTEM_PROMPT + tool
                        # inventory + metadata, so it never reads the orchestrator's
                        # local `system_prompt` variable. Setting the attributes
                        # routes the augmentation/directive into that property.
                        #
                        # Order in the agent: augmentation first, resume directive
                        # last (so the resume directive overrides initial gate intent).
                        if plan_mode_augmentation:
                            unified_agent._plan_mode_augmentation = plan_mode_augmentation
                        if plan_mode_resume_directive:
                            unified_agent._plan_mode_resume_directive = plan_mode_resume_directive

                    # Classify follow-up intent (TRANSFORM vs NEW_DATA)
                    from app.services.chat.follow_up_classifier import FollowUpIntent, classify_follow_up
                    from app.services.chat.result_cache import (
                        CachedResult,
                        _cache_result_sync,
                        get_latest_result,
                    )

                    _follow_up_intent = FollowUpIntent.NEW_DATA
                    _cached_result = None
                    if not _is_chitchat and not is_financial and session.messages:
                        _cached_result = await get_latest_result(str(session.id))
                        _follow_up_intent = classify_follow_up(
                            message=sanitized_input,
                            has_previous_result=_cached_result is not None,
                        )
                        if _follow_up_intent == FollowUpIntent.TRANSFORM:
                            print("[ORCHESTRATOR] TRANSFORM intent — using cached result", flush=True)

                    # Augment task for financial report queries or transform requests
                    unified_task = sanitized_input
                    if not _is_chitchat and is_financial:
                        unified_task = _build_financial_mode_task(sanitized_input)
                        print("[UNIFIED] Financial report mode activated (SuiteQL + CONSOLIDATE)", flush=True)
                    elif is_web_search:
                        unified_task = (
                            f"{sanitized_input}\n\n"
                            "OVERRIDE: The user explicitly asked for a web search. You MUST call the web_search "
                            "tool FIRST before any database queries. Extract the search topic from their message "
                            "and search the web."
                        )
                        print("[ORCHESTRATOR] Web search override activated", flush=True)
                    elif _has_data_reference and _cached_result:
                        unified_task = (
                            f"{sanitized_input}\n\n"
                            f"[CACHED DATA AVAILABLE] The user is referencing data from earlier in this conversation. "
                            f"Previous result ({_cached_result.row_count} rows, type: {_cached_result.result_type}) "
                            f"is available via the reference_previous_result tool. "
                            f"Use it instead of re-querying. Do NOT run a new database query."
                        )
                        print("[ORCHESTRATOR] Data reference detected — using cached result", flush=True)
                    elif _follow_up_intent == FollowUpIntent.TRANSFORM and _cached_result:
                        unified_task = (
                            f"{sanitized_input}\n\n"
                            f"[CACHED DATA AVAILABLE] The previous result ({_cached_result.row_count} rows, "
                            f"type: {_cached_result.result_type}, columns: {_cached_result.columns[:10]}) "
                            f"is available via the reference_previous_result tool. "
                            f"Use it instead of re-querying NetSuite or BigQuery."
                        )

                    if attached_file_context:
                        unified_task = f"{unified_task}\n\n{attached_file_context}"

                    streamed_text_parts: list[str] = []
                    agent_result = None
                    last_structured_output: dict | None = None
                    suppress_streamed_text = False
                    _charts_output: list[dict] = []

                    unified_model = model if is_byok else settings.MULTI_AGENT_SQL_MODEL

                    # Route simple lookups to Haiku for 10x speed + cost savings
                    # Only for non-BYOK tenants (BYOK users chose their model)
                    if (
                        not is_byok
                        and not _is_chitchat
                        and not is_financial
                        and importance_tier.value <= 2
                        and _is_simple_lookup(sanitized_input)
                    ):
                        unified_model = HAIKU_MODEL
                        print("[ORCHESTRATOR] Simple lookup detected — routing to Haiku", flush=True)

                    # Track whether we're inside a <chart> block during streaming
                    _in_chart_block = False
                    _chart_buffer = ""

                    # Cache callback: write intercepted results to Redis IMMEDIATELY.
                    # Same-turn follow-ups (e.g. pricing_export → pricing_to_sheets in
                    # one assistant message) need to read each other's writes via
                    # get_latest_result_by_type before the agent loop completes — a
                    # deferred flush would always show them stale data. Each entry
                    # gets a unique synthetic id so multiple tool calls within one
                    # turn don't collide on a shared message_id; an alias under
                    # assistant_msg.id is added at flush time below for
                    # reference_previous_result(message_id=...) lookups.
                    _pending_caches: list[CachedResult] = []

                    def _on_tool_intercepted(tool_name: str, event_type_str: str, event_data: dict):
                        cr = _build_intercept_cache_entry(
                            tool_name=tool_name,
                            event_type_str=event_type_str,
                            event_data=event_data,
                            conversation_id=str(session.id),
                        )
                        if cr is None:
                            return
                        # Eager write so same-turn follow-ups can read it.
                        _cache_result_sync(str(session.id), cr.message_id, cr)
                        _pending_caches.append(cr)

                    # Emit Drive source map once per turn so the frontend can resolve
                    # [source_name] citations to clickable Drive links.
                    if context.get("drive_sources"):
                        yield {"type": "drive_sources", "sources": context["drive_sources"]}

                    # ── Plan Mode hard gate ─────────────────────────────────
                    # When financial-ambiguity regex matches AND flag is on AND
                    # clarify is in the inventory, pass `plan_mode_clarify_only=True`
                    # to run_streaming (it filters _tool_defs to clarify-only AFTER
                    # _setup_context, so we don't fight the rebuild) and force
                    # tool_choice=clarify on the adapter. plan_mode_enabled is only
                    # set on the non-chitchat path, so guard with _is_chitchat.
                    _plan_mode_active = False
                    _plan_mode_tool_choice: dict | None = None
                    # Mutual exclusion: when EITHER resume variant is active
                    # (source-pick or manual clarify), the user already
                    # disambiguated. Re-firing the new-clarify gate would
                    # force another card on top of the resumed turn.
                    if (
                        not _is_chitchat
                        and plan_mode_enabled
                        and not _plan_mode_resume_active
                        and is_financial_ambiguous(sanitized_input)
                    ):
                        _has_clarify = any(t.get("name") == "clarify" for t in (tool_definitions or []))
                        if _has_clarify:
                            _plan_mode_tool_choice = try_force_tool_choice(
                                specialist_adapter, "clarify", model=unified_model
                            )
                            if _plan_mode_tool_choice is not None:
                                _plan_mode_active = True
                            # else: adapter unsupported (e.g., Gemini 1.0) — gate stays off.
                        else:
                            logger.warning("[PLAN_MODE] flag on but clarify tool missing — skipping gate")

                    async for event_type, payload in unified_agent.run_streaming(
                        task=unified_task,
                        context=context,
                        db=db,
                        adapter=specialist_adapter,
                        model=unified_model,
                        conversation_history=history_messages,
                        tool_choice=_plan_mode_tool_choice,
                        plan_mode_clarify_only=_plan_mode_active,
                        plan_mode_resume_source=plan_mode_resume_source,
                        tool_result_interceptor=_make_tool_interceptor(
                            context_need, cache_callback=_on_tool_intercepted
                        ),
                        session_id=str(session.id),
                        run_id=run_id,
                    ):
                        if event_type == "text":
                            streamed_text_parts.append(payload)
                            if suppress_streamed_text:
                                continue
                            # Suppress <chart> blocks from streaming text output
                            _chart_buffer += payload
                            while True:
                                if not _in_chart_block:
                                    # Look for <chart> tag start
                                    idx = _chart_buffer.find("<chart>")
                                    if idx == -1:
                                        # No chart tag — yield buffered text (keep last 7 chars in case of partial tag)
                                        safe = _chart_buffer[:-7] if len(_chart_buffer) > 7 else ""
                                        _chart_buffer = _chart_buffer[len(safe) :]
                                        if safe:
                                            yield {"type": "text", "content": safe}
                                        break
                                    else:
                                        # Yield text before the chart tag
                                        if idx > 0:
                                            yield {"type": "text", "content": _chart_buffer[:idx]}
                                        _chart_buffer = _chart_buffer[idx:]
                                        _in_chart_block = True
                                else:
                                    # Inside chart block — look for </chart>
                                    end_idx = _chart_buffer.find("</chart>")
                                    if end_idx == -1:
                                        break  # Wait for more data
                                    # Found end — discard the entire <chart>...</chart> block
                                    _chart_buffer = _chart_buffer[end_idx + 8 :]
                                    _in_chart_block = False
                        elif event_type == "tool_status":
                            # Flush chart buffer before tool execution — text is paused
                            if _chart_buffer and not _in_chart_block:
                                yield {"type": "text", "content": _chart_buffer}
                                _chart_buffer = ""
                            yield {"type": "tool_status", "content": payload}
                        elif event_type == "tool_start":
                            yield {"type": "tool_start", **payload}
                        elif event_type == "tool_end":
                            yield {"type": "tool_end", **payload}
                        elif event_type == "tool_intercept":
                            # payload is (event_type_str, event_data_dict)
                            last_structured_output = {"type": payload[0], "data": payload[1]}
                            yield {"type": payload[0], "data": payload[1]}
                            if _is_pricing_task_output(last_structured_output):
                                suppress_streamed_text = True
                                _chart_buffer = ""
                                _in_chart_block = False

                            # Auto-generate chart from financial reports (deterministic, no LLM)
                            if payload[0] == "financial_report":
                                from app.services.chat.financial_chart_builder import build_financial_chart

                                _fr_data = payload[1]
                                _fr_chart = build_financial_chart(
                                    report_type=_fr_data.get("report_type", ""),
                                    summary=_fr_data.get("summary", {}),
                                )
                                if _fr_chart:
                                    _chart_dict = _fr_chart.model_dump()
                                    yield {"type": "chart", "data": _chart_dict}
                                    _charts_output.append(_chart_dict)
                                    print(
                                        f"[ORCHESTRATOR] Auto-generated chart for {_fr_data.get('report_type')}",
                                        flush=True,
                                    )
                        elif event_type == "confirmation_required":
                            last_structured_output = {"type": "write_confirmation", **payload}
                        elif event_type == "clarification_required":
                            # Plan Mode HITL clarification — payload IS the structured_output
                            last_structured_output = dict(payload)
                            # Yield SSE event so frontend renders the card
                            yield {"type": "clarification_required", "data": payload}
                        elif event_type == "response":
                            agent_result = payload

                    # Flush remaining buffered text (non-chart content)
                    if _chart_buffer and not _in_chart_block and not suppress_streamed_text:
                        yield {"type": "text", "content": _chart_buffer}

                    if agent_result is None:
                        final_text = (
                            _sanitize_assistant_text("".join(streamed_text_parts))
                            or "I wasn't able to process that request."
                        )
                        coord_result_tokens = (0, 0)
                        coord_result_cache = (0, 0)
                        coord_result_tool_calls: list[dict] = []
                    else:
                        final_text = _sanitize_assistant_text(agent_result.data or "")
                        coord_result_tokens = (
                            agent_result.tokens_used.input_tokens,
                            agent_result.tokens_used.output_tokens,
                        )
                        coord_result_cache = (
                            agent_result.tokens_used.cache_creation_input_tokens,
                            agent_result.tokens_used.cache_read_input_tokens,
                        )
                        coord_result_tool_calls = agent_result.tool_calls_log

                    # Extract charts from agent response text → emit as SSE events
                    from app.services.chat.chart_extractor import extract_charts

                    final_text, charts = extract_charts(final_text)
                    for chart in charts:
                        _chart_dict = chart.model_dump()
                        yield {"type": "chart", "data": _chart_dict}
                        _charts_output.append(_chart_dict)
                    if charts:
                        print(f"[ORCHESTRATOR] Extracted {len(charts)} chart(s) from agent response", flush=True)

                    # Extract confidence score from agent result
                    confidence_val = (
                        agent_result.confidence_score
                        if agent_result and agent_result.confidence_score is not None
                        else None
                    )
                    if confidence_val is not None:
                        yield {
                            "type": "confidence",
                            "score": confidence_val,
                            "explanation": _get_confidence_explanation(confidence_val),
                        }

                    # Emit importance tier SSE event
                    yield {
                        "type": "importance",
                        "tier": importance_tier.value,
                        "label": importance_tier.label,
                        "needs_review": (
                            agent_result is not None
                            and any(
                                isinstance(tc.get("result"), dict)
                                and tc.get("result", {}).get("judge_verdict", {}).get("needs_review")
                                for tc in (agent_result.tool_calls_log or [])
                            )
                        ),
                    }

                    # Persist charts alongside structured_output (backward compatible)
                    _persisted_output = last_structured_output
                    if _charts_output:
                        if _persisted_output:
                            _persisted_output = {**_persisted_output, "charts": _charts_output}
                        else:
                            _persisted_output = {"charts": _charts_output}

                    assistant_msg = ChatMessage(
                        tenant_id=tenant_id,
                        session_id=session.id,
                        role="assistant",
                        content=_coerce_assistant_content(
                            final_text,
                            _persisted_output,
                            tool_calls=coord_result_tool_calls,
                        ),
                        tool_calls=coord_result_tool_calls if coord_result_tool_calls else None,
                        citations=citations if citations else None,
                        token_count=coord_result_tokens[0] + coord_result_tokens[1],
                        input_tokens=coord_result_tokens[0],
                        output_tokens=coord_result_tokens[1],
                        cache_creation_tokens=coord_result_cache[0],
                        cache_read_tokens=coord_result_cache[1],
                        model_used=unified_model,
                        provider_used=provider if is_byok else settings.MULTI_AGENT_SPECIALIST_PROVIDER,
                        is_byok=is_byok,
                        confidence_score=confidence_val,
                        query_importance=importance_tier.value,
                        structured_output=_persisted_output,
                        agent_id=_selected_agent_id if _selected_agent_id else None,
                        created_at=datetime.now(timezone.utc),
                    )
                    db.add(assistant_msg)

                    # Plan Mode telemetry — clarification_pending event row.
                    # The chat_disclosure_events table survives chat history compaction
                    # (the system-prompt assembler re-reads it each turn) so the agent
                    # remembers prior clarifications across a session.
                    if isinstance(_persisted_output, dict) and _persisted_output.get("type") == "clarification":
                        from app.models.chat_disclosure_event import ChatDisclosureEvent

                        # codex round 6 Bug 2 — `assistant_msg.id` uses
                        # SQLAlchemy's Python-side ``default=uuid.uuid4`` which
                        # only fires at flush time. Without this explicit flush
                        # we'd persist ``chat_message_id=NULL`` on the
                        # disclosure event and lose the link from the telemetry
                        # row to the clarification message it describes.
                        await db.flush()

                        db.add(
                            ChatDisclosureEvent(
                                tenant_id=tenant_id,
                                chat_session_id=session.id,
                                chat_message_id=assistant_msg.id,
                                event_type="clarification_pending",
                                payload={
                                    "options": _persisted_output.get("options", []),
                                    "default_id": _persisted_output.get("default_id"),
                                    "ambiguity_summary": _persisted_output.get("ambiguity_summary", ""),
                                },
                            )
                        )

                    # Eager writes already happened in _on_tool_intercepted under
                    # unique synthetic ids (so same-turn follow-ups can read each
                    # other's writes). Now also write an ALIAS under the persisted
                    # assistant_msg.id so reference_previous_result(message_id=...)
                    # still resolves. Multi-entry turns share the alias key — the
                    # latest entry wins, matching prior behavior.
                    for _pc in _pending_caches:
                        _pc.message_id = str(assistant_msg.id)
                        _cache_result_sync(str(session.id), str(assistant_msg.id), _pc)

                    if not session.title:
                        session.title = user_message[:100].strip()
                    session.updated_at = func.now()

                    audit_payload: dict[str, Any] = {
                        "mode": "unified_agent",
                        "provider": settings.MULTI_AGENT_SPECIALIST_PROVIDER,
                        "model": settings.MULTI_AGENT_SQL_MODEL,
                        "steps": len(coord_result_tool_calls),
                        "input_tokens": coord_result_tokens[0],
                        "output_tokens": coord_result_tokens[1],
                        "cache_creation_tokens": coord_result_cache[0],
                        "cache_read_tokens": coord_result_cache[1],
                        "doc_chunks_count": len(state.doc_chunks) if state.doc_chunks else 0,
                        "tools_called": [t["tool"] for t in coord_result_tool_calls],
                    }
                    if coord_result_cache[1] > 0:
                        # Log cache savings for visibility
                        print(f"[ORCHESTRATOR] Cache hit: {coord_result_cache[1]:,} tokens read from cache", flush=True)
                    await log_event(
                        db=db,
                        tenant_id=tenant_id,
                        category="chat",
                        action="chat.turn",
                        actor_id=user_id,
                        resource_type="chat_session",
                        resource_id=str(session.id),
                        correlation_id=correlation_id,
                        payload=audit_payload,
                    )

                    if not is_byok:
                        await deduct_chat_credits(db, tenant_id, settings.MULTI_AGENT_SQL_MODEL)

                    await db.commit()

                    # Auto-pin after a successful turn so follow-ups stick with the
                    # data source the user was actually working with.
                    try:
                        _pin_update = _compute_source_pin_update(coord_result_tool_calls)
                        if _pin_update == "leave_pin":
                            pass
                        elif _pin_update is None:
                            if getattr(session, "source_pin", None) is not None:
                                session.source_pin = None
                                await db.commit()
                        elif getattr(session, "source_pin", None) != _pin_update:
                            session.source_pin = _pin_update
                            await db.commit()
                    except Exception:
                        logger.warning("auto_source_pin_update_failed", exc_info=True)

                    asyncio.create_task(
                        _dispatch_memory_update(
                            tenant_id=tenant_id,
                            user_id=user_id,
                            db=db,
                            user_message=sanitized_input,
                            assistant_message=final_text,
                        )
                    )
                    asyncio.create_task(
                        _dispatch_content_summary(
                            db=db,
                            message_id=assistant_msg.id,
                            user_message=sanitized_input,
                            assistant_message=final_text,
                        )
                    )

                    result_msg = {
                        "id": str(assistant_msg.id),
                        "role": assistant_msg.role,
                        "content": assistant_msg.content,
                        "tool_calls": assistant_msg.tool_calls,
                        "citations": assistant_msg.citations,
                    }
                    if hasattr(assistant_msg, "created_at") and assistant_msg.created_at:
                        result_msg["created_at"] = assistant_msg.created_at.isoformat()
                    if confidence_val is not None:
                        result_msg["confidence_score"] = confidence_val
                    result_msg["query_importance"] = importance_tier.value
                    if _selected_agent_id:
                        result_msg["agent_id"] = _selected_agent_id
                    # Mirror `assistant_msg.structured_output` into the SSE
                    # payload so the frontend can render clarification /
                    # write-confirmation cards immediately on the terminal
                    # `message` event without waiting for a session refetch.
                    if assistant_msg.structured_output:
                        result_msg["structured_output"] = assistant_msg.structured_output
                    yield {"type": "message", "message": result_msg}
                    return

                # Legacy multi-agent path removed — unified agent handles all queries

        # ── Split system prompt for Anthropic prompt caching ──
        prompt_parts = split_system_prompt(system_prompt)

        # ── Single-agent agentic loop (default path) ──
        tool_calls_log: list[dict] = []
        final_text = ""
        total_input_tokens = 0
        total_output_tokens = 0
        total_cache_creation_tokens = 0
        total_cache_read_tokens = 0
        last_structured_output: dict | None = None
        suppress_streamed_text = False
        # Dedup workspace_propose_patch per canonical file path across the
        # whole turn so a single model response that emits two identical
        # patch tool_uses doesn't create two draft changesets. Maps
        # canonical_path -> changeset_id so the skip result can echo the real
        # prior id back to the model instead of asking it to invent one.
        # Only populated after a *successful* execution so a failed first
        # patch doesn't block a corrected retry in the same turn.
        # Mirrors the guard in BaseSpecialistAgent.run_streaming.
        patched_files: dict[str, str] = {}

        for step in range(MAX_STEPS):
            response = None

            # Determine if we should stream using stream_message or fallback
            # (Assuming all adapters implemented stream_message, else this would fail)
            async for event_type, payload in adapter.stream_message(
                model=model,
                max_tokens=16384,
                system=prompt_parts.static,
                system_dynamic=prompt_parts.dynamic,
                messages=messages,
                tools=tool_definitions if tool_definitions else None,
            ):
                if event_type == "text":
                    if not suppress_streamed_text:
                        yield {"type": "text", "content": payload}
                elif event_type == "response":
                    response = payload

            if not response:
                break

            total_input_tokens += response.usage.input_tokens
            total_output_tokens += response.usage.output_tokens
            total_cache_creation_tokens += response.usage.cache_creation_input_tokens
            total_cache_read_tokens += response.usage.cache_read_input_tokens

            if not response.tool_use_blocks:
                # Pure text response — we're done
                final_text = "\n".join(response.text_blocks) if response.text_blocks else ""
                break

            # Append assistant message
            messages.append(adapter.build_assistant_message(response))

            # Execute each tool call and collect results
            tool_results_content = []
            for block in response.tool_use_blocks:
                # Auto-inject workspace_id for workspace tools
                if block.name.startswith("workspace_"):
                    if "workspace_id" not in block.input or not _is_valid_uuid(block.input.get("workspace_id", "")):
                        if workspace_context:
                            block.input["workspace_id"] = workspace_context["workspace_id"]
                        elif not workspace_context:
                            # Auto-resolve default workspace for non-workspace sessions
                            resolved_ws_id = await _resolve_default_workspace(db, tenant_id)
                            if resolved_ws_id:
                                block.input["workspace_id"] = resolved_ws_id
                                if not workspace_context:
                                    workspace_context = {"workspace_id": resolved_ws_id, "name": "auto-resolved"}

                # Dedup: skip duplicate workspace_propose_patch for same file
                # within this turn. Without this, the LLM occasionally emits two
                # identical tool_use blocks in a single response and the user
                # ends up with two draft changesets to approve. The agent still
                # sees a tool_result so the conversation flows normally.
                #
                # Use the canonicalised path (matching workspace_service
                # .validate_path) as the key so the LLM can't bypass the guard
                # by varying the prefix ('./foo.js' vs 'foo.js' vs 'foo.js ').
                # Falls back to the raw input if normalization raises — the
                # tool call would fail anyway, so the dedup match doesn't
                # matter.
                _dedup_patch_key: str | None = None
                if block.name == "workspace_propose_patch":
                    from app.services.workspace_service import validate_path as _ws_validate_path

                    _raw_patch_path = block.input.get("file_path", "")
                    try:
                        _dedup_patch_key = _ws_validate_path(_raw_patch_path) if _raw_patch_path else None
                    except ValueError:
                        _dedup_patch_key = _raw_patch_path or None
                    if _dedup_patch_key and _dedup_patch_key in patched_files:
                        _prior_cs_id = patched_files[_dedup_patch_key]
                        print(
                            f"[WORKSPACE] Skipping duplicate patch for {_dedup_patch_key} "
                            f"(reusing changeset_id={_prior_cs_id})",
                            flush=True,
                        )
                        tool_results_content.append(
                            {
                                "type": "tool_result",
                                "tool_use_id": block.id,
                                "content": json.dumps(
                                    {
                                        "skipped": "duplicate_patch_same_turn",
                                        "message": (
                                            f"Already proposed a patch for '{_dedup_patch_key}' this turn. "
                                            "Use the changeset_id below; do not call workspace_propose_patch again."
                                        ),
                                        "changeset_id": _prior_cs_id,
                                    }
                                ),
                            }
                        )
                        tool_calls_log.append(
                            {
                                "step": step,
                                "tool": block.name,
                                "params": block.input,
                                "result_summary": json.dumps(
                                    {
                                        "skipped": "duplicate_patch_same_turn",
                                        "changeset_id": _prior_cs_id,
                                    }
                                ),
                                "duration_ms": 0,
                            }
                        )
                        continue

                yield {"type": "tool_status", "content": f"Executing {block.name}..."}
                yield {
                    "type": "tool_start",
                    "tool_name": block.name,
                    "tool_input": block.input,
                    "step": step,
                }

                # Route tool execution: onboarding tools vs standard tools
                t0 = time.monotonic()
                if is_onboarding and block.name in (
                    "save_onboarding_profile",
                    "start_netsuite_oauth",
                    "check_netsuite_connection",
                ):
                    result_str = await execute_onboarding_tool(
                        tool_name=block.name,
                        tool_input=block.input,
                        tenant_id=tenant_id,
                        user_id=user_id,
                        db=db,
                    )
                else:
                    # Policy evaluation: check if tool call is allowed
                    policy_result = policy_evaluate(active_policy, block.name, block.input)
                    if not policy_result["allowed"]:
                        result_str = json.dumps(
                            {"error": f"Policy blocked: {policy_result.get('reason', 'Not allowed')}"}
                        )
                    else:
                        result_str = await execute_tool_call(
                            tool_name=block.name,
                            tool_input=block.input,
                            tenant_id=tenant_id,
                            actor_id=user_id,
                            correlation_id=correlation_id,
                            db=db,
                            session_id=str(session.id),
                        )

                        # Output redaction: strip blocked fields from tool results
                        if active_policy and active_policy.blocked_fields:
                            try:
                                parsed = json.loads(result_str)
                                parsed = redact_output(active_policy, parsed)
                                result_str = json.dumps(parsed, default=str)
                            except (json.JSONDecodeError, TypeError):
                                pass  # Non-JSON result, skip redaction

                # Intercept data tool results: emit SSE event with full data,
                # condense result_str to summary-only for the LLM. Routed
                # through _intercept_with_cache so pricing follow-up tools
                # (pricing_revise / pricing_to_sheets) can read the cached
                # pricing_state in this single-agent / legacy path too.
                intercept_type, intercept_data, result_str = _intercept_with_cache(
                    block.name,
                    result_str,
                    context_need=ContextNeed.FULL,
                    session_id=str(session.id),
                )
                if intercept_type is not None:
                    last_structured_output = {"type": intercept_type, "data": intercept_data}
                    yield {"type": intercept_type, "data": intercept_data}
                    if _is_pricing_task_output(last_structured_output):
                        suppress_streamed_text = True

                tool_results_content.append(
                    {
                        "type": "tool_result",
                        "tool_use_id": block.id,
                        "content": result_str,
                    }
                )

                # Log for audit
                elapsed_ms = int((time.monotonic() - t0) * 1000)

                _result_dict = {"result_summary": result_str}
                _row_count = tool_call_row_count(_result_dict)
                _had_error = tool_call_had_error(_result_dict)
                _summary = (
                    f"{_row_count} rows returned"
                    if _row_count and not _had_error
                    else ("Error" if _had_error else "Done")
                )
                yield {
                    "type": "tool_end",
                    "tool_name": block.name,
                    "step": step,
                    "duration_ms": elapsed_ms,
                    "success": not _had_error,
                    "result_summary": _summary,
                }

                tool_calls_log.append(
                    build_tool_call_log_entry(
                        step=step,
                        tool_name=block.name,
                        params=block.input,
                        result_str=result_str,
                        duration_ms=elapsed_ms,
                    )
                )

                # Record successful propose_patch into dedup map AFTER execution
                # succeeds (Codex review #2): if the first call failed
                # (policy block, parse error, transient DB error), we want a
                # corrected second call in the same turn to actually run, not
                # be silently skipped. _had_error covers `{"error": ...}` and
                # the tool-call helper's truthy error signal.
                if block.name == "workspace_propose_patch" and _dedup_patch_key and not _had_error:
                    try:
                        _parsed_result = json.loads(result_str) if isinstance(result_str, str) else {}
                    except (json.JSONDecodeError, TypeError):
                        _parsed_result = {}
                    _new_cs_id = _parsed_result.get("changeset_id") if isinstance(_parsed_result, dict) else None
                    if _new_cs_id:
                        patched_files[_dedup_patch_key] = _new_cs_id

            messages.append(adapter.build_tool_result_message(tool_results_content))

        else:
            # Loop exhausted — make one final call without tools to force text
            logger.warning(
                "Agentic loop exhausted %d steps, forcing final response",
                MAX_STEPS,
            )
            response = None
            async for event_type, payload in adapter.stream_message(
                model=model,
                max_tokens=8192,
                system=prompt_parts.static,
                system_dynamic=prompt_parts.dynamic,
                messages=messages,
            ):
                if event_type == "text":
                    if not suppress_streamed_text:
                        yield {"type": "text", "content": payload}
                elif event_type == "response":
                    response = payload

            if response:
                total_input_tokens += response.usage.input_tokens
                total_output_tokens += response.usage.output_tokens
                total_cache_creation_tokens += response.usage.cache_creation_input_tokens
                total_cache_read_tokens += response.usage.cache_read_input_tokens
                final_text = "\n".join(response.text_blocks) if response.text_blocks else ""

        # Strip raw tool reference tags / leaked XML the LLM may include
        final_text = _sanitize_assistant_text(final_text)

        # ── Save assistant message ──
        assistant_msg = ChatMessage(
            tenant_id=tenant_id,
            session_id=session.id,
            role="assistant",
            content=_coerce_assistant_content(
                final_text,
                last_structured_output,
                tool_calls=tool_calls_log,
            ),
            tool_calls=tool_calls_log if tool_calls_log else None,
            citations=citations if citations else None,
            token_count=total_input_tokens + total_output_tokens,
            input_tokens=total_input_tokens,
            output_tokens=total_output_tokens,
            cache_creation_tokens=total_cache_creation_tokens,
            cache_read_tokens=total_cache_read_tokens,
            model_used=model,
            provider_used=provider,
            is_byok=is_byok,
            query_importance=importance_tier.value,
            structured_output=last_structured_output,
            created_at=datetime.now(timezone.utc),
        )
        db.add(assistant_msg)

        # Auto-title from first message
        if not session.title:
            session.title = user_message[:100].strip()
        # Always bump updated_at so session re-sorts to top of list
        session.updated_at = func.now()

        # Audit
        audit_payload: dict[str, Any] = {
            "mode": "agentic",
            "provider": provider,
            "model": model,
            "steps": len(tool_calls_log),
            "input_tokens": total_input_tokens,
            "output_tokens": total_output_tokens,
            "doc_chunks_count": 0 if is_onboarding else (len(state.doc_chunks) if state.doc_chunks else 0),
            "tools_called": [t["tool"] for t in tool_calls_log],
        }
        if workspace_context:
            audit_payload["workspace_id"] = workspace_context["workspace_id"]
        await log_event(
            db=db,
            tenant_id=tenant_id,
            category="chat",
            action="chat.turn",
            actor_id=user_id,
            resource_type="chat_session",
            resource_id=str(session.id),
            correlation_id=correlation_id,
            payload=audit_payload,
        )

        # Tollbooth: deduct credits before commit (skip for BYOK users)
        if not is_byok:
            await deduct_chat_credits(db, tenant_id, model)

        await db.commit()

        # Auto-pin after a successful turn so follow-ups stick with the
        # data source the user was actually working with.
        try:
            _pin_update = _compute_source_pin_update(tool_calls_log)
            if _pin_update == "leave_pin":
                pass
            elif _pin_update is None:
                if getattr(session, "source_pin", None) is not None:
                    session.source_pin = None
                    await db.commit()
            elif getattr(session, "source_pin", None) != _pin_update:
                session.source_pin = _pin_update
                await db.commit()
        except Exception:
            logger.warning("auto_source_pin_update_failed", exc_info=True)

        # Fire-and-forget background tasks
        asyncio.create_task(
            _dispatch_memory_update(
                tenant_id=tenant_id,
                user_id=user_id,
                db=db,
                user_message=sanitized_input,
                assistant_message=final_text,
            )
        )
        asyncio.create_task(
            _dispatch_content_summary(
                db=db,
                message_id=assistant_msg.id,
                user_message=sanitized_input,
                assistant_message=final_text,
            )
        )

        result_msg = {
            "id": str(assistant_msg.id),
            "role": assistant_msg.role,
            "content": assistant_msg.content,
            "tool_calls": assistant_msg.tool_calls,
            "citations": assistant_msg.citations,
        }
        if hasattr(assistant_msg, "created_at") and assistant_msg.created_at:
            result_msg["created_at"] = assistant_msg.created_at.isoformat()
        result_msg["query_importance"] = importance_tier.value
        # Mirror `assistant_msg.structured_output` into the SSE payload so the
        # frontend can render clarification / write-confirmation cards
        # immediately on the terminal `message` event without waiting for a
        # session refetch.
        if assistant_msg.structured_output:
            result_msg["structured_output"] = assistant_msg.structured_output

        yield {"type": "message", "message": result_msg}
    except (Exception, asyncio.CancelledError):
        # Round 8 Bug 1: catch ``asyncio.CancelledError`` explicitly. On
        # Python 3.11+ ``CancelledError`` does NOT inherit from
        # ``Exception`` — when the outer
        # ``asyncio.wait_for(_run_chat_background, timeout=300)`` fires,
        # the bare ``except Exception`` was bypassed and the
        # clarification stayed at status='chosen' despite the failed
        # turn. We must catch both, run the revert, then re-raise so
        # cancellation still propagates to the asyncio loop. (Note: do
        # NOT use ``except BaseException`` — that would also swallow
        # ``SystemExit``/``KeyboardInterrupt``.)
        if _revert_message_id_on_failure is not None:
            try:
                # Clear any failed-transaction state on the session before
                # issuing the revert (mirrors the audit-failure rollback
                # added in round 7 Bug 2).
                await db.rollback()
            except Exception:
                logger.exception(
                    "[PLAN_MODE] db.rollback() failed before revert — "
                    "clarification %s may be stranded at status='chosen'",
                    _revert_message_id_on_failure,
                )
            try:
                from app.services.chat.plan_mode.short_circuit import (
                    revert_clarification_to_pending,
                )

                await revert_clarification_to_pending(
                    message_id=_revert_message_id_on_failure,
                    tenant_id=tenant_id,
                    db=db,
                )
            except Exception:
                logger.exception(
                    "[PLAN_MODE] revert_clarification_to_pending failed "
                    "after resume turn body raised — clarification %s "
                    "may be stranded at status='chosen'",
                    _revert_message_id_on_failure,
                )
        raise
