"""Template filler — detect currency columns, fill Excel templates, generate output."""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from decimal import Decimal

from openpyxl import Workbook

from app.schemas.pricing import PricingOutput

# ---------------------------------------------------------------------------
# Currency alias mapping
# ---------------------------------------------------------------------------
_DEFAULT_CURRENCY_ALIASES: dict[str, str] = {
    "USD": "USD", "US DOLLAR": "USD",
    "GBP": "GBP", "EUR": "EUR", "CAD": "CAD", "AUD": "AUD", "JPY": "JPY",
    "KRW": "KRW", "INR": "INR", "AED": "AED", "SEK": "SEK", "NOK": "NOK",
    "DKK": "DKK", "PLN": "PLN", "CZK": "CZK", "CHF": "CHF", "HUF": "HUF", "RON": "RON",
    "BRITISH POUND": "GBP", "POUND STERLING": "GBP", "STERLING": "GBP",
    "EURO": "EUR", "CANADIAN DOLLAR": "CAD", "AUSTRALIAN DOLLAR": "AUD",
    "JAPANESE YEN": "JPY", "YEN": "JPY", "KOREAN WON": "KRW", "WON": "KRW",
    "INDIAN RUPEE": "INR", "RUPEE": "INR", "UAE DIRHAM": "AED", "DIRHAM": "AED",
    "SWEDISH KRONA": "SEK", "NORWEGIAN KRONE": "NOK", "DANISH KRONE": "DKK",
    "POLISH ZLOTY": "PLN", "ZLOTY": "PLN", "CZECH KORUNA": "CZK", "KORUNA": "CZK",
    "SWISS FRANC": "CHF", "FRANC": "CHF", "HUNGARIAN FORINT": "HUF", "FORINT": "HUF",
    "ROMANIAN LEU": "RON", "LEU": "RON",
}

# ---------------------------------------------------------------------------
# Regex patterns
# ---------------------------------------------------------------------------
_SKU_PATTERNS = re.compile(
    r"(?i)^(sku|item(\s*id)?|external\s*id|product\s*(code|id)|part\s*(number|no))$"
)
_PRICE_PATTERNS = re.compile(
    r"(?i)(^usd$|usd\s*price|price\s*usd|base\s*price|unit\s*price)"
)
_USD_PATTERN = re.compile(
    r"(?i)(^usd$|^us\s*dollar$|usd\s*price|price\s*usd|base.*price|unit.*price)"
)


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------
@dataclass
class ColumnMapping:
    """Result of column detection."""

    sku_col: int  # 0-based index
    price_col: int  # 0-based index
    currency_cols: dict[int, str] = field(default_factory=dict)  # 0-based col → currency code


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _normalize_header(h: str) -> str:
    """Strip 'Price', 'Local', 'Base', parentheses, extra whitespace → uppercase."""
    if not h:
        return ""
    # Extract content from parentheses if present
    paren_match = re.search(r"\(([^)]+)\)", h)
    if paren_match:
        h = paren_match.group(1)
    # Remove common prefixes/suffixes
    h = re.sub(r"(?i)\b(price|local|base)\b", "", h)
    # Clean up whitespace
    h = h.strip()
    return h.upper()


def _match_currency(token: str, aliases: dict[str, str]) -> str | None:
    """Check token against aliases dict, return currency code or None."""
    if not token:
        return None
    upper = token.upper().strip()
    if upper in aliases:
        return aliases[upper]
    return None


# ---------------------------------------------------------------------------
# Main class
# ---------------------------------------------------------------------------

class TemplateFiller:
    """Reads Excel templates, detects currency columns, fills with pricing data."""

    def __init__(self, aliases: dict[str, str] | None = None):
        self._aliases = dict(_DEFAULT_CURRENCY_ALIASES)
        if aliases:
            # Add custom aliases (uppercase keys)
            for k, v in aliases.items():
                self._aliases[k.upper()] = v.upper()

    def detect_columns(self, wb: Workbook) -> ColumnMapping:
        """Detect SKU, price, and currency columns from row 1 headers."""
        ws = wb.active
        headers: list[str] = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            headers.append(str(val) if val is not None else "")

        sku_col = self._detect_sku_col(headers)
        price_col = self._detect_price_col(headers)
        currency_cols = self._detect_currency_cols(headers, price_col)

        return ColumnMapping(
            sku_col=sku_col,
            price_col=price_col,
            currency_cols=currency_cols,
        )

    def _detect_sku_col(self, headers: list[str]) -> int:
        """Find the SKU/Item ID column. Raises ValueError if not found."""
        for i, h in enumerate(headers):
            if h and _SKU_PATTERNS.match(h.strip()):
                return i
        raise ValueError("Could not detect SKU column. Expected: SKU, Item, Item ID, External ID, Product Code")

    def _detect_price_col(self, headers: list[str]) -> int:
        """Find the USD/base price column. Raises ValueError if not found."""
        # First pass: direct pattern match on raw header
        for i, h in enumerate(headers):
            if h and _PRICE_PATTERNS.search(h.strip()):
                return i
        # Second pass: match via USD pattern on raw header (e.g. "US Dollar")
        for i, h in enumerate(headers):
            if h and _USD_PATTERN.search(h.strip()):
                return i
        # Third pass: normalize and check for USD alias (e.g. "Base (USD)")
        for i, h in enumerate(headers):
            if not h:
                continue
            normalized = _normalize_header(h)
            currency = _match_currency(normalized, self._aliases)
            if currency == "USD":
                return i
        raise ValueError("Could not detect USD/price column. Expected: USD, USD Price, Base Price, Price USD")

    def _detect_currency_cols(self, headers: list[str], price_col: int) -> dict[int, str]:
        """Detect target currency columns (excludes USD/source)."""
        currency_cols: dict[int, str] = {}
        for i, h in enumerate(headers):
            if i == price_col:
                continue
            if not h:
                continue
            normalized = _normalize_header(h)
            if not normalized:
                continue
            currency = _match_currency(normalized, self._aliases)
            if currency and currency != "USD":
                currency_cols[i] = currency
        return currency_cols

    def fill(
        self,
        wb: Workbook,
        results: list[PricingOutput],
        mapping: ColumnMapping,
    ) -> None:
        """Fill currency cells in the workbook with pricing results."""
        ws = wb.active
        # Build SKU → PricingOutput lookup
        sku_lookup: dict[str, PricingOutput] = {r.sku: r for r in results}

        for row in range(2, ws.max_row + 1):
            sku_val = ws.cell(row=row, column=mapping.sku_col + 1).value
            if not sku_val:
                continue
            sku = str(sku_val).strip()
            output = sku_lookup.get(sku)
            if not output:
                continue

            for col_idx, currency_code in mapping.currency_cols.items():
                if currency_code in output.results:
                    final_price = output.results[currency_code].final_price
                    ws.cell(row=row, column=col_idx + 1, value=final_price)

    def generate_default_output(self, results: list[PricingOutput]) -> Workbook:
        """Create a 3-sheet workbook: Prices, Conversion Details, Config Snapshot."""
        wb = Workbook()

        # --- Sheet 1: Prices ---
        ws_prices = wb.active
        ws_prices.title = "Prices"

        # Collect all currencies across results
        all_currencies: list[str] = []
        seen: set[str] = set()
        for r in results:
            for code in r.results:
                if code not in seen:
                    all_currencies.append(code)
                    seen.add(code)
        all_currencies.sort()

        # Headers
        price_headers = ["SKU", "Item Name", "USD"] + all_currencies
        for ci, h in enumerate(price_headers, 1):
            ws_prices.cell(row=1, column=ci, value=h)

        # Data
        for ri, output in enumerate(results, 2):
            ws_prices.cell(row=ri, column=1, value=output.sku)
            ws_prices.cell(row=ri, column=2, value=output.item_name)
            ws_prices.cell(row=ri, column=3, value=float(output.usd_price))
            for ci, code in enumerate(all_currencies, 4):
                if code in output.results:
                    ws_prices.cell(row=ri, column=ci, value=float(output.results[code].final_price))

        # --- Sheet 2: Conversion Details ---
        ws_audit = wb.create_sheet("Conversion Details")
        audit_headers = [
            "SKU", "Currency", "FX Rate", "Tier", "Converted Amount",
            "VAT Rate", "VAT Amount", "Pre-Round Amount", "Final Price", "Rounding Rule",
        ]
        for ci, h in enumerate(audit_headers, 1):
            ws_audit.cell(row=1, column=ci, value=h)

        audit_row = 2
        for output in results:
            for code in sorted(output.results.keys()):
                cr = output.results[code]
                ws_audit.cell(row=audit_row, column=1, value=output.sku)
                ws_audit.cell(row=audit_row, column=2, value=cr.currency)
                ws_audit.cell(row=audit_row, column=3, value=float(cr.fx_rate))
                ws_audit.cell(row=audit_row, column=4, value=cr.tier)
                ws_audit.cell(row=audit_row, column=5, value=float(cr.converted_amount))
                ws_audit.cell(row=audit_row, column=6, value=float(cr.vat_rate) if cr.vat_rate else None)
                ws_audit.cell(row=audit_row, column=7, value=float(cr.vat_amount) if cr.vat_amount else None)
                ws_audit.cell(row=audit_row, column=8, value=float(cr.pre_round_amount))
                ws_audit.cell(row=audit_row, column=9, value=float(cr.final_price))
                ws_audit.cell(row=audit_row, column=10, value=cr.rounding_rule)
                audit_row += 1

        # --- Sheet 3: Config Snapshot ---
        ws_config = wb.create_sheet("Config Snapshot")
        ws_config.cell(row=1, column=1, value="Parameter")
        ws_config.cell(row=1, column=2, value="Value")
        ws_config.cell(row=2, column=1, value="Base Currency")
        ws_config.cell(row=2, column=2, value="USD")
        ws_config.cell(row=3, column=1, value="SKU Count")
        ws_config.cell(row=3, column=2, value=len(results))
        ws_config.cell(row=4, column=1, value="Target Currencies")
        ws_config.cell(row=4, column=2, value=", ".join(sorted(seen)))

        return wb

    def generate_netsuite_csv(self, results: list[PricingOutput]) -> str:
        """Generate CSV string in NetSuite import format."""
        output = io.StringIO(newline="")
        writer = csv.writer(output, lineterminator="\n")
        writer.writerow(["External ID", "Item", "Price Level", "Currency", "Rate"])

        for pricing_output in results:
            for code in sorted(pricing_output.results.keys()):
                cr = pricing_output.results[code]
                writer.writerow([
                    pricing_output.sku,
                    pricing_output.sku,
                    "Base Price",
                    cr.currency,
                    str(cr.final_price),
                ])

        return output.getvalue()
