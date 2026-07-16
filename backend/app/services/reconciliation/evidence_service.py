"""Generate evidence packs from reconciliation results.

Produces Excel workbooks with:
- Summary sheet: run metadata + match/exception/unmatched counts
- All Results sheet: full detail table
- Exceptions sheet: only needs_review items (bucket-authoritative)
"""

from __future__ import annotations

import collections
import io
from datetime import date, datetime, timezone
from decimal import Decimal
from typing import Any

import structlog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.reconciliation.four_bucket_classifier import (
    BUCKET_AUTO_CLASSIFICATIONS,
    BUCKET_MATCHES,
    BUCKET_NEEDS_REVIEW,
    BUCKET_RULES,
)

logger = structlog.get_logger()

_HEADER_FILL = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
_HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
_MATCHED_FILL = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
_EXCEPTION_FILL = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
_UNMATCHED_FILL = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
# Light blue: deterministic/rule-based suggested rows (auto_classifications + rules).
# Distinct from green (matches), yellow (needs_review/exceptions), red (unmatched).
_SUGGESTED_FILL = PatternFill(start_color="d1ecf1", end_color="d1ecf1", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class EvidencePackGenerator:
    """Generate evidence pack Excel files from reconciliation results."""

    def generate_excel(
        self,
        results: list[dict],
        run_id: str,
        date_from: date,
        date_to: date,
        tenant_name: str | None = None,
        proposals: list[dict] | None = None,
    ) -> io.BytesIO:
        """Generate the full evidence pack workbook."""
        wb = Workbook()

        # --- Summary sheet ---
        # needs_review = anything NOT in the three auto-approvable buckets:
        #   matches / auto_classifications / rules.
        # This folds None + unknown bucket values into needs_review, consistent
        # with the yellow row-fill else-branch and keeping the partition valid.
        needs_review_results = [
            r for r in results if r.get("bucket") not in (BUCKET_MATCHES, BUCKET_AUTO_CLASSIFICATIONS, BUCKET_RULES)
        ]
        self._write_summary(wb, results, needs_review_results, run_id, date_from, date_to, tenant_name)

        # --- All Results sheet ---
        self._write_results(wb, results, "All Results")

        # --- Exceptions sheet ---
        # Categorize on the authoritative four-bucket value, NOT advisory confidence.
        # needs_review already covers unmatched + material-variance auto_matched rows.
        self._write_results(wb, needs_review_results, "Exceptions")

        # --- Proposals sheet (summary-first recon rework, Phase 1) ---
        if proposals:
            self._write_proposals_sheet(wb, proposals)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def _write_proposals_sheet(self, wb: Workbook, proposals: list[dict]) -> None:
        ws = wb.create_sheet("Proposals")
        headers = [
            "Group",
            "Root Cause",
            "Action",
            "Vehicle",
            "Status",
            "Source",
            "Amount",
            "Currency",
            "Above Materiality",
            "Narrative",
            "Order Ref",
            "Stripe Charge",
            "NetSuite ID",
        ]
        ws.append(headers)
        for p in proposals:
            ws.append(
                [
                    p.get("group_key"),
                    p.get("root_cause"),
                    p.get("action"),
                    p.get("booking_vehicle"),
                    p.get("status"),
                    p.get("source"),
                    float(p.get("proposed_amount") or 0),
                    p.get("currency"),
                    "YES" if p.get("above_materiality") else "no",
                    p.get("narrative"),
                    p.get("order_reference"),
                    p.get("stripe_charge_id"),
                    p.get("netsuite_internal_id"),
                ]
            )

    def _write_summary(
        self,
        wb: Workbook,
        results: list[dict],
        needs_review_results: list[dict],
        run_id: str,
        date_from: date,
        date_to: date,
        tenant_name: str | None,
    ) -> None:
        """Write the summary sheet with run metadata and counts.

        Count by authoritative bucket — the four buckets PARTITION the run:
          matches + auto_classifications + rules + needs_review == total_results.
        "Unmatched" is a sub-detail of needs_review (Unmatched ⊆ Needs Review).
        """
        ws = wb.active
        ws.title = "Summary"

        # Count by authoritative bucket in a single pass over results.
        # needs_review is already pre-computed (any bucket not in the three
        # auto-approvable buckets — folds None/unknown into needs_review).
        bucket_counts = collections.Counter(r.get("bucket") for r in results)
        matched = bucket_counts[BUCKET_MATCHES]
        auto_classified = bucket_counts[BUCKET_AUTO_CLASSIFICATIONS]
        rules_count = bucket_counts[BUCKET_RULES]
        needs_review = len(needs_review_results)
        unmatched = len([r for r in results if r.get("match_type") == "unmatched"])
        total_variance = sum(Decimal(str(r.get("variance_amount", 0))) for r in results)

        title_font = Font(name="Arial", size=14, bold=True)
        label_font = Font(name="Arial", size=11, bold=True)
        value_font = Font(name="Arial", size=11)

        row = 1
        ws.cell(row=row, column=1, value="Reconciliation Evidence Pack").font = title_font
        row += 1
        if tenant_name:
            ws.cell(row=row, column=1, value=f"Tenant: {tenant_name}").font = value_font
            row += 1

        row += 1
        summary_data = [
            ("Run ID", run_id),
            ("Period", f"{date_from.isoformat()} to {date_to.isoformat()}"),
            ("Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
            ("", ""),
            ("Auto-Matched", matched),
            ("Auto-Classified", auto_classified),
            ("Rules (Fuzzy)", rules_count),
            ("Needs Review (Exceptions)", needs_review),
            # Unmatched ⊆ Needs Review — sub-detail kept for drill-down clarity.
            ("Unmatched (within Needs Review)", unmatched),
            ("Total Results", len(results)),
            ("Total Variance", f"${total_variance:,.2f}"),
        ]

        for label, value in summary_data:
            ws.cell(row=row, column=1, value=label).font = label_font
            # Render integer 0 counts as "0" (not blank); spacer "" rows stay blank.
            ws.cell(row=row, column=2, value=str(value) if value is not None else "").font = value_font
            row += 1

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 45

    def _write_results(self, wb: Workbook, results: list[dict], sheet_name: str) -> None:
        """Write a results table sheet."""
        ws = wb.create_sheet(title=sheet_name)

        headers = [
            "Match Type",
            # Advisory amount+temporal composite (R2) — display only, never a verdict.
            # Categorization (Exceptions sheet, summary counts, row fills) is keyed on
            # the authoritative four-bucket `bucket`, NOT this value.
            "Advisory Match Score",
            "Status",
            "Stripe Amount",
            "NetSuite Amount",
            "Variance",
            "Variance Type",
            "Explanation",
            "Currency",
            "Match Rule",
            "Payout ID",
            "Deposit ID(s)",
        ]

        # Header row
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, result in enumerate(results, 2):
            evidence = result.get("evidence", {}) or {}
            row_data: list[Any] = [
                result.get("match_type", ""),
                float(result.get("confidence", 0)),
                result.get("status", ""),
                float(result["stripe_amount"]) if result.get("stripe_amount") is not None else "",
                float(result["netsuite_amount"]) if result.get("netsuite_amount") is not None else "",
                float(result.get("variance_amount", 0)),
                result.get("variance_type", ""),
                result.get("variance_explanation", ""),
                result.get("currency", ""),
                result.get("match_rule", ""),
                evidence.get("payout_source_id", ""),
                ", ".join(evidence.get("deposit_ids", [])),
            ]

            # Row fill: clean 1:1 mapping from the authoritative four buckets.
            # "Unmatched" is treated as a special case on match_type (before bucket check)
            # because it may sit inside needs_review and we want the red visual signal.
            match_type = result.get("match_type", "")
            bucket = result.get("bucket", "")
            if match_type == "unmatched":
                fill = _UNMATCHED_FILL  # red
            elif bucket == BUCKET_MATCHES:
                fill = _MATCHED_FILL  # green
            elif bucket == BUCKET_NEEDS_REVIEW:
                fill = _EXCEPTION_FILL  # yellow
            elif bucket in (BUCKET_AUTO_CLASSIFICATIONS, BUCKET_RULES):
                fill = _SUGGESTED_FILL  # light blue — matched, not review-required
            else:
                # None/unknown bucket — conservative fallback to yellow
                fill = _EXCEPTION_FILL

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = fill
                cell.border = _THIN_BORDER

        # Auto-width columns
        for col_idx in range(1, len(headers) + 1):
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = max(15, len(headers[col_idx - 1]) + 4)

        # Wider column for explanations
        ws.column_dimensions[get_column_letter(8)].width = 50
