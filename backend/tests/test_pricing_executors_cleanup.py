"""Step 4 — pricing_convert / pricing_export cleanup + pricing_state emit.

Two changes verified here:

1. The legacy ``response_instruction`` markdown table is removed (CLAUDE.md
   Mistakes #41 / feedback_no_llm_numbers.md). The frontend renders the
   pricing preview from the existing task_output card; the LLM must NOT be
   instructed to print a row-by-row markdown table.

2. Both executors emit ``pricing_state`` in their result dict so the
   orchestrator interceptor (Step 5) can persist it as the cache payload for
   pricing_revise / pricing_to_sheets follow-ups.
"""

from __future__ import annotations

import io
import uuid
from decimal import Decimal
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from openpyxl import Workbook

from app.mcp.tools.pricing_tools import (
    pricing_convert_execute,
    pricing_export_execute,
)
from app.schemas.pricing import CurrencyConfig, TenantPricingConfig


def _three_currency_config_dict() -> dict:
    config = TenantPricingConfig(
        base_currency="USD",
        eur_fx_rate=Decimal("0.92"),
        currencies={
            "GBP": CurrencyConfig(
                fx_rate=Decimal("0.79"),
                tier="usd_based",
                vat_rate=Decimal("0.20"),
                rounding_rule="nearest_9",
            ),
            "EUR": CurrencyConfig(
                fx_rate=Decimal("0.92"),
                tier="eur_based",
                vat_rate=Decimal("0.23"),
                rounding_rule="nearest_9",
            ),
            "CAD": CurrencyConfig(
                fx_rate=Decimal("1.36"),
                tier="usd_based",
                vat_rate=None,
                rounding_rule="nearest_9",
            ),
        },
    )
    return config.model_dump(mode="json")


def _build_excel_bytes(rows: list[tuple[str, float]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "SKU"
    ws["B1"] = "USD Price"
    for i, (sku, price) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=sku)
        ws.cell(row=i, column=2, value=price)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def pricing_context():
    return {"db": AsyncMock(), "tenant_id": uuid.uuid4(), "user_id": uuid.uuid4()}


def _patch_save_output(saved_files: list):
    """Helper — patch _file_svc.save_output to record saves and return mock files."""

    async def _save_output(*, db, tenant_id, user_id, filename, content, related_message_id=None):
        file_obj = MagicMock()
        file_obj.id = uuid.uuid4()
        file_obj.filename = filename
        saved_files.append(file_obj)
        return file_obj

    return _save_output


class TestPricingConvertCleanup:
    """pricing_convert_execute — drops markdown table, emits pricing_state."""

    @pytest.mark.asyncio
    async def test_no_markdown_table_in_result(self, pricing_context):
        config_row = MagicMock()
        config_row.config = _three_currency_config_dict()
        task_file = MagicMock()
        task_file.filename = "input.xlsx"
        excel_bytes = _build_excel_bytes([("SKU-1", 99.00), ("SKU-2", 199.00)])

        saved_files: list = []

        with (
            patch(
                "app.mcp.tools.pricing_tools.get_config",
                new_callable=AsyncMock,
                return_value=config_row,
            ),
            patch(
                "app.mcp.tools.pricing_tools._file_svc.get_file",
                new_callable=AsyncMock,
                return_value=(task_file, excel_bytes),
            ),
            patch(
                "app.mcp.tools.pricing_tools._file_svc.save_output",
                new=_patch_save_output(saved_files),
            ),
        ):
            result = await pricing_convert_execute({"file_id": str(uuid.uuid4())}, pricing_context)

        assert result["success"] is True
        # Markdown table absence — no separator row, no "show this EXACT" text.
        ri = result.get("response_instruction", "")
        assert "|---" not in ri, "response_instruction must not contain a markdown table"
        assert "EXACT table" not in ri, "instructions must not tell the LLM to print prices verbatim"
        assert "verbatim" not in ri.lower()

    @pytest.mark.asyncio
    async def test_emits_pricing_state(self, pricing_context):
        config_row = MagicMock()
        config_row.config = _three_currency_config_dict()
        task_file = MagicMock()
        task_file.filename = "input.xlsx"
        excel_bytes = _build_excel_bytes([("SKU-1", 99.00), ("SKU-2", 199.00)])

        saved_files: list = []

        with (
            patch(
                "app.mcp.tools.pricing_tools.get_config",
                new_callable=AsyncMock,
                return_value=config_row,
            ),
            patch(
                "app.mcp.tools.pricing_tools._file_svc.get_file",
                new_callable=AsyncMock,
                return_value=(task_file, excel_bytes),
            ),
            patch(
                "app.mcp.tools.pricing_tools._file_svc.save_output",
                new=_patch_save_output(saved_files),
            ),
        ):
            result = await pricing_convert_execute({"file_id": str(uuid.uuid4())}, pricing_context)

        assert "pricing_state" in result
        ps = result["pricing_state"]
        # Seed inputs preserved
        assert len(ps["seed_items"]) == 2
        assert ps["seed_items"][0]["sku"] == "SKU-1"
        # Effective state seeded from inputs
        assert ps["effective_items"] == ps["seed_items"]
        # Effective currencies match the tenant config order (sorted for stability)
        assert set(ps["effective_currencies"]) == {"GBP", "EUR", "CAD"}
        # Empty override dicts on the seed run
        assert ps["effective_fx_overrides"] == {}
        assert ps["effective_vat_overrides"] == {}
        assert ps["effective_rounding_overrides"] == {}
        assert ps["effective_uplift_by_currency"] == {}
        assert ps["applied_overrides_log"] == []
        # File pointers
        assert ps["excel_file_id"] == str(saved_files[0].id)
        assert ps["netsuite_csv_file_id"] == str(saved_files[1].id)
        assert ps["row_count"] == 2
        # Header columns include SKU + USD + currencies
        assert "SKU" in ps["header_columns"]
        assert "USD" in ps["header_columns"]
        assert "GBP" in ps["header_columns"]

    @pytest.mark.asyncio
    async def test_preview_field_unchanged(self, pricing_context):
        """The frontend card relies on `preview` shape — list of dicts with
        SKU/USD/currency keys. That contract must NOT change."""
        config_row = MagicMock()
        config_row.config = _three_currency_config_dict()
        task_file = MagicMock()
        task_file.filename = "input.xlsx"
        excel_bytes = _build_excel_bytes([("SKU-1", 99.00)])

        saved_files: list = []

        with (
            patch(
                "app.mcp.tools.pricing_tools.get_config",
                new_callable=AsyncMock,
                return_value=config_row,
            ),
            patch(
                "app.mcp.tools.pricing_tools._file_svc.get_file",
                new_callable=AsyncMock,
                return_value=(task_file, excel_bytes),
            ),
            patch(
                "app.mcp.tools.pricing_tools._file_svc.save_output",
                new=_patch_save_output(saved_files),
            ),
        ):
            result = await pricing_convert_execute({"file_id": str(uuid.uuid4())}, pricing_context)

        assert isinstance(result["preview"], list)
        assert len(result["preview"]) == 1
        row = result["preview"][0]
        assert "SKU" in row and row["SKU"] == "SKU-1"
        assert "USD" in row
        assert "GBP" in row


class TestPricingExportCleanup:
    """pricing_export_execute — drops markdown table, emits pricing_state."""

    @pytest.mark.asyncio
    async def test_no_markdown_table_in_result(self, pricing_context):
        config_row = MagicMock()
        config_row.config = _three_currency_config_dict()
        saved_files: list = []

        items = [
            {"sku": "SKU-1", "usd_price": 99.0},
            {"sku": "SKU-2", "usd_price": 199.0, "item_name": "Gadget"},
        ]

        with (
            patch(
                "app.mcp.tools.pricing_tools.get_config",
                new_callable=AsyncMock,
                return_value=config_row,
            ),
            patch(
                "app.mcp.tools.pricing_tools._file_svc.save_output",
                new=_patch_save_output(saved_files),
            ),
        ):
            result = await pricing_export_execute({"items": items}, pricing_context)

        assert result["success"] is True
        ri = result.get("response_instruction", "")
        assert "|---" not in ri
        assert "EXACT table" not in ri
        assert "verbatim" not in ri.lower()

    @pytest.mark.asyncio
    async def test_emits_pricing_state(self, pricing_context):
        config_row = MagicMock()
        config_row.config = _three_currency_config_dict()
        saved_files: list = []

        items = [{"sku": "SKU-1", "usd_price": 99.0}]

        with (
            patch(
                "app.mcp.tools.pricing_tools.get_config",
                new_callable=AsyncMock,
                return_value=config_row,
            ),
            patch(
                "app.mcp.tools.pricing_tools._file_svc.save_output",
                new=_patch_save_output(saved_files),
            ),
        ):
            result = await pricing_export_execute({"items": items}, pricing_context)

        assert "pricing_state" in result
        ps = result["pricing_state"]
        assert len(ps["seed_items"]) == 1
        assert ps["seed_items"][0]["sku"] == "SKU-1"
        assert ps["effective_items"] == ps["seed_items"]
        assert ps["row_count"] == 1
        assert ps["excel_file_id"] == str(saved_files[0].id)
        assert ps["netsuite_csv_file_id"] == str(saved_files[1].id)
