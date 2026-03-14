"""Tests for excel_export_service."""

import io

import openpyxl
import pytest

from app.services.excel_export_service import (
    ExcelExportConfig,
    generate_excel,
    _humanize_header,
    _detect_column_types,
    _to_number,
)


class TestHumanizeHeader:
    def test_snake_case(self):
        assert _humanize_header("order_count") == "Order Count"

    def test_camel_case(self):
        assert _humanize_header("orderCount") == "Order Count"

    def test_already_spaced(self):
        assert _humanize_header("Total Amount") == "Total Amount"

    def test_single_word(self):
        assert _humanize_header("revenue") == "Revenue"


class TestToNumber:
    def test_int(self):
        assert _to_number(42) == 42

    def test_float(self):
        assert _to_number(3.14) == 3.14

    def test_string_int(self):
        assert _to_number("1000") == 1000

    def test_string_float(self):
        assert _to_number("1,234.56") == 1234.56

    def test_dollar_sign(self):
        assert _to_number("$500") == 500

    def test_accounting_parens(self):
        assert _to_number("(123.45)") == -123.45

    def test_non_numeric(self):
        assert _to_number("hello") is None

    def test_none(self):
        assert _to_number(None) is None


class TestDetectColumnTypes:
    def test_currency_by_name(self):
        types = _detect_column_types(["amount", "name"], [[100, "Alice"]])
        assert types["amount"] == "currency"
        assert types["name"] == "text"

    def test_id_stays_text(self):
        types = _detect_column_types(["internalid", "tranid"], [[1001, "SO123"]])
        assert types["internalid"] == "text"
        assert types["tranid"] == "text"

    def test_percent_by_name(self):
        types = _detect_column_types(["margin_rate"], [[0.15]])
        assert types["margin_rate"] == "percent"

    def test_date_by_name(self):
        types = _detect_column_types(["trandate"], [["2026-01-15"]])
        assert types["trandate"] == "date"

    def test_numeric_by_data_sampling(self):
        rows = [[i] for i in range(20)]
        types = _detect_column_types(["custom_col"], rows)
        assert types["custom_col"] == "number"

    def test_text_by_data_sampling(self):
        rows = [["foo"], ["bar"], ["baz"]]
        types = _detect_column_types(["notes"], rows)
        assert types["notes"] == "text"


class TestGenerateExcel:
    def _load(self, buf: io.BytesIO) -> openpyxl.Workbook:
        buf.seek(0)
        return openpyxl.load_workbook(buf)

    def test_basic_export(self):
        buf = generate_excel(
            columns=["name", "amount"],
            rows=[["Alice", 100], ["Bob", 200]],
        )
        wb = self._load(buf)
        ws = wb.active
        assert ws is not None
        assert ws.max_row >= 6

    def test_currency_detection(self):
        buf = generate_excel(
            columns=["total_amount"],
            rows=[[1234.56]],
        )
        wb = self._load(buf)
        ws = wb.active
        data_cell = None
        for row in ws.iter_rows(min_row=1, max_col=1):
            for cell in row:
                if cell.value == 1234.56:
                    data_cell = cell
                    break
        assert data_cell is not None
        assert "#,##0.00" in (data_cell.number_format or "")

    def test_id_stays_text(self):
        buf = generate_excel(
            columns=["internalid"],
            rows=[[4000]],
        )
        wb = self._load(buf)
        ws = wb.active
        data_cell = None
        for row in ws.iter_rows(min_row=1, max_col=1):
            for cell in row:
                if str(cell.value) == "4000":
                    data_cell = cell
                    break
        assert data_cell is not None
        assert isinstance(data_cell.value, str)

    def test_negative_numbers(self):
        buf = generate_excel(
            columns=["amount"],
            rows=[[-1234.56]],
        )
        wb = self._load(buf)
        ws = wb.active
        data_cell = None
        for row in ws.iter_rows(min_row=1, max_col=1):
            for cell in row:
                if cell.value == -1234.56:
                    data_cell = cell
                    break
        assert data_cell is not None
        assert "(#,##0.00)" in (data_cell.number_format or "")

    def test_null_handling(self):
        buf = generate_excel(
            columns=["name"],
            rows=[[None]],
        )
        wb = self._load(buf)
        ws = wb.active
        # Find the data row (after title, metadata, header rows)
        # Null values should NOT appear as the string "None"
        for row in ws.iter_rows(min_row=1, max_col=1):
            for cell in row:
                assert cell.value != "None", "Null should not be written as string 'None'"
        # The data cell should be empty (openpyxl round-trips "" as None)
        header_row = None
        for row in ws.iter_rows(min_row=1, max_col=1):
            for cell in row:
                if cell.value == "Name":
                    header_row = cell.row
        assert header_row is not None
        data_cell = ws.cell(row=header_row + 1, column=1)
        assert data_cell.value is None  # openpyxl converts "" to None on save/load

    def test_empty_rows(self):
        buf = generate_excel(columns=["a", "b"], rows=[])
        wb = self._load(buf)
        assert wb.active is not None

    def test_large_dataset(self):
        cols = ["id", "name", "amount", "rate"]
        rows = [[i, f"Item {i}", i * 10.5, 0.15] for i in range(10_000)]
        buf = generate_excel(columns=cols, rows=rows)
        wb = self._load(buf)
        assert wb.active is not None

    def test_column_width_clamping(self):
        buf = generate_excel(
            columns=["x"],
            rows=[["a"]],
        )
        wb = self._load(buf)
        ws = wb.active
        width = ws.column_dimensions["A"].width
        assert width >= 8
        assert width <= 45

    def test_metadata_block(self):
        buf = generate_excel(
            columns=["a"],
            rows=[[1]],
            title="My Report",
            metadata={"Period": "Jan 2026", "Account": "Framework"},
        )
        wb = self._load(buf)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "My Report"
        cell_values = [ws.cell(row=r, column=1).value for r in range(2, 10)]
        assert "Period:" in cell_values
        assert "Account:" in cell_values

    def test_column_type_override(self):
        buf = generate_excel(
            columns=["ref"],
            rows=[[500.0]],
            column_types={"ref": "currency"},
        )
        wb = self._load(buf)
        ws = wb.active
        data_cell = None
        for row in ws.iter_rows(min_row=1, max_col=1):
            for cell in row:
                if cell.value == 500.0:
                    data_cell = cell
                    break
        assert data_cell is not None
        assert "#,##0.00" in (data_cell.number_format or "")

    def test_percent_detection(self):
        buf = generate_excel(
            columns=["margin_rate"],
            rows=[[0.25]],
        )
        wb = self._load(buf)
        ws = wb.active
        data_cell = None
        for row in ws.iter_rows(min_row=1, max_col=1):
            for cell in row:
                if cell.value == 0.25:
                    data_cell = cell
                    break
        assert data_cell is not None
        assert "%" in (data_cell.number_format or "")

    def test_stripe_pattern(self):
        buf = generate_excel(
            columns=["a"],
            rows=[[1], [2], [3], [4]],
        )
        wb = self._load(buf)
        ws = wb.active
        data_rows = []
        for row in ws.iter_rows(min_row=1, max_col=1):
            for cell in row:
                if cell.value in (1, 2, 3, 4):
                    data_rows.append(cell)
        assert len(data_rows) == 4
        stripe_cell = data_rows[1]
        assert stripe_cell.fill.fgColor.rgb is not None
        assert "F8F9FA" in (stripe_cell.fill.fgColor.rgb or "").upper()
