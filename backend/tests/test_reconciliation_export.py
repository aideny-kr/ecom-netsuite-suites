"""Per-section CSV/Excel export endpoint (GET .../runs/{run_id}/export).

Follows the calling convention already used by test_resolution_summary_api.py /
test_resolution_flag_and_evidence.py: call the FastAPI endpoint function
directly against seeded DB fixtures, and for streamed bytes consume
``response.body_iterator``.
"""

import csv
import io
from decimal import Decimal

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import select

from app.api.v1.reconciliation import (
    export_run_section,
    plan_resolutions,
    reject_resolution_group,
)
from app.models.reconciliation import ReconResolutionProposal
from tests.conftest import (
    create_test_netsuite_posting,
    create_test_recon_result,
    create_test_recon_run,
    create_test_user,
    enable_feature_flag,
)

_GROUPS_HEADERS = [
    "group_key",
    "root_cause",
    "action",
    "booking_vehicle",
    "currency",
    "count",
    "proposed_count",
    "approved_count",
    "above_materiality_count",
    "total_amount",
]

_PROPOSALS_HEADERS = [
    "order_reference",
    "stripe_charge_id",
    "netsuite_internal_id",
    "netsuite_record_type",
    "stripe_amount",
    "netsuite_amount",
    "variance_amount",
    "proposed_amount",
    "currency",
    "status",
    "above_materiality",
    "root_cause",
    "action",
    "booking_vehicle",
    "narrative",
]

_PROPOSALS_XLSX_EXTRA_HEADERS = ["proposal_id", "run_id", "source", "decided_by", "decided_at", "created_at"]

_RESULTS_HEADERS = [
    "match_type",
    "confidence",
    "status",
    "bucket",
    "stripe_amount",
    "netsuite_amount",
    "variance_amount",
    "variance_type",
    "variance_explanation",
    "currency",
    "match_rule",
]


async def _body_bytes(response) -> bytes:
    return b"".join([chunk async for chunk in response.body_iterator])


def _csv_rows(body: bytes) -> list[list[str]]:
    return list(csv.reader(io.StringIO(body.decode("utf-8"))))


async def _seed_run_with_groups(db, tenant):
    """fees (matched, above-materiality-eligible amounts), timing, and a
    chargeback (-> needs_human) group, plus a clean match with no proposal —
    mirrors the fixture shape in test_resolution_summary_api._seed."""
    user, _ = await create_test_user(db, tenant)
    await enable_feature_flag(db, tenant.id, "recon_resolution_ui")
    run = await create_test_recon_run(db, tenant.id, status="completed")
    posting = await create_test_netsuite_posting(db, tenant.id, netsuite_internal_id="98765", record_type="custdep")
    await create_test_recon_result(
        db,
        tenant.id,
        run.id,
        status="pending",
        bucket="auto_classifications",
        match_type="deterministic",
        variance_type="fees",
        variance_amount=Decimal("9.00"),
        stripe_amount=Decimal("1000.00"),
        netsuite_amount=Decimal("991.00"),
        evidence={"charge_source_id": "ch_fee1", "order_reference": "R1"},
        deposit_id=posting.id,
    )
    await create_test_recon_result(
        db,
        tenant.id,
        run.id,
        status="pending",
        bucket="rules",
        match_type="fuzzy",
        variance_type="timing",
        variance_amount=Decimal("0"),
        stripe_amount=Decimal("50.00"),
        netsuite_amount=Decimal("50.00"),
        evidence={"charge_source_id": "ch_t", "order_reference": "R2"},
    )
    await create_test_recon_result(
        db,
        tenant.id,
        run.id,
        status="pending",
        bucket="needs_review",
        match_type="deterministic",
        variance_type="chargeback",
        variance_amount=Decimal("42.00"),
        stripe_amount=Decimal("42.00"),
        netsuite_amount=Decimal("0"),
        evidence={"charge_source_id": "ch_c", "order_reference": "R3"},
    )
    await create_test_recon_result(
        db,
        tenant.id,
        run.id,
        status="auto_matched",
        bucket="matches",
        match_type="deterministic",
        variance_type=None,
        variance_amount=Decimal("0"),
        stripe_amount=Decimal("10.00"),
        netsuite_amount=Decimal("10.00"),
        evidence={"charge_source_id": "ch_m", "order_reference": "R4"},
    )
    run.matches_count = 1
    await db.flush()
    await plan_resolutions(str(run.id), user=user, db=db)
    return user, run


async def _seed_multi_currency_fees(db, tenant):
    user, _ = await create_test_user(db, tenant)
    await enable_feature_flag(db, tenant.id, "recon_resolution_ui")
    run = await create_test_recon_run(db, tenant.id, status="completed")
    await create_test_recon_result(
        db,
        tenant.id,
        run.id,
        status="pending",
        bucket="auto_classifications",
        match_type="deterministic",
        variance_type="fees",
        variance_amount=Decimal("5.00"),
        stripe_amount=Decimal("500.00"),
        netsuite_amount=Decimal("495.00"),
        currency="USD",
        evidence={"charge_source_id": "ch_usd", "order_reference": "R1"},
    )
    await create_test_recon_result(
        db,
        tenant.id,
        run.id,
        status="pending",
        bucket="auto_classifications",
        match_type="deterministic",
        variance_type="fees",
        variance_amount=Decimal("7.00"),
        stripe_amount=Decimal("700.00"),
        netsuite_amount=Decimal("693.00"),
        currency="EUR",
        evidence={"charge_source_id": "ch_eur", "order_reference": "R2"},
    )
    await db.flush()
    await plan_resolutions(str(run.id), user=user, db=db)
    return user, run


async def _seed_needs_human_cross_group(db, tenant):
    user, _ = await create_test_user(db, tenant)
    await enable_feature_flag(db, tenant.id, "recon_resolution_ui")
    run = await create_test_recon_run(db, tenant.id, status="completed")
    # chargeback -> needs_human (planner rule 3).
    await create_test_recon_result(
        db,
        tenant.id,
        run.id,
        status="pending",
        bucket="needs_review",
        match_type="deterministic",
        variance_type="chargeback",
        variance_amount=Decimal("42"),
        stripe_amount=Decimal("42"),
        netsuite_amount=Decimal("0"),
        evidence={"charge_source_id": "ch_c", "order_reference": "R3"},
    )
    # manual_adjustment -> needs_human (planner rule 10, catch-all) — a
    # DIFFERENT root_cause/group_key, proving the action filter spans groups.
    await create_test_recon_result(
        db,
        tenant.id,
        run.id,
        status="pending",
        bucket="needs_review",
        match_type="exception",
        variance_type="manual_adjustment",
        variance_amount=Decimal("15"),
        stripe_amount=Decimal("115"),
        netsuite_amount=Decimal("100"),
        evidence={"charge_source_id": "ch_m", "order_reference": "R5"},
    )
    run.matches_count = 0
    await db.flush()
    await plan_resolutions(str(run.id), user=user, db=db)
    return user, run


# ---------------------------------------------------------------------------
# section=groups
# ---------------------------------------------------------------------------
async def test_export_groups_csv_header_and_row(db, tenant_a):
    user, run = await _seed_run_with_groups(db, tenant_a)
    response = await export_run_section(str(run.id), user=user, db=db, section="groups", format="csv")
    rows = _csv_rows(await _body_bytes(response))
    assert rows[0] == _GROUPS_HEADERS
    by_group_key = {r[0]: r for r in rows[1:]}
    fees_row = by_group_key["fees:book_fee_line:deposit"]
    assert fees_row == [
        "fees:book_fee_line:deposit",
        "fees",
        "book_fee_line",
        "deposit",
        "USD",
        "1",
        "1",
        "0",
        "0",
        "9.00",
    ]


async def test_export_groups_xlsx_loads_with_headers(db, tenant_a):
    user, run = await _seed_run_with_groups(db, tenant_a)
    response = await export_run_section(str(run.id), user=user, db=db, section="groups", format="xlsx")
    wb = load_workbook(io.BytesIO(await _body_bytes(response)))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers == _GROUPS_HEADERS
    assert ws.max_row == 4  # header + fees + timing + chargeback (needs_human)


# ---------------------------------------------------------------------------
# section=proposals
# ---------------------------------------------------------------------------
async def test_export_proposals_csv_row_has_exact_decimal_strings(db, tenant_a):
    user, run = await _seed_run_with_groups(db, tenant_a)
    response = await export_run_section(
        str(run.id),
        user=user,
        db=db,
        section="proposals",
        format="csv",
        group_key="fees:book_fee_line:deposit",
    )
    rows = _csv_rows(await _body_bytes(response))
    assert rows[0] == _PROPOSALS_HEADERS
    assert len(rows) == 2
    row = dict(zip(_PROPOSALS_HEADERS, rows[1]))
    assert row["order_reference"] == "R1"
    assert row["stripe_charge_id"] == "ch_fee1"
    assert row["netsuite_internal_id"] == "98765"
    assert row["netsuite_record_type"] == "custdep"
    # exact Decimal strings — never float notation ("991.0")
    assert row["stripe_amount"] == "1000.00"
    assert row["netsuite_amount"] == "991.00"
    assert row["variance_amount"] == "9.00"
    assert row["proposed_amount"] == "9.00"
    assert row["above_materiality"] == "False"


async def test_export_proposals_xlsx_has_extra_columns(db, tenant_a):
    user, run = await _seed_run_with_groups(db, tenant_a)
    response = await export_run_section(
        str(run.id),
        user=user,
        db=db,
        section="proposals",
        format="xlsx",
        group_key="fees:book_fee_line:deposit",
    )
    wb = load_workbook(io.BytesIO(await _body_bytes(response)))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers == _PROPOSALS_HEADERS + _PROPOSALS_XLSX_EXTRA_HEADERS
    row = dict(zip(headers, [c.value for c in ws[2]]))
    assert row["source"] == "planner"
    assert row["run_id"] == str(run.id)
    assert row["decided_by"] is None
    assert row["created_at"] is not None


async def test_export_proposals_group_key_filter_narrows_rows(db, tenant_a):
    user, run = await _seed_run_with_groups(db, tenant_a)
    response = await export_run_section(
        str(run.id),
        user=user,
        db=db,
        section="proposals",
        format="csv",
        group_key="timing:carry_forward:none",
    )
    rows = _csv_rows(await _body_bytes(response))
    assert len(rows) == 2  # header + one timing proposal
    row = dict(zip(_PROPOSALS_HEADERS, rows[1]))
    assert row["root_cause"] == "timing"
    assert row["action"] == "carry_forward"


async def test_export_proposals_currency_filter_narrows_rows(db, tenant_a):
    user, run = await _seed_multi_currency_fees(db, tenant_a)
    response = await export_run_section(
        str(run.id), user=user, db=db, section="proposals", format="csv", currency="EUR"
    )
    rows = _csv_rows(await _body_bytes(response))
    assert len(rows) == 2
    row = dict(zip(_PROPOSALS_HEADERS, rows[1]))
    assert row["currency"] == "EUR"
    assert row["order_reference"] == "R2"


async def test_export_proposals_action_filter_spans_groups(db, tenant_a):
    user, run = await _seed_needs_human_cross_group(db, tenant_a)
    response = await export_run_section(
        str(run.id), user=user, db=db, section="proposals", format="csv", action="needs_human"
    )
    rows = _csv_rows(await _body_bytes(response))
    assert len(rows) == 3  # header + 2 needs_human proposals across 2 root_causes
    root_causes = {dict(zip(_PROPOSALS_HEADERS, r))["root_cause"] for r in rows[1:]}
    assert root_causes == {"chargeback", "manual_adjustment"}


async def test_export_proposals_excludes_superseded_and_rejected(db, tenant_a):
    user, run = await _seed_run_with_groups(db, tenant_a)
    await reject_resolution_group(str(run.id), "fees:book_fee_line:deposit", user=user, db=db)

    response = await export_run_section(
        str(run.id),
        user=user,
        db=db,
        section="proposals",
        format="csv",
        group_key="fees:book_fee_line:deposit",
    )
    rows = _csv_rows(await _body_bytes(response))
    assert len(rows) == 1  # header only — the rejected proposal is excluded


# ---------------------------------------------------------------------------
# section=results
# ---------------------------------------------------------------------------
async def test_export_results_csv_header_and_row(db, tenant_a):
    user, run = await _seed_run_with_groups(db, tenant_a)
    response = await export_run_section(str(run.id), user=user, db=db, section="results", format="csv")
    rows = _csv_rows(await _body_bytes(response))
    assert rows[0] == _RESULTS_HEADERS
    assert len(rows) == 5  # header + 4 seeded results
    fee_row = next(dict(zip(_RESULTS_HEADERS, r)) for r in rows[1:] if r[7] == "fees")
    assert fee_row["match_type"] == "deterministic"
    assert fee_row["bucket"] == "auto_classifications"
    assert fee_row["stripe_amount"] == "1000.00"
    assert fee_row["netsuite_amount"] == "991.00"
    assert fee_row["variance_amount"] == "9.00"
    assert fee_row["currency"] == "USD"
    assert fee_row["match_rule"] == "order_reference_exact"


async def test_export_results_xlsx_loads(db, tenant_a):
    user, run = await _seed_run_with_groups(db, tenant_a)
    response = await export_run_section(str(run.id), user=user, db=db, section="results", format="xlsx")
    wb = load_workbook(io.BytesIO(await _body_bytes(response)))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers == _RESULTS_HEADERS
    assert ws.max_row == 5


# ---------------------------------------------------------------------------
# Validation + tenant scoping
# ---------------------------------------------------------------------------
async def test_export_bad_section_400(db, tenant_a):
    user, run = await _seed_run_with_groups(db, tenant_a)
    with pytest.raises(HTTPException) as exc:
        await export_run_section(str(run.id), user=user, db=db, section="bogus", format="csv")
    assert exc.value.status_code == 400


async def test_export_bad_format_400(db, tenant_a):
    user, run = await _seed_run_with_groups(db, tenant_a)
    with pytest.raises(HTTPException) as exc:
        await export_run_section(str(run.id), user=user, db=db, section="groups", format="pdf")
    assert exc.value.status_code == 400


async def test_export_other_tenant_run_404(db, tenant_a, tenant_b):
    user, _ = await create_test_user(db, tenant_a)
    run_b = await create_test_recon_run(db, tenant_b.id, status="completed")
    await db.flush()
    with pytest.raises(HTTPException) as exc:
        await export_run_section(str(run_b.id), user=user, db=db, section="groups", format="csv")
    assert exc.value.status_code == 404


# ---------------------------------------------------------------------------
# CSV-injection escaping (Task 6 hardening) — OWASP formula-injection
# mitigation on this new export surface only.
# ---------------------------------------------------------------------------
async def test_export_proposals_narrative_formula_injection_escaped_csv_and_xlsx(db, tenant_a):
    """A free-text narrative starting with '=' must be quote-prefixed in both
    the CSV and XLSX outputs, and openpyxl must never auto-type it as a
    formula cell."""
    user, run = await _seed_run_with_groups(db, tenant_a)
    proposal = (
        (
            await db.execute(
                select(ReconResolutionProposal).where(
                    ReconResolutionProposal.run_id == run.id,
                    ReconResolutionProposal.root_cause == "fees",
                )
            )
        )
        .scalars()
        .first()
    )
    malicious = '=HYPERLINK("http://evil.example","click")'
    proposal.narrative = malicious
    await db.flush()

    csv_response = await export_run_section(
        str(run.id), user=user, db=db, section="proposals", format="csv", group_key="fees:book_fee_line:deposit"
    )
    rows = _csv_rows(await _body_bytes(csv_response))
    row = dict(zip(_PROPOSALS_HEADERS, rows[1]))
    assert row["narrative"] == f"'{malicious}"

    xlsx_response = await export_run_section(
        str(run.id), user=user, db=db, section="proposals", format="xlsx", group_key="fees:book_fee_line:deposit"
    )
    wb = load_workbook(io.BytesIO(await _body_bytes(xlsx_response)))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    narrative_col = headers.index("narrative") + 1
    cell = ws.cell(row=2, column=narrative_col)
    assert cell.value == f"'{malicious}"
    assert cell.data_type == "s"  # never auto-typed as a formula


async def test_export_results_negative_variance_not_escaped_stays_numeric(db, tenant_a):
    """A negative Decimal amount is a NUMBER, never a formula-injection
    string — CSV keeps the exact numeric string, XLSX keeps a numeric cell.
    Uses a non-whole fraction (-9.35, not -9.00) so the assertion isn't
    confused by openpyxl's own load-time int/float coercion for whole numbers."""
    user, _ = await create_test_user(db, tenant_a)
    run = await create_test_recon_run(db, tenant_a.id, status="completed")
    await create_test_recon_result(
        db,
        tenant_a.id,
        run.id,
        status="pending",
        bucket="needs_review",
        match_type="deterministic",
        variance_type="fees",
        variance_amount=Decimal("-9.35"),
        stripe_amount=Decimal("990.65"),
        netsuite_amount=Decimal("1000.00"),
        evidence={"charge_source_id": "ch_neg", "order_reference": "R9"},
    )
    await db.flush()

    csv_response = await export_run_section(str(run.id), user=user, db=db, section="results", format="csv")
    rows = _csv_rows(await _body_bytes(csv_response))
    row = dict(zip(_RESULTS_HEADERS, rows[1]))
    assert row["variance_amount"] == "-9.35"

    xlsx_response = await export_run_section(str(run.id), user=user, db=db, section="results", format="xlsx")
    wb = load_workbook(io.BytesIO(await _body_bytes(xlsx_response)))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    variance_col = headers.index("variance_amount") + 1
    cell = ws.cell(row=2, column=variance_col)
    assert cell.value == -9.35
    assert isinstance(cell.value, float)
    assert cell.data_type == "n"


async def test_export_filename_includes_section_dates_and_group_key(db, tenant_a):
    """group_key's own colons (root_cause:action:booking_vehicle) are illegal
    in a Windows filename, so they must come out sanitized to dashes — the
    filename is a download artifact, not the raw query param."""
    user, run = await _seed_run_with_groups(db, tenant_a)
    response = await export_run_section(
        str(run.id),
        user=user,
        db=db,
        section="proposals",
        format="csv",
        group_key="fees:book_fee_line:deposit",
    )
    disposition = response.headers["content-disposition"]
    expected = f"recon-proposals-{run.date_from.isoformat()}-{run.date_to.isoformat()}-fees-book_fee_line-deposit.csv"
    assert expected in disposition
    assert ":" not in disposition


async def test_export_filename_sanitizes_any_char_outside_safe_set(db, tenant_a):
    """Every char outside [A-Za-z0-9._-] becomes '-' — not just colons.
    group_key is caller-supplied and lands straight in a response header, so
    this also closes off header-injection/path-traversal characters, not
    only the Windows-filename concern. Still root_cause:action:vehicle (3
    colon-separated parts, per `_parse_group_key`) — the unsafe chars (/, space,
    *) live inside the parts themselves."""
    user, run = await _seed_run_with_groups(db, tenant_a)
    response = await export_run_section(
        str(run.id),
        user=user,
        db=db,
        section="proposals",
        format="csv",
        group_key="fe/es:bo ok:de*it",
    )
    disposition = response.headers["content-disposition"]
    expected = f'filename="recon-proposals-{run.date_from.isoformat()}-{run.date_to.isoformat()}-fe-es-bo-ok-de-it.csv"'
    assert expected in disposition
