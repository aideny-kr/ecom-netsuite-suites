"""Integration test — full pricing pipeline: config → engine → template fill → output."""

from decimal import Decimal

from app.schemas.pricing import CurrencyConfig, PricingInput, TenantPricingConfig
from app.services.pricing_engine import PricingEngine
from app.services.template_filler import TemplateFiller


class TestPricingIntegration:
    def test_full_pipeline_default_output(self):
        """Config → PricingEngine.convert_batch → TemplateFiller.generate_default_output → verify Excel."""
        config = TenantPricingConfig(
            base_currency="USD",
            eur_fx_rate=Decimal("0.92"),
            currencies={
                "GBP": CurrencyConfig(
                    fx_rate=Decimal("0.79"), tier="usd_based", vat_rate=Decimal("0.20"), rounding_rule="nearest_9"
                ),
                "JPY": CurrencyConfig(
                    fx_rate=Decimal("149.50"), tier="usd_based", vat_rate=Decimal("0.10"), rounding_rule="nearest_100"
                ),
            },
        )
        items = [
            PricingInput(sku="SKU-001", item_name="Widget", usd_price=Decimal("99.00")),
            PricingInput(sku="SKU-002", item_name="Gadget", usd_price=Decimal("199.00")),
        ]
        engine = PricingEngine()
        results = engine.convert_batch(items, config)
        assert len(results) == 2

        filler = TemplateFiller()
        wb = filler.generate_default_output(results)

        # Verify the 3 expected sheets
        assert "Prices" in wb.sheetnames
        assert "Conversion Details" in wb.sheetnames
        assert "Config Snapshot" in wb.sheetnames
        assert len(wb.sheetnames) == 3

        # Prices sheet: row 1 = headers, rows 2-3 = data
        ws_prices = wb["Prices"]
        assert ws_prices.cell(row=1, column=1).value == "SKU"
        assert ws_prices.cell(row=2, column=1).value == "SKU-001"
        assert ws_prices.cell(row=3, column=1).value == "SKU-002"

        # Config Snapshot sheet should record SKU count
        ws_config = wb["Config Snapshot"]
        assert ws_config.cell(row=3, column=2).value == 2

    def test_full_pipeline_template_fill(self):
        """Config → Engine → TemplateFiller.fill with template → verify filled values."""
        import openpyxl

        template_wb = openpyxl.Workbook()
        ws = template_wb.active
        ws["A1"] = "SKU"
        ws["B1"] = "USD"
        ws["C1"] = "GBP"
        ws["A2"] = "SKU-001"
        ws["B2"] = 99.00

        config = TenantPricingConfig(
            base_currency="USD",
            eur_fx_rate=Decimal("0.92"),
            currencies={
                "GBP": CurrencyConfig(
                    fx_rate=Decimal("0.79"), tier="usd_based", vat_rate=Decimal("0.20"), rounding_rule="nearest_9"
                ),
            },
        )
        items = [PricingInput(sku="SKU-001", usd_price=Decimal("99.00"))]

        engine = PricingEngine()
        results = engine.convert_batch(items, config)

        filler = TemplateFiller()
        mapping = filler.detect_columns(template_wb)
        filler.fill(template_wb, results, mapping)

        ws = template_wb.active
        # GBP: 99.00 * 0.79 * 1.20 = 93.852 → nearest_9 → 99
        gbp_val = ws.cell(row=2, column=3).value
        assert gbp_val is not None
        assert float(gbp_val) > 0
        assert float(gbp_val) == 99.0

    def test_netsuite_csv_output(self):
        """Config → Engine → TemplateFiller.generate_netsuite_csv → verify CSV format."""
        config = TenantPricingConfig(
            base_currency="USD",
            eur_fx_rate=Decimal("0.92"),
            currencies={
                "GBP": CurrencyConfig(
                    fx_rate=Decimal("0.79"), tier="usd_based", vat_rate=Decimal("0.20"), rounding_rule="nearest_9"
                ),
            },
        )
        items = [PricingInput(sku="SKU-001", usd_price=Decimal("99.00"))]

        engine = PricingEngine()
        results = engine.convert_batch(items, config)

        filler = TemplateFiller()
        csv_content = filler.generate_netsuite_csv(results)
        assert isinstance(csv_content, str)
        lines = csv_content.strip().split("\n")
        assert len(lines) >= 2  # header + at least 1 row
        assert "SKU-001" in csv_content
        assert "GBP" in csv_content

        # Header columns
        header = lines[0]
        assert "External ID" in header
        assert "Currency" in header
        assert "Rate" in header

    def test_two_currencies_produce_two_csv_rows(self):
        """Each currency generates a separate CSV row per SKU."""
        config = TenantPricingConfig(
            base_currency="USD",
            eur_fx_rate=Decimal("0.92"),
            currencies={
                "GBP": CurrencyConfig(fx_rate=Decimal("0.79"), tier="usd_based", rounding_rule="nearest_9"),
                "JPY": CurrencyConfig(fx_rate=Decimal("149.50"), tier="usd_based", rounding_rule="nearest_100"),
            },
        )
        items = [PricingInput(sku="SKU-001", usd_price=Decimal("100.00"))]
        engine = PricingEngine()
        results = engine.convert_batch(items, config)

        filler = TemplateFiller()
        csv_content = filler.generate_netsuite_csv(results)
        lines = [l for l in csv_content.strip().split("\n") if l]
        # 1 header + 2 data rows (GBP + JPY)
        assert len(lines) == 3

    def test_default_output_conversion_details_populated(self):
        """Conversion Details sheet has one row per SKU-currency combination."""
        config = TenantPricingConfig(
            base_currency="USD",
            eur_fx_rate=Decimal("0.92"),
            currencies={
                "GBP": CurrencyConfig(
                    fx_rate=Decimal("0.79"), tier="usd_based", vat_rate=Decimal("0.20"), rounding_rule="nearest_9"
                ),
                "JPY": CurrencyConfig(
                    fx_rate=Decimal("149.50"), tier="usd_based", vat_rate=Decimal("0.10"), rounding_rule="nearest_100"
                ),
            },
        )
        items = [
            PricingInput(sku="SKU-001", item_name="Widget", usd_price=Decimal("99.00")),
        ]
        engine = PricingEngine()
        results = engine.convert_batch(items, config)
        filler = TemplateFiller()
        wb = filler.generate_default_output(results)

        ws_audit = wb["Conversion Details"]
        # Row 1 = headers; rows 2-3 = GBP + JPY for SKU-001 (sorted alphabetically)
        assert ws_audit.cell(row=2, column=1).value == "SKU-001"
        assert ws_audit.cell(row=2, column=2).value == "GBP"
        assert ws_audit.cell(row=3, column=2).value == "JPY"
