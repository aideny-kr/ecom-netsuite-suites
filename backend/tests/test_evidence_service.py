"""Tests for evidence pack generation."""

import io
import uuid
from datetime import date
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from app.services.reconciliation.evidence_service import (
    _EXCEPTION_FILL,
    _SUGGESTED_FILL,
    EvidencePackGenerator,
)


@pytest.fixture
def sample_results() -> list[dict]:
    """Sample reconciliation results for evidence generation.

    Fixtures carry bucket + status consistent with the four-bucket taxonomy:
      - deterministic clean match  → bucket="matches"
      - fuzzy timing gap           → bucket="rules"   (fuzzy/rule-based)
      - unmatched                  → bucket="needs_review"
    """
    return [
        {
            "id": str(uuid.uuid4()),
            "match_type": "deterministic",
            "confidence": Decimal("1.0"),
            "status": "auto_matched",
            "bucket": "matches",
            "stripe_amount": Decimal("1455.00"),
            "netsuite_amount": Decimal("1455.00"),
            "variance_amount": Decimal("0.00"),
            "variance_type": None,
            "variance_explanation": None,
            "currency": "USD",
            "match_rule": "exact_payout_id",
            "evidence": {
                "payout_source_id": "po_test001",
                "deposit_ids": ["12001"],
            },
        },
        {
            "id": str(uuid.uuid4()),
            "match_type": "fuzzy",
            "confidence": Decimal("0.85"),
            "status": "suggested",
            "bucket": "rules",
            "stripe_amount": Decimal("3104.00"),
            "netsuite_amount": Decimal("3104.00"),
            "variance_amount": Decimal("0.00"),
            "variance_type": "timing",
            "variance_explanation": "Deposit is 2 days after payout arrival",
            "currency": "USD",
            "match_rule": "amount_exact+within_2_days",
            "evidence": {
                "payout_source_id": "po_test005",
                "deposit_ids": ["12005"],
            },
        },
        {
            "id": str(uuid.uuid4()),
            "match_type": "unmatched",
            "confidence": Decimal("0.0"),
            "status": "pending",
            "bucket": "needs_review",
            "stripe_amount": Decimal("863.30"),
            "netsuite_amount": None,
            "variance_amount": Decimal("863.30"),
            "variance_type": "missing",
            "variance_explanation": "No matching deposit found",
            "currency": "USD",
            "match_rule": "no_match",
            "evidence": {
                "payout_source_id": "po_test007",
                "deposit_ids": [],
            },
        },
    ]


class TestEvidencePackGenerator:
    def test_generates_excel(self, sample_results):
        """Evidence pack should produce a valid Excel file."""
        generator = EvidencePackGenerator()
        excel_bytes = generator.generate_excel(
            results=sample_results,
            run_id="test-run-001",
            date_from=date(2026, 3, 1),
            date_to=date(2026, 3, 31),
        )

        assert isinstance(excel_bytes, io.BytesIO)
        wb = load_workbook(excel_bytes)
        assert len(wb.sheetnames) >= 2  # Summary + Details at minimum

    def test_summary_sheet_has_totals(self, sample_results):
        """Summary sheet should contain match/exception/unmatched counts."""
        generator = EvidencePackGenerator()
        excel_bytes = generator.generate_excel(
            results=sample_results,
            run_id="test-run-002",
            date_from=date(2026, 3, 1),
            date_to=date(2026, 3, 31),
        )

        wb = load_workbook(excel_bytes)
        summary_ws = wb["Summary"]

        # Check that summary data exists (look for key labels)
        all_values = [cell.value for row in summary_ws.iter_rows() for cell in row if cell.value]
        text = " ".join(str(v) for v in all_values)

        assert "Matched" in text or "matched" in text
        assert "Exception" in text or "exception" in text or "Unmatched" in text

    def test_exceptions_sheet_exists(self, sample_results):
        """Should have an Exceptions sheet with needs_review items."""
        generator = EvidencePackGenerator()
        excel_bytes = generator.generate_excel(
            results=sample_results,
            run_id="test-run-003",
            date_from=date(2026, 3, 1),
            date_to=date(2026, 3, 31),
        )

        wb = load_workbook(excel_bytes)
        assert "Exceptions" in wb.sheetnames

    def test_empty_results_still_generates(self):
        """Even with no results, should produce a valid Excel."""
        generator = EvidencePackGenerator()
        excel_bytes = generator.generate_excel(
            results=[],
            run_id="test-run-empty",
            date_from=date(2026, 3, 1),
            date_to=date(2026, 3, 31),
        )

        assert isinstance(excel_bytes, io.BytesIO)
        wb = load_workbook(excel_bytes)
        assert "Summary" in wb.sheetnames

    # ---------------------------------------------------------------------------
    # Bucket-based categorization tests (the core regression suite)
    # ---------------------------------------------------------------------------

    def test_summary_counts_by_bucket(self, sample_results):
        """Summary counts are driven by the four authoritative buckets, not confidence.

        The four buckets PARTITION the run:
          matches + auto_classifications + rules + needs_review == total_results.
        sample_results has: 1 matches, 0 auto_classifications, 1 rules, 1 needs_review.
        """
        generator = EvidencePackGenerator()
        excel_bytes = generator.generate_excel(
            results=sample_results,
            run_id="test-run-counts",
            date_from=date(2026, 3, 1),
            date_to=date(2026, 3, 31),
        )

        wb = load_workbook(excel_bytes)
        summary_ws = wb["Summary"]

        # Collect label→value pairs from the summary sheet (col A = label, col B = value)
        summary = {}
        for row in summary_ws.iter_rows(min_col=1, max_col=2):
            label_cell, value_cell = row
            if label_cell.value:
                summary[str(label_cell.value)] = value_cell.value

        # sample_results has 1 matches, 0 auto_classifications, 1 rules, 1 needs_review (unmatched)
        assert summary.get("Auto-Matched") == "1", (
            "Auto-Matched should count bucket='matches' rows only (got: %r)" % summary
        )
        assert summary.get("Auto-Classified") == "0", (
            "Auto-Classified should count bucket='auto_classifications' rows (got: %r)" % summary
        )
        assert summary.get("Rules (Fuzzy)") == "1", f"Rules (Fuzzy) should count bucket='rules' rows (got: {summary!r})"
        assert summary.get("Needs Review (Exceptions)") == "1", (
            "Needs Review should count bucket='needs_review' rows (got: %r)" % summary
        )
        assert summary.get("Unmatched (within Needs Review)") == "1", (
            "Unmatched should still count match_type='unmatched' rows (got: %r)" % summary
        )

        # The four primary buckets PARTITION the run:
        #   matches + auto_classifications + rules + needs_review == total_results.
        # (Unmatched ⊆ Needs Review — it is a sub-detail, not an extra partition.)
        matched = int(summary["Auto-Matched"])
        auto_classified = int(summary["Auto-Classified"])
        rules = int(summary["Rules (Fuzzy)"])
        needs_review = int(summary["Needs Review (Exceptions)"])
        total = int(summary["Total Results"])
        assert matched + auto_classified + rules + needs_review == total, (
            "Summary buckets must partition the run "
            "(matches+auto_classifications+rules+needs_review==total); got %r" % summary
        )

    def test_exceptions_sheet_contains_needs_review_rows(self, sample_results):
        """Exceptions sheet must contain the needs_review (unmatched) row."""
        generator = EvidencePackGenerator()
        excel_bytes = generator.generate_excel(
            results=sample_results,
            run_id="test-run-exc-content",
            date_from=date(2026, 3, 1),
            date_to=date(2026, 3, 31),
        )

        wb = load_workbook(excel_bytes)
        exc_ws = wb["Exceptions"]

        # Row 1 is headers; data rows start at 2.
        data_rows = list(exc_ws.iter_rows(min_row=2, values_only=True))
        non_empty = [r for r in data_rows if any(v is not None for v in r)]

        # Only the unmatched/needs_review row should appear.
        assert len(non_empty) == 1, "Exceptions sheet should contain exactly 1 needs_review row (got %d)" % len(
            non_empty
        )
        # First column is match_type; should be "unmatched"
        assert non_empty[0][0] == "unmatched"

    def test_exceptions_sheet_excludes_matches_bucket(self, sample_results):
        """bucket='matches' rows must NOT appear in Exceptions regardless of confidence."""
        # Find the deterministic/matches row payout id
        matches_row = next(r for r in sample_results if r["bucket"] == "matches")
        payout_id = matches_row["evidence"]["payout_source_id"]

        generator = EvidencePackGenerator()
        excel_bytes = generator.generate_excel(
            results=sample_results,
            run_id="test-run-exc-excl",
            date_from=date(2026, 3, 1),
            date_to=date(2026, 3, 31),
        )

        wb = load_workbook(excel_bytes)
        exc_ws = wb["Exceptions"]

        # Look up the Payout ID column by header name (don't hard-code the index).
        header_values = [cell.value for cell in exc_ws[1]]
        payout_idx = header_values.index("Payout ID")  # 0-based for values_only rows
        payout_ids_in_exc = [
            row[payout_idx] for row in exc_ws.iter_rows(min_row=2, values_only=True) if row[payout_idx] is not None
        ]
        assert payout_id not in payout_ids_in_exc, "bucket='matches' row should not appear in Exceptions sheet"

    # ---------------------------------------------------------------------------
    # Regression test: low-confidence match stays in auto-matched, not exceptions
    # ---------------------------------------------------------------------------

    def test_low_confidence_matches_bucket_not_exception(self):
        """A row with bucket='matches' and low advisory confidence (0.60) must be
        counted as Auto-Matched and must NOT appear in the Exceptions sheet.

        This is the core regression: the old code keyed on confidence >= 0.95
        and would have mis-filed this row into Exceptions.
        """
        low_conf_match = {
            "id": str(uuid.uuid4()),
            "match_type": "deterministic",
            "confidence": Decimal("0.6000"),  # low advisory composite (date-gap)
            "status": "auto_matched",
            "bucket": "matches",
            "stripe_amount": Decimal("2500.00"),
            "netsuite_amount": Decimal("2500.00"),
            "variance_amount": Decimal("0.00"),
            "variance_type": None,
            "variance_explanation": None,
            "currency": "USD",
            "match_rule": "exact_payout_id",
            "evidence": {
                "payout_source_id": "po_lowconf",
                "deposit_ids": ["99001"],
            },
        }
        results = [low_conf_match]

        generator = EvidencePackGenerator()
        excel_bytes = generator.generate_excel(
            results=results,
            run_id="test-run-lowconf",
            date_from=date(2026, 3, 1),
            date_to=date(2026, 3, 31),
        )

        wb = load_workbook(excel_bytes)

        # (a) Counted in Auto-Matched
        summary_ws = wb["Summary"]
        summary = {}
        for row in summary_ws.iter_rows(min_col=1, max_col=2):
            label_cell, value_cell = row
            if label_cell.value:
                summary[str(label_cell.value)] = value_cell.value
        assert summary.get("Auto-Matched") == "1", (
            "Low-confidence bucket='matches' row must be counted as Auto-Matched; got %r" % summary
        )

        # (b) NOT present in Exceptions sheet
        exc_ws = wb["Exceptions"]
        data_rows = list(exc_ws.iter_rows(min_row=2, values_only=True))
        non_empty = [r for r in data_rows if any(v is not None for v in r)]
        assert len(non_empty) == 0, (
            "Low-confidence bucket='matches' row must NOT appear in Exceptions sheet; got %d rows" % len(non_empty)
        )

        # (c) Row fill: green (_MATCHED_FILL) in All Results sheet
        all_ws = wb["All Results"]
        data_cell_fill = all_ws.cell(row=2, column=1).fill.fgColor.rgb
        # _MATCHED_FILL is start_color="d4edda"
        assert data_cell_fill.upper().endswith("D4EDDA"), (
            "Low-confidence bucket='matches' row should have green (matched) fill; got %r" % data_cell_fill
        )

    def test_high_confidence_needs_review_is_exception(self):
        """A row with bucket='needs_review' and HIGH advisory confidence (0.99) —
        i.e. a confident-but-material-variance auto_matched row — must appear in
        Exceptions and must NOT be counted as Auto-Matched.

        Proves that bucket, not confidence/status, drives categorization.
        """
        high_conf_needs_review = {
            "id": str(uuid.uuid4()),
            "match_type": "deterministic",
            "confidence": Decimal("0.9900"),  # high confidence, but material variance
            "status": "auto_matched",
            "bucket": "needs_review",
            "stripe_amount": Decimal("5000.00"),
            "netsuite_amount": Decimal("4900.00"),
            "variance_amount": Decimal("100.00"),
            "variance_type": "amount",
            "variance_explanation": "Material variance: $100",
            "currency": "USD",
            "match_rule": "amount_approx",
            "evidence": {
                "payout_source_id": "po_highconf_material",
                "deposit_ids": ["99002"],
            },
        }
        results = [high_conf_needs_review]

        generator = EvidencePackGenerator()
        excel_bytes = generator.generate_excel(
            results=results,
            run_id="test-run-highconf-needs-review",
            date_from=date(2026, 3, 1),
            date_to=date(2026, 3, 31),
        )

        wb = load_workbook(excel_bytes)

        # Auto-Matched count must be 0 (not counted despite high confidence)
        summary_ws = wb["Summary"]
        summary = {}
        for row in summary_ws.iter_rows(min_col=1, max_col=2):
            label_cell, value_cell = row
            if label_cell.value:
                summary[str(label_cell.value)] = value_cell.value
        # A 0 count now renders "0" exactly (no longer blank).
        assert summary.get("Auto-Matched") == "0", (
            "bucket='needs_review' row must NOT be counted as Auto-Matched; got %r" % summary
        )

        # Must appear in Exceptions sheet
        exc_ws = wb["Exceptions"]
        data_rows = list(exc_ws.iter_rows(min_row=2, values_only=True))
        non_empty = [r for r in data_rows if any(v is not None for v in r)]
        assert len(non_empty) == 1, "bucket='needs_review' row must appear in Exceptions sheet; got %d rows" % len(
            non_empty
        )

        # Row fill in All Results: yellow (_EXCEPTION_FILL = "fff3cd")
        all_ws = wb["All Results"]
        data_cell_fill = all_ws.cell(row=2, column=1).fill.fgColor.rgb
        assert data_cell_fill.upper().endswith("FFF3CD"), (
            "bucket='needs_review' row should have yellow (exception) fill; got %r" % data_cell_fill
        )

    def test_summary_needs_review_count_includes_material_match(self):
        """A material needs_review row (deterministic match_type, NOT unmatched) must be
        counted in 'Needs Review (Exceptions)' but NOT in 'Unmatched' — proving the
        Needs Review count is the authoritative exception total, and the buckets
        still partition the run.
        """
        results = [
            # bucket='matches'
            {
                "id": str(uuid.uuid4()),
                "match_type": "deterministic",
                "confidence": Decimal("1.0"),
                "status": "auto_matched",
                "bucket": "matches",
                "stripe_amount": Decimal("1000.00"),
                "netsuite_amount": Decimal("1000.00"),
                "variance_amount": Decimal("0.00"),
                "variance_type": None,
                "variance_explanation": None,
                "currency": "USD",
                "match_rule": "exact_payout_id",
                "evidence": {"payout_source_id": "po_m1", "deposit_ids": ["1"]},
            },
            # bucket='needs_review' but match_type='deterministic' (material variance, NOT unmatched)
            {
                "id": str(uuid.uuid4()),
                "match_type": "deterministic",
                "confidence": Decimal("0.99"),
                "status": "auto_matched",
                "bucket": "needs_review",
                "stripe_amount": Decimal("5000.00"),
                "netsuite_amount": Decimal("4800.00"),
                "variance_amount": Decimal("200.00"),
                "variance_type": "amount",
                "variance_explanation": "Material variance: $200",
                "currency": "USD",
                "match_rule": "amount_approx",
                "evidence": {"payout_source_id": "po_material", "deposit_ids": ["2"]},
            },
        ]

        generator = EvidencePackGenerator()
        excel_bytes = generator.generate_excel(
            results=results,
            run_id="test-run-material-needs-review",
            date_from=date(2026, 3, 1),
            date_to=date(2026, 3, 31),
        )

        wb = load_workbook(excel_bytes)
        summary_ws = wb["Summary"]
        summary = {}
        for row in summary_ws.iter_rows(min_col=1, max_col=2):
            label_cell, value_cell = row
            if label_cell.value:
                summary[str(label_cell.value)] = value_cell.value

        # Needs Review captures the material match; Unmatched does NOT (it's a real match).
        assert summary.get("Needs Review (Exceptions)") == "1", (
            "Material needs_review row must count in Needs Review; got %r" % summary
        )
        assert summary.get("Unmatched (within Needs Review)") == "0", (
            "Material needs_review row is a match, not unmatched; got %r" % summary
        )
        assert summary.get("Auto-Matched") == "1", summary
        assert summary.get("Auto-Classified") == "0", summary
        assert summary.get("Rules (Fuzzy)") == "0", summary

        # Four-way partition still holds.
        matched = int(summary["Auto-Matched"])
        auto_classified = int(summary["Auto-Classified"])
        rules = int(summary["Rules (Fuzzy)"])
        needs_review = int(summary["Needs Review (Exceptions)"])
        total = int(summary["Total Results"])
        assert matched + auto_classified + rules + needs_review == total, (
            "Buckets must partition the run; got %r" % summary
        )

    def test_confidence_column_still_present(self, sample_results):
        """The Confidence data column must still be present (advisory display only)."""
        generator = EvidencePackGenerator()
        excel_bytes = generator.generate_excel(
            results=sample_results,
            run_id="test-run-conf-col",
            date_from=date(2026, 3, 1),
            date_to=date(2026, 3, 31),
        )

        wb = load_workbook(excel_bytes)
        all_ws = wb["All Results"]

        # Check headers row for "Confidence"
        header_values = [cell.value for cell in all_ws[1]]
        assert "Confidence" in header_values, "Confidence column must remain in All Results sheet headers"

        # And the actual value for the first row should be the float confidence
        conf_col_idx = header_values.index("Confidence") + 1  # 1-based
        conf_value = all_ws.cell(row=2, column=conf_col_idx).value
        assert conf_value is not None and isinstance(conf_value, (int, float)), (
            "Confidence cell should contain a numeric advisory value; got %r" % conf_value
        )

    # ---------------------------------------------------------------------------
    # Fix 1 — four-bucket fill coloring: auto_classifications + rules get BLUE
    # (not yellow _EXCEPTION_FILL), and are counted under their own labels.
    # ---------------------------------------------------------------------------

    def test_suggested_buckets_get_blue_fill_not_yellow(self):
        """auto_classifications and rules rows must receive _SUGGESTED_FILL (blue),
        NOT _EXCEPTION_FILL (yellow). They are deterministic/rule-based matches,
        not exceptions, so yellow would be misleading.
        """
        auto_class_row = {
            "id": str(uuid.uuid4()),
            "match_type": "deterministic",
            "confidence": Decimal("0.95"),
            "status": "auto_matched",
            "bucket": "auto_classifications",
            "stripe_amount": Decimal("300.00"),
            "netsuite_amount": Decimal("300.20"),
            "variance_amount": Decimal("0.20"),
            "variance_type": "amount_mismatch",
            "variance_explanation": "Immaterial variance: $0.20",
            "currency": "USD",
            "match_rule": "order_reference_exact",
            "evidence": {"payout_source_id": "po_ac1", "deposit_ids": ["ac001"]},
        }
        rules_row = {
            "id": str(uuid.uuid4()),
            "match_type": "fuzzy",
            "confidence": Decimal("0.85"),
            "status": "suggested",
            "bucket": "rules",
            "stripe_amount": Decimal("500.00"),
            "netsuite_amount": Decimal("500.00"),
            "variance_amount": Decimal("0.00"),
            "variance_type": None,
            "variance_explanation": None,
            "currency": "USD",
            "match_rule": "amount+date+currency",
            "evidence": {"payout_source_id": "po_r1", "deposit_ids": ["r001"]},
        }
        results = [auto_class_row, rules_row]

        generator = EvidencePackGenerator()
        excel_bytes = generator.generate_excel(
            results=results,
            run_id="test-run-blue-fill",
            date_from=date(2026, 3, 1),
            date_to=date(2026, 3, 31),
        )

        wb = load_workbook(excel_bytes)
        all_ws = wb["All Results"]

        # Both data rows (row 2 = auto_classifications, row 3 = rules) must be BLUE.
        blue_rgb = _SUGGESTED_FILL.fgColor.rgb.upper()
        yellow_rgb = _EXCEPTION_FILL.fgColor.rgb.upper()

        for row_idx, label in ((2, "auto_classifications"), (3, "rules")):
            cell_rgb = all_ws.cell(row=row_idx, column=1).fill.fgColor.rgb.upper()
            assert cell_rgb == blue_rgb, (
                f"bucket='{label}' row should have blue (_SUGGESTED_FILL={blue_rgb!r}) fill; got {cell_rgb!r}"
            )
            assert cell_rgb != yellow_rgb, (
                f"bucket='{label}' row must NOT have yellow (_EXCEPTION_FILL) fill; got {cell_rgb!r}"
            )

        # Neither row should appear on the Exceptions sheet.
        exc_ws = wb["Exceptions"]
        data_rows = list(exc_ws.iter_rows(min_row=2, values_only=True))
        non_empty = [r for r in data_rows if any(v is not None for v in r)]
        assert len(non_empty) == 0, (
            "auto_classifications + rules rows must NOT appear in Exceptions sheet; got %d rows" % len(non_empty)
        )

        # Summary: auto_classifications counted under "Auto-Classified" (not "Needs Review").
        summary_ws = wb["Summary"]
        summary = {}
        for row in summary_ws.iter_rows(min_col=1, max_col=2):
            label_cell, value_cell = row
            if label_cell.value:
                summary[str(label_cell.value)] = value_cell.value

        assert summary.get("Auto-Classified") == "1", (
            "auto_classifications row must be counted under 'Auto-Classified'; got %r" % summary
        )
        assert summary.get("Rules (Fuzzy)") == "1", "rules row must be counted under 'Rules (Fuzzy)'; got %r" % summary
        assert summary.get("Needs Review (Exceptions)") == "0", (
            "Neither auto_classifications nor rules should count as Needs Review; got %r" % summary
        )

    # ---------------------------------------------------------------------------
    # Fix 2 — unknown / None bucket is treated as needs_review for counts + Exceptions
    # ---------------------------------------------------------------------------

    def test_unknown_bucket_counted_as_needs_review(self):
        """A row with bucket=None or an unrecognised bucket string must be:
        (a) counted under 'Needs Review (Exceptions)' in the summary,
        (b) present on the Exceptions sheet,
        (c) painted yellow (_EXCEPTION_FILL) in All Results,
        (d) the four bucket counts still sum to total_results (partition holds).
        """
        none_bucket_row = {
            "id": str(uuid.uuid4()),
            "match_type": "deterministic",
            "confidence": Decimal("0.80"),
            "status": "auto_matched",
            "bucket": None,  # missing bucket
            "stripe_amount": Decimal("100.00"),
            "netsuite_amount": Decimal("100.00"),
            "variance_amount": Decimal("0.00"),
            "variance_type": None,
            "variance_explanation": None,
            "currency": "USD",
            "match_rule": "exact_payout_id",
            "evidence": {"payout_source_id": "po_none_bucket", "deposit_ids": ["nb001"]},
        }
        weird_bucket_row = {
            "id": str(uuid.uuid4()),
            "match_type": "fuzzy",
            "confidence": Decimal("0.70"),
            "status": "suggested",
            "bucket": "weird",  # unknown bucket string
            "stripe_amount": Decimal("200.00"),
            "netsuite_amount": Decimal("200.00"),
            "variance_amount": Decimal("0.00"),
            "variance_type": None,
            "variance_explanation": None,
            "currency": "USD",
            "match_rule": "amount+date",
            "evidence": {"payout_source_id": "po_weird_bucket", "deposit_ids": ["wb001"]},
        }
        known_matches_row = {
            "id": str(uuid.uuid4()),
            "match_type": "deterministic",
            "confidence": Decimal("1.00"),
            "status": "auto_matched",
            "bucket": "matches",
            "stripe_amount": Decimal("300.00"),
            "netsuite_amount": Decimal("300.00"),
            "variance_amount": Decimal("0.00"),
            "variance_type": None,
            "variance_explanation": None,
            "currency": "USD",
            "match_rule": "exact_payout_id",
            "evidence": {"payout_source_id": "po_known", "deposit_ids": ["k001"]},
        }
        results = [none_bucket_row, weird_bucket_row, known_matches_row]

        generator = EvidencePackGenerator()
        excel_bytes = generator.generate_excel(
            results=results,
            run_id="test-run-unknown-bucket",
            date_from=date(2026, 3, 1),
            date_to=date(2026, 3, 31),
        )

        wb = load_workbook(excel_bytes)

        # (a) Both unknown-bucket rows counted under Needs Review
        summary_ws = wb["Summary"]
        summary = {}
        for row in summary_ws.iter_rows(min_col=1, max_col=2):
            label_cell, value_cell = row
            if label_cell.value:
                summary[str(label_cell.value)] = value_cell.value

        assert summary.get("Needs Review (Exceptions)") == "2", (
            "Both None and unknown-string bucket rows must be counted as Needs Review; got %r" % summary
        )
        assert summary.get("Auto-Matched") == "1", (
            "Only the known 'matches' row should be Auto-Matched; got %r" % summary
        )

        # (d) Partition holds: matches + auto_classified + rules + needs_review == total
        matched = int(summary["Auto-Matched"])
        auto_classified = int(summary["Auto-Classified"])
        rules = int(summary["Rules (Fuzzy)"])
        needs_review = int(summary["Needs Review (Exceptions)"])
        total = int(summary["Total Results"])
        assert matched + auto_classified + rules + needs_review == total, (
            "Buckets must partition the run (including unknown bucket rows); got %r" % summary
        )

        # (b) Both unknown-bucket rows present on Exceptions sheet
        exc_ws = wb["Exceptions"]
        data_rows = list(exc_ws.iter_rows(min_row=2, values_only=True))
        non_empty = [r for r in data_rows if any(v is not None for v in r)]
        assert len(non_empty) == 2, (
            "Both None and unknown-string bucket rows must appear on Exceptions sheet; got %d rows" % len(non_empty)
        )

        # (c) Both unknown-bucket rows get yellow (_EXCEPTION_FILL) in All Results
        all_ws = wb["All Results"]
        yellow_rgb = _EXCEPTION_FILL.fgColor.rgb.upper()
        # rows 2 and 3 are none_bucket_row and weird_bucket_row
        for row_idx in (2, 3):
            cell_rgb = all_ws.cell(row=row_idx, column=1).fill.fgColor.rgb.upper()
            assert cell_rgb == yellow_rgb, "Unknown-bucket row %d should have yellow (_EXCEPTION_FILL) fill; got %r" % (
                row_idx,
                cell_rgb,
            )
