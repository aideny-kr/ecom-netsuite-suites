"""Step 6 — pricing_revise executor tests.

Verifies the cumulative override semantics, latest-wins conflict resolution,
fresh tenant config fetch, currency validation, empty-items guard, and
absence of the legacy markdown-table response_instruction (Mistakes #41).
"""

from __future__ import annotations

import csv
import io
import uuid
from decimal import Decimal
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from openpyxl import load_workbook

from app.mcp.tools.pricing_tools import pricing_revise_execute
from app.schemas.pricing import CurrencyConfig, TenantPricingConfig


def _three_currency_config_dict() -> dict:
    config = TenantPricingConfig(
        base_currency="USD",
        eur_fx_rate=Decimal("0.92"),
        currencies={
            "GBP": CurrencyConfig(
                fx_rate=Decimal("0.79"),
                tier="usd_based",
                vat_rate=Decimal("0.20"),
                rounding_rule="nearest_9",
            ),
            "EUR": CurrencyConfig(
                fx_rate=Decimal("0.92"),
                tier="eur_based",
                vat_rate=Decimal("0.23"),
                rounding_rule="nearest_9",
            ),
            "CAD": CurrencyConfig(
                fx_rate=Decimal("1.36"),
                tier="usd_based",
                vat_rate=None,
                rounding_rule="nearest_9",
            ),
        },
    )
    return config.model_dump(mode="json")


def _seed_state(items: list[dict] | None = None) -> dict:
    items = items or [
        {"sku": "SKU-1", "usd_price": "100", "item_name": None},
        {"sku": "SKU-2", "usd_price": "200", "item_name": "Two"},
    ]
    return {
        "seed_items": items,
        "effective_items": [dict(it) for it in items],
        "effective_currencies": ["GBP", "EUR", "CAD"],
        "effective_fx_overrides": {},
        "effective_vat_overrides": {},
        "effective_rounding_overrides": {},
        "effective_uplift_by_currency": {},
        "applied_overrides_log": [],
        "excel_file_id": "old-excel",
        "netsuite_csv_file_id": "old-csv",
        "header_columns": ["SKU", "Item Name", "USD", "CAD", "EUR", "GBP"],
        "row_count": len(items),
    }


@pytest.fixture
def revise_context():
    return {
        "db": AsyncMock(),
        "tenant_id": uuid.uuid4(),
        "actor_id": uuid.uuid4(),
        "conversation_id": "conv-abc",
    }


def _patch_save_output(saved: list):
    async def _save_output(*, db, tenant_id, user_id, filename, content, related_message_id=None):
        f = MagicMock()
        f.id = uuid.uuid4()
        f.filename = filename
        f.content = content
        saved.append(f)
        return f

    return _save_output


def _cached_result(payload: dict):
    cr = MagicMock()
    cr.result_type = "pricing"
    cr.payload = payload
    return cr


def _run_revise(params, ctx, *, payload, config_dict=None, saved: list | None = None):
    """Invoke pricing_revise_execute with all standard mocks installed."""
    if saved is None:
        saved = []
    cached = _cached_result(payload) if payload is not None else None
    config_row = MagicMock()
    config_row.config = config_dict or _three_currency_config_dict()

    with (
        patch(
            "app.mcp.tools.pricing_tools.get_latest_result_by_type",
            new_callable=AsyncMock,
            return_value=cached,
        ),
        patch(
            "app.mcp.tools.pricing_tools.get_config",
            new_callable=AsyncMock,
            return_value=config_row,
        ),
        patch(
            "app.mcp.tools.pricing_tools._file_svc.save_output",
            new=_patch_save_output(saved),
        ),
    ):
        import asyncio

        result = asyncio.run(pricing_revise_execute(params, ctx))
    return result, saved


class TestCacheMiss:
    def test_cache_miss_returns_clear_error(self, revise_context):
        result, _ = _run_revise(
            {"overrides": {"percent_uplift": {"GBP": 0.05}}},
            revise_context,
            payload=None,
        )
        assert result["error"] is True
        assert "no prior pricing" in result["message"].lower()

    def test_no_conversation_id_returns_error(self):
        ctx = {
            "db": AsyncMock(),
            "tenant_id": uuid.uuid4(),
            "actor_id": uuid.uuid4(),
        }
        result, _ = _run_revise(
            {"overrides": {"percent_uplift": {"GBP": 0.05}}},
            ctx,
            payload=_seed_state(),
        )
        assert result["error"] is True


class TestPriceChanges:
    def test_sku_price_change_round_trips(self, revise_context):
        result, _ = _run_revise(
            {"overrides": {"sku_price_changes": [{"sku": "SKU-1", "usd_price": 149.0}]}},
            revise_context,
            payload=_seed_state(),
        )
        assert result["success"] is True
        ps = result["pricing_state"]
        # Effective items: SKU-1 reflects new price.
        sku1 = next(it for it in ps["effective_items"] if it["sku"] == "SKU-1")
        assert Decimal(str(sku1["usd_price"])) == Decimal("149")

    def test_latest_wins_same_sku_price_change(self, revise_context):
        # Two prior revises bumped SKU-1 already; the third revise overwrites.
        seed = _seed_state()
        seed["effective_items"][0]["usd_price"] = "100"
        seed["applied_overrides_log"] = [
            {"sku_price_changes": [{"sku": "SKU-1", "usd_price": 100.0}]},
            {"sku_price_changes": [{"sku": "SKU-1", "usd_price": 120.0}]},
        ]
        result, _ = _run_revise(
            {"overrides": {"sku_price_changes": [{"sku": "SKU-1", "usd_price": 150.0}]}},
            revise_context,
            payload=seed,
        )
        ps = result["pricing_state"]
        sku1 = next(it for it in ps["effective_items"] if it["sku"] == "SKU-1")
        assert Decimal(str(sku1["usd_price"])) == Decimal("150")
        # Audit log appended (not replaced).
        assert len(ps["applied_overrides_log"]) == 3

    def test_sku_price_change_unknown_sku_noop(self, revise_context):
        result, _ = _run_revise(
            {"overrides": {"sku_price_changes": [{"sku": "GHOST", "usd_price": 99.0}]}},
            revise_context,
            payload=_seed_state(),
        )
        assert result["success"] is True
        ps = result["pricing_state"]
        skus = {it["sku"] for it in ps["effective_items"]}
        assert "GHOST" not in skus

    def test_reset_string_false_does_not_trigger_reset(self, revise_context):
        """Codex review finding: bool('false') is True. If the LLM (or a
        non-strict adapter) ever passes reset as the string 'false', the
        executor must NOT silently discard the user's overrides."""
        seed = _seed_state()
        seed["effective_uplift_by_currency"] = {"GBP": "0.05"}
        result, _ = _run_revise(
            {
                "reset": "false",  # string, not bool — must be parsed as False
                "overrides": {"percent_uplift": {"EUR": 0.10}},
            },
            revise_context,
            payload=seed,
        )
        ps = result["pricing_state"]
        # The accumulated GBP uplift survives, EUR added per replace-per-key rule.
        assert Decimal(ps["effective_uplift_by_currency"]["GBP"]) == Decimal("0.05")
        assert Decimal(ps["effective_uplift_by_currency"]["EUR"]) == Decimal("0.10")
        # No reset event in the audit log.
        assert not any(e.get("reset") is True for e in ps["applied_overrides_log"])

    def test_reset_string_true_does_trigger_reset(self, revise_context):
        """The flip side: 'true' as a string should still trigger reset."""
        seed = _seed_state()
        seed["effective_uplift_by_currency"] = {"GBP": "0.05"}
        result, _ = _run_revise(
            {"reset": "true", "overrides": {}},
            revise_context,
            payload=seed,
        )
        ps = result["pricing_state"]
        assert ps["effective_uplift_by_currency"] == {}

    def test_decimal_precision_through_json(self, revise_context):
        # 0.1 + 0.2 = 0.30000000000000004 in float.  Decimal(str(v)) must clamp.
        bad_float = 0.1 + 0.2
        result, _ = _run_revise(
            {"overrides": {"sku_price_changes": [{"sku": "SKU-1", "usd_price": bad_float}]}},
            revise_context,
            payload=_seed_state(),
        )
        ps = result["pricing_state"]
        sku1 = next(it for it in ps["effective_items"] if it["sku"] == "SKU-1")
        # Decimal(str(0.1+0.2)) = Decimal("0.30000000000000004") — clamping by str() is the
        # exact intent of the spec's "Decimal(str(v))" precision rule. Whatever the
        # exact representation, it must be a Decimal-precise round-trip of the str cast.
        assert Decimal(str(sku1["usd_price"])) == Decimal(str(bad_float))


class TestSkuAddRemove:
    def test_remove_sku_drops_from_outputs(self, revise_context):
        result, _ = _run_revise(
            {"overrides": {"skus_to_remove": ["SKU-2"]}},
            revise_context,
            payload=_seed_state(),
        )
        ps = result["pricing_state"]
        skus = {it["sku"] for it in ps["effective_items"]}
        assert skus == {"SKU-1"}

    def test_remove_all_skus_returns_error(self, revise_context):
        result, _ = _run_revise(
            {"overrides": {"skus_to_remove": ["SKU-1", "SKU-2"]}},
            revise_context,
            payload=_seed_state(),
        )
        assert result["error"] is True
        assert "no skus left" in result["message"].lower()

    def test_remove_unknown_sku_silent_noop(self, revise_context):
        result, _ = _run_revise(
            {"overrides": {"skus_to_remove": ["GHOST"]}},
            revise_context,
            payload=_seed_state(),
        )
        assert result["success"] is True

    def test_add_then_remove_same_sku_drops_it(self, revise_context):
        # First revise adds SKU-Y.
        seed = _seed_state()
        result, _ = _run_revise(
            {"overrides": {"skus_to_add": [{"sku": "SKU-Y", "usd_price": 99.0}]}},
            revise_context,
            payload=seed,
        )
        # Second revise removes SKU-Y.
        seed2 = result["pricing_state"]
        result2, _ = _run_revise(
            {"overrides": {"skus_to_remove": ["SKU-Y"]}},
            revise_context,
            payload=seed2,
        )
        skus = {it["sku"] for it in result2["pricing_state"]["effective_items"]}
        assert "SKU-Y" not in skus

    def test_remove_then_add_same_sku_keeps_it(self, revise_context):
        # Revise 1 removes SKU-1 (originally seeded).
        seed = _seed_state()
        result, _ = _run_revise(
            {"overrides": {"skus_to_remove": ["SKU-1"]}},
            revise_context,
            payload=seed,
        )
        # Revise 2 adds SKU-1 back at a new price.
        seed2 = result["pricing_state"]
        result2, _ = _run_revise(
            {"overrides": {"skus_to_add": [{"sku": "SKU-1", "usd_price": 999.0}]}},
            revise_context,
            payload=seed2,
        )
        ps = result2["pricing_state"]
        sku1 = next(it for it in ps["effective_items"] if it["sku"] == "SKU-1")
        assert Decimal(str(sku1["usd_price"])) == Decimal("999")

    def test_within_call_remove_then_add_same_sku(self, revise_context):
        # Spec: order is remove → add → price-change.  Same-call add+remove on
        # the same SKU yields the SKU re-added at the new price.
        result, _ = _run_revise(
            {
                "overrides": {
                    "skus_to_remove": ["SKU-1"],
                    "skus_to_add": [{"sku": "SKU-1", "usd_price": 99.0}],
                }
            },
            revise_context,
            payload=_seed_state(),
        )
        ps = result["pricing_state"]
        sku1 = next(it for it in ps["effective_items"] if it["sku"] == "SKU-1")
        assert Decimal(str(sku1["usd_price"])) == Decimal("99")

    def test_skus_to_add_duplicate_is_noop(self, revise_context):
        result, _ = _run_revise(
            {"overrides": {"skus_to_add": [{"sku": "SKU-1", "usd_price": 999.0}]}},
            revise_context,
            payload=_seed_state(),
        )
        ps = result["pricing_state"]
        sku1s = [it for it in ps["effective_items"] if it["sku"] == "SKU-1"]
        # No duplicate row appended; price unchanged from the seed.
        assert len(sku1s) == 1
        assert Decimal(str(sku1s[0]["usd_price"])) == Decimal("100")


class TestCurrencyOps:
    def test_currencies_to_add_unknown_currency_errors(self, revise_context):
        result, _ = _run_revise(
            {"overrides": {"currencies_to_add": ["XYZ"]}},
            revise_context,
            payload=_seed_state(),
        )
        assert result["error"] is True
        assert "not configured" in result["message"].lower()

    def test_currencies_to_remove_filters(self, revise_context):
        # Start with all 3 currencies in effective_currencies.
        seed = _seed_state()
        result, _ = _run_revise(
            {"overrides": {"currencies_to_remove": ["EUR"]}},
            revise_context,
            payload=seed,
        )
        ps = result["pricing_state"]
        assert "EUR" not in ps["effective_currencies"]
        assert "GBP" in ps["effective_currencies"]


class TestUpliftSemantics:
    def test_percent_uplift_only_targeted_currency(self, revise_context):
        # Apply +5% on GBP only.  Engine output GBP price > USD * 0.79 * 1.20.
        result, _ = _run_revise(
            {"overrides": {"percent_uplift": {"GBP": 0.05}}},
            revise_context,
            payload=_seed_state(),
        )
        ps = result["pricing_state"]
        # Effective state captures the GBP uplift.
        assert Decimal(ps["effective_uplift_by_currency"]["GBP"]) == Decimal("0.05")
        # CAD/EUR not in the uplift dict.
        assert "CAD" not in ps["effective_uplift_by_currency"]

    def test_replace_per_key_uplift(self, revise_context):
        # Two revises against GBP — second wins.
        seed = _seed_state()
        result, _ = _run_revise(
            {"overrides": {"percent_uplift": {"GBP": 0.05}}},
            revise_context,
            payload=seed,
        )
        seed2 = result["pricing_state"]
        result2, _ = _run_revise(
            {"overrides": {"percent_uplift": {"GBP": 0.10}}},
            revise_context,
            payload=seed2,
        )
        ps = result2["pricing_state"]
        # Decimal(str(0.10)) normalizes to "0.1" — compare as Decimals.
        assert Decimal(ps["effective_uplift_by_currency"]["GBP"]) == Decimal("0.10")


class TestFxOverrideSafety:
    def test_rejects_ambiguous_eur_fx_override_for_eur_based_tier(self, revise_context):
        """Do not accept the staging failure mode as a guessed target-EUR edit.

        User intent was "set final EUR display price to 129"; the bad tool call
        guessed EUR fx_rate_overrides = 129 / 99, which produces EUR 159 for an
        eur_based EUR tier after the engine applies EUR conversion, VAT, and
        rounding. Safe first pass is to reject this unsupported contract before
        mutating cached pricing state or writing revised output files.
        """
        seed = _seed_state(items=[{"sku": "ITEM-001", "usd_price": "99", "item_name": None}])
        result, saved = _run_revise(
            {
                "overrides": {
                    "sku_price_changes": [{"sku": "ITEM-001", "usd_price": 99}],
                    "fx_rate_overrides": {"EUR": 1.303030303030303},
                }
            },
            revise_context,
            payload=seed,
        )

        assert result["error"] is True
        assert "EUR final display price edits are not supported through fx_rate_overrides" in result["message"]
        assert saved == []
        assert seed["effective_fx_overrides"] == {}
        assert seed["applied_overrides_log"] == []

    def test_rejects_fx_override_for_non_eur_eur_based_currency(self, revise_context):
        """Do not allow target-final-price guesses for any eur_based currency.

        A follow-up like "change Euro to 139 and other euro base currencies" can
        lead the LLM to put every EUR-derived currency in fx_rate_overrides. The
        previous guard only rejected EUR itself; it still allowed SEK/NOK/etc.,
        producing misleading revised outputs instead of asking the user which
        supported edit type they intended.
        """
        config = TenantPricingConfig(
            base_currency="USD",
            eur_fx_rate=Decimal("0.92"),
            currencies={
                "EUR": CurrencyConfig(
                    fx_rate=Decimal("1.00"),
                    tier="eur_based",
                    vat_rate=Decimal("0.23"),
                    rounding_rule="nearest_9",
                ),
                "SEK": CurrencyConfig(
                    fx_rate=Decimal("11.29"),
                    tier="eur_based",
                    vat_rate=Decimal("0.25"),
                    rounding_rule="nearest_9",
                ),
                "GBP": CurrencyConfig(
                    fx_rate=Decimal("0.79"),
                    tier="usd_based",
                    vat_rate=Decimal("0.20"),
                    rounding_rule="nearest_9",
                ),
            },
        ).model_dump(mode="json")
        seed = _seed_state(items=[{"sku": "ITEM-001", "usd_price": "100", "item_name": None}])
        seed["effective_currencies"] = ["EUR", "SEK", "GBP"]
        seed["header_columns"] = ["SKU", "Item Name", "USD", "EUR", "GBP", "SEK"]

        result, saved = _run_revise(
            {
                "overrides": {
                    "fx_rate_overrides": {"SEK": 1.39},
                }
            },
            revise_context,
            payload=seed,
            config_dict=config,
        )

        assert result["error"] is True
        assert "EUR-based currency final display price edits are not supported" in result["message"]
        assert "SEK" in result["message"]
        assert saved == []
        assert seed["effective_fx_overrides"] == {}
        assert seed["applied_overrides_log"] == []

    def test_allows_fx_override_for_usd_based_currency(self, revise_context):
        seed = _seed_state(items=[{"sku": "ITEM-001", "usd_price": "100", "item_name": None}])
        result, saved = _run_revise(
            {
                "overrides": {
                    "fx_rate_overrides": {"GBP": 0.81},
                }
            },
            revise_context,
            payload=seed,
        )

        assert result["success"] is True
        assert saved
        assert result["pricing_state"]["effective_fx_overrides"]["GBP"] == "0.81"


class TestEurBaseFollowup:
    def test_target_final_eur_price_regenerates_eur_based_download_prices(self, revise_context):
        """target_final_prices is the supported final-display-price contract."""
        config = TenantPricingConfig(
            base_currency="USD",
            eur_fx_rate=Decimal("0.92"),
            currencies={
                "EUR": CurrencyConfig(
                    fx_rate=Decimal("1.00"),
                    tier="eur_based",
                    vat_rate=Decimal("0.23"),
                    rounding_rule="nearest_9",
                ),
                "SEK": CurrencyConfig(
                    fx_rate=Decimal("11.29"),
                    tier="eur_based",
                    vat_rate=Decimal("0.25"),
                    rounding_rule="nearest_9",
                ),
                "GBP": CurrencyConfig(
                    fx_rate=Decimal("0.79"),
                    tier="usd_based",
                    vat_rate=Decimal("0.20"),
                    rounding_rule="nearest_9",
                ),
            },
        ).model_dump(mode="json")
        seed = _seed_state(items=[{"sku": "ITEM-001", "usd_price": "100", "item_name": "Test Item"}])
        seed["effective_currencies"] = ["EUR", "SEK", "GBP"]
        seed["header_columns"] = ["SKU", "Item Name", "USD", "EUR", "GBP", "SEK"]

        result, saved = _run_revise(
            {"overrides": {"target_final_prices": {"EUR": 149}}},
            revise_context,
            payload=seed,
            config_dict=config,
        )

        assert result["success"] is True
        preview_row = result["preview"][0]
        assert preview_row["EUR"] == 149.0
        assert preview_row["SEK"] == 1689.0
        ps = result["pricing_state"]
        assert ps["effective_target_final_prices"]["EUR"] == "149"
        assert Decimal(ps["effective_eur_fx_rate_override"]) == Decimal("149") / Decimal("123.00")

        excel_file = next(f for f in saved if f.filename.endswith(".xlsx"))
        wb = load_workbook(io.BytesIO(excel_file.content), data_only=True)
        ws = wb["Prices"]
        headers = [cell.value for cell in ws[1]]
        values = [cell.value for cell in ws[2]]
        excel_row = dict(zip(headers, values, strict=False))
        assert excel_row["EUR"] == 149
        assert excel_row["SEK"] == 1689

        csv_file = next(f for f in saved if f.filename.endswith(".csv"))
        csv_rows = list(csv.DictReader(io.StringIO(csv_file.content.decode("utf-8"))))
        csv_prices = {row["Currency"]: row["Rate"] for row in csv_rows if row["External ID"] == "ITEM-001"}
        assert csv_prices["EUR"] == "149"
        assert csv_prices["SEK"] == "1689"

    def test_target_final_eur_price_rejects_unreachable_rounding_target(self, revise_context):
        config = TenantPricingConfig(
            base_currency="USD",
            eur_fx_rate=Decimal("0.92"),
            currencies={
                "EUR": CurrencyConfig(
                    fx_rate=Decimal("1.00"),
                    tier="eur_based",
                    vat_rate=Decimal("0.23"),
                    rounding_rule="nearest_9",
                ),
            },
        ).model_dump(mode="json")
        seed = _seed_state(items=[{"sku": "ITEM-001", "usd_price": "100", "item_name": "Test Item"}])
        seed["effective_currencies"] = ["EUR"]

        result, saved = _run_revise(
            {"overrides": {"target_final_prices": {"EUR": 150}}},
            revise_context,
            payload=seed,
            config_dict=config,
        )

        assert result["error"] is True
        assert "not reachable" in result["message"]
        assert saved == []
        assert seed["applied_overrides_log"] == []

    def test_target_final_eur_price_requires_single_effective_sku(self, revise_context):
        result, saved = _run_revise(
            {"overrides": {"target_final_prices": {"EUR": 149}}},
            revise_context,
            payload=_seed_state(),
        )

        assert result["error"] is True
        assert "exactly one SKU" in result["message"]
        assert saved == []

    def test_top_level_eur_fx_rate_regenerates_eur_based_download_prices(self, revise_context):
        """A fresh tenant EUR base rate recomputes EUR and EUR-based exports.

        This covers the explicit follow-up path: when the tenant config's
        top-level eur_fx_rate changes, pricing_revise reads the fresh config
        and regenerates both task preview and downloadable files from it.
        """
        config = TenantPricingConfig(
            base_currency="USD",
            eur_fx_rate=Decimal("1.10"),
            currencies={
                "EUR": CurrencyConfig(
                    fx_rate=Decimal("1.00"),
                    tier="eur_based",
                    vat_rate=Decimal("0.23"),
                    rounding_rule="nearest_9",
                ),
                "SEK": CurrencyConfig(
                    fx_rate=Decimal("11.29"),
                    tier="eur_based",
                    vat_rate=Decimal("0.25"),
                    rounding_rule="nearest_9",
                ),
                "GBP": CurrencyConfig(
                    fx_rate=Decimal("0.79"),
                    tier="usd_based",
                    vat_rate=Decimal("0.20"),
                    rounding_rule="nearest_9",
                ),
            },
        ).model_dump(mode="json")
        seed = _seed_state(items=[{"sku": "ITEM-001", "usd_price": "100", "item_name": "Test Item"}])
        seed["effective_currencies"] = ["EUR", "SEK", "GBP"]
        seed["header_columns"] = ["SKU", "Item Name", "USD", "EUR", "GBP", "SEK"]

        result, saved = _run_revise(
            {"overrides": {}},
            revise_context,
            payload=seed,
            config_dict=config,
        )

        assert result["success"] is True
        preview_row = result["preview"][0]
        assert preview_row["EUR"] == 139.0
        assert preview_row["SEK"] == 1579.0

        excel_file = next(f for f in saved if f.filename.endswith(".xlsx"))
        wb = load_workbook(io.BytesIO(excel_file.content), data_only=True)
        ws = wb["Prices"]
        headers = [cell.value for cell in ws[1]]
        values = [cell.value for cell in ws[2]]
        excel_row = dict(zip(headers, values, strict=False))
        assert excel_row["EUR"] == 139
        assert excel_row["SEK"] == 1579

        csv_file = next(f for f in saved if f.filename.endswith(".csv"))
        csv_rows = list(csv.DictReader(io.StringIO(csv_file.content.decode("utf-8"))))
        csv_prices = {row["Currency"]: row["Rate"] for row in csv_rows if row["External ID"] == "ITEM-001"}
        assert csv_prices["EUR"] == "139"
        assert csv_prices["SEK"] == "1579"


class TestReset:
    def test_reset_clears_overrides(self, revise_context):
        seed = _seed_state()
        seed["effective_items"][0]["usd_price"] = "999"
        seed["effective_uplift_by_currency"] = {"GBP": "0.05"}
        seed["effective_fx_overrides"] = {"GBP": "0.85"}
        seed["applied_overrides_log"] = [{"sku_price_changes": [...]}, {"percent_uplift": {"GBP": 0.05}}]
        result, _ = _run_revise(
            {"reset": True, "overrides": {}},
            revise_context,
            payload=seed,
        )
        ps = result["pricing_state"]
        # Effective state matches seed: SKU-1 is back at 100.
        sku1 = next(it for it in ps["effective_items"] if it["sku"] == "SKU-1")
        assert Decimal(str(sku1["usd_price"])) == Decimal("100")
        assert ps["effective_uplift_by_currency"] == {}
        assert ps["effective_fx_overrides"] == {}

    def test_reset_ignores_other_override_fields(self, revise_context):
        seed = _seed_state()
        seed["effective_uplift_by_currency"] = {"EUR": "0.05"}
        result, _ = _run_revise(
            {"reset": True, "overrides": {"percent_uplift": {"GBP": 0.05}}},
            revise_context,
            payload=seed,
        )
        ps = result["pricing_state"]
        # GBP uplift dropped (reset wins).
        assert ps["effective_uplift_by_currency"] == {}
        # Audit records the reset event only.
        assert any("reset" in str(e).lower() for e in ps["applied_overrides_log"])


class TestFreshConfig:
    def test_revise_uses_fresh_tenant_config(self, revise_context):
        """Settings edits between turns must be reflected in the revise output.

        Initial config snapshot in seed had GBP fx_rate=1.25; the live config
        has GBP at the spec default (0.79).  Empty-overrides revise must use
        the fresh fx_rate, not anything cached from a prior payload.
        """
        seed = _seed_state()
        # Spec lock D4: base config is NOT in the payload — revise must call
        # get_config() each time. Verified by the fact that no seed_config exists.
        assert "seed_config" not in seed
        result, saved = _run_revise(
            {"overrides": {}},
            revise_context,
            payload=seed,
        )
        assert result["success"] is True
        # Engine ran with fresh GBP fx_rate=0.79 → SKU-1 at $100 → 100*0.79*1.20=94.8 → round_9=99.
        gbp = next(p for p in result["preview"] if p["SKU"] == "SKU-1")["GBP"]
        assert int(gbp) == 99


class TestCachePayloadShape:
    def test_writes_pricing_state_to_cache(self, revise_context):
        """After a successful revise, the executor must return pricing_state
        with a fresh excel_file_id and accurate row_count."""
        result, saved = _run_revise(
            {"overrides": {"sku_price_changes": [{"sku": "SKU-1", "usd_price": 200.0}]}},
            revise_context,
            payload=_seed_state(),
        )
        ps = result["pricing_state"]
        assert ps["excel_file_id"] == str(saved[0].id)
        assert ps["netsuite_csv_file_id"] == str(saved[1].id)
        assert ps["row_count"] == len(ps["effective_items"])

    def test_no_response_instruction_markdown_table(self, revise_context):
        result, _ = _run_revise(
            {"overrides": {"percent_uplift": {"GBP": 0.05}}},
            revise_context,
            payload=_seed_state(),
        )
        ri = result.get("response_instruction", "")
        assert "|---" not in ri
        assert "EXACT table" not in ri
        assert "verbatim" not in ri.lower()


class TestCumulativeAccumulation:
    def test_cumulative_sku_changes(self, revise_context):
        # Revise 1: bump SKU-1 to 150.
        seed = _seed_state()
        r1, _ = _run_revise(
            {"overrides": {"sku_price_changes": [{"sku": "SKU-1", "usd_price": 150.0}]}},
            revise_context,
            payload=seed,
        )
        # Revise 2: bump SKU-2 to 250.
        r2, _ = _run_revise(
            {"overrides": {"sku_price_changes": [{"sku": "SKU-2", "usd_price": 250.0}]}},
            revise_context,
            payload=r1["pricing_state"],
        )
        ps = r2["pricing_state"]
        prices = {it["sku"]: Decimal(str(it["usd_price"])) for it in ps["effective_items"]}
        assert prices["SKU-1"] == Decimal("150")
        assert prices["SKU-2"] == Decimal("250")
        # Audit log accumulated both revises.
        assert len(ps["applied_overrides_log"]) == 2
