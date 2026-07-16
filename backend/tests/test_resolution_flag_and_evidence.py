"""recon_resolution_ui flag registration + evidence pack Proposals sheet.

No existing DB-backed test exercises the download_evidence endpoint (only a
route-registration string check in test_reconciliation_api.py and an MCP tool
wrapper test that never touches generate_excel). So the endpoint-wiring
assertion below is a new test, following the pattern already used by
test_resolution_summary_api.py / test_resolution_group_actions.py of calling
the FastAPI endpoint function directly against seeded DB fixtures.
"""

import io
from decimal import Decimal

from openpyxl import Workbook, load_workbook

from app.api.v1.reconciliation import download_evidence, plan_resolutions
from app.services.feature_flag_service import DEFAULT_FLAGS
from app.services.reconciliation.evidence_service import EvidencePackGenerator
from tests.conftest import (
    create_test_netsuite_posting,
    create_test_recon_result,
    create_test_recon_run,
    create_test_user,
    enable_feature_flag,
)


def test_recon_resolution_ui_flag_registered_default_off():
    assert DEFAULT_FLAGS.get("recon_resolution_ui") is False


def test_proposals_sheet_writer():
    gen = EvidencePackGenerator()
    proposals = [
        {
            "group_key": "fees:book_fee_line:deposit",
            "root_cause": "fees",
            "action": "book_fee_line",
            "booking_vehicle": "deposit",
            "status": "proposed",
            "narrative": "Stripe processing fee — book as a fee line.",
            "proposed_amount": Decimal("3.20"),
            "currency": "USD",
            "above_materiality": False,
            "source": "planner",
            "order_reference": "R946866359",
            "stripe_charge_id": "ch_3Nxxx",
            "netsuite_internal_id": "12345",
        }
    ]
    wb = Workbook()
    gen._write_proposals_sheet(wb, proposals)
    assert "Proposals" in wb.sheetnames
    ws = wb["Proposals"]
    headers = [c.value for c in ws[1]]
    assert "Group" in headers and "Action" in headers and "Narrative" in headers
    assert "Order Ref" in headers and "Stripe Charge" in headers and "NetSuite ID" in headers
    assert ws.max_row == 2  # header + one proposal row
    row = dict(zip(headers, [c.value for c in ws[2]]))
    assert row["Order Ref"] == "R946866359"
    assert row["Stripe Charge"] == "ch_3Nxxx"
    assert row["NetSuite ID"] == "12345"


async def test_download_evidence_includes_proposals_sheet(db, tenant_a):
    """Endpoint-level: a run with a live (non-superseded/rejected) proposal
    must produce an evidence pack whose workbook has a "Proposals" sheet."""
    user, _ = await create_test_user(db, tenant_a)
    await enable_feature_flag(db, tenant_a.id, "recon_resolution_ui")
    run = await create_test_recon_run(db, tenant_a.id, status="completed")
    await create_test_recon_result(
        db,
        tenant_a.id,
        run.id,
        status="pending",
        bucket="auto_classifications",
        match_type="deterministic",
        variance_type="fees",
        variance_amount=Decimal("120.00"),
        stripe_amount=Decimal("10000"),
        netsuite_amount=Decimal("9880.00"),
        evidence={"charge_source_id": "ch_evp1", "order_reference": "R1"},
    )
    await db.flush()
    await plan_resolutions(str(run.id), user=user, db=db)

    response = await download_evidence(str(run.id), user=user, db=db)
    body = b"".join([chunk async for chunk in response.body_iterator])
    wb = load_workbook(io.BytesIO(body))
    assert "Proposals" in wb.sheetnames


async def test_download_evidence_proposals_sheet_has_identifiers(db, tenant_a):
    """A3: the Proposals sheet's identifier columns are populated end-to-end
    from the endpoint's dict-ify — order ref + charge id always, NetSuite id
    only for a proposal whose result has a linked deposit."""
    user, _ = await create_test_user(db, tenant_a)
    await enable_feature_flag(db, tenant_a.id, "recon_resolution_ui")
    run = await create_test_recon_run(db, tenant_a.id, status="completed")
    posting = await create_test_netsuite_posting(db, tenant_a.id, netsuite_internal_id="55555")
    await create_test_recon_result(
        db,
        tenant_a.id,
        run.id,
        status="pending",
        bucket="auto_classifications",
        match_type="deterministic",
        variance_type="fees",
        variance_amount=Decimal("120.00"),
        stripe_amount=Decimal("10000"),
        netsuite_amount=Decimal("9880.00"),
        evidence={"charge_source_id": "ch_evp2", "order_reference": "R2"},
        deposit_id=posting.id,
    )
    await db.flush()
    await plan_resolutions(str(run.id), user=user, db=db)

    response = await download_evidence(str(run.id), user=user, db=db)
    body = b"".join([chunk async for chunk in response.body_iterator])
    wb = load_workbook(io.BytesIO(body))
    ws = wb["Proposals"]
    headers = [c.value for c in ws[1]]
    row = dict(zip(headers, [c.value for c in ws[2]]))
    assert row["Order Ref"] == "R2"
    assert row["Stripe Charge"] == "ch_evp2"
    assert row["NetSuite ID"] == "55555"
