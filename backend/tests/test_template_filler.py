"""Tests for template filler — column detection, Excel fill, default output, NetSuite CSV."""

from decimal import Decimal

import pytest
from openpyxl import Workbook

from app.schemas.pricing import CurrencyResult, PricingOutput
from app.services.template_filler import TemplateFiller


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_workbook(headers, rows=None):
    wb = Workbook()
    ws = wb.active
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    for ri, row in enumerate(rows or [], 2):
        for ci, v in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=v)
    return wb


def _make_output(sku, usd, currencies):
    results = {}
    for code, price in currencies.items():
        results[code] = CurrencyResult(
            currency=code,
            fx_rate=Decimal("1"),
            tier="usd_based",
            converted_amount=Decimal(str(price)),
            vat_rate=None,
            vat_amount=None,
            pre_round_amount=Decimal(str(price)),
            final_price=Decimal(str(price)),
            rounding_rule="nearest_9",
        )
    return PricingOutput(
        sku=sku, item_name=sku, usd_price=Decimal(str(usd)), results=results
    )


# ---------------------------------------------------------------------------
# Class 1: TestColumnDetection
# ---------------------------------------------------------------------------

class TestColumnDetection:
    """Test TemplateFiller().detect_columns() on various header patterns."""

    def setup_method(self):
        self.filler = TemplateFiller()

    def test_exact_currency_codes(self):
        wb = _make_workbook(["SKU", "USD", "GBP", "EUR", "JPY"])
        mapping = self.filler.detect_columns(wb)
        assert mapping.currency_cols == {2: "GBP", 3: "EUR", 4: "JPY"}

    def test_currency_with_price_suffix(self):
        wb = _make_workbook(["SKU", "USD Price", "GBP Price", "EUR Price"])
        mapping = self.filler.detect_columns(wb)
        assert mapping.currency_cols == {2: "GBP", 3: "EUR"}

    def test_currency_with_price_prefix(self):
        wb = _make_workbook(["Item", "Price USD", "Price GBP", "Price EUR"])
        mapping = self.filler.detect_columns(wb)
        assert mapping.currency_cols == {2: "GBP", 3: "EUR"}

    def test_currency_in_parentheses(self):
        wb = _make_workbook(["SKU", "Base (USD)", "Local (GBP)", "Local (EUR)"])
        mapping = self.filler.detect_columns(wb)
        assert mapping.currency_cols == {2: "GBP", 3: "EUR"}

    def test_full_currency_names(self):
        wb = _make_workbook(["SKU", "US Dollar", "British Pound", "Euro", "Japanese Yen"])
        mapping = self.filler.detect_columns(wb)
        assert mapping.currency_cols == {2: "GBP", 3: "EUR", 4: "JPY"}

    def test_case_insensitive(self):
        wb = _make_workbook(["sku", "usd", "gbp", "eur"])
        mapping = self.filler.detect_columns(wb)
        assert mapping.currency_cols == {2: "GBP", 3: "EUR"}

    def test_extra_non_currency_columns(self):
        wb = _make_workbook(["SKU", "Item Name", "Category", "USD", "GBP", "Notes"])
        mapping = self.filler.detect_columns(wb)
        assert mapping.currency_cols == {4: "GBP"}

    def test_no_currency_columns(self):
        wb = _make_workbook(["SKU", "Item Name", "USD Price"])
        mapping = self.filler.detect_columns(wb)
        # Only USD found = source, no targets
        assert mapping.currency_cols == {}

    def test_custom_aliases(self):
        wb = _make_workbook(["SKU", "USD", "Sterling", "Franc"])
        filler = TemplateFiller(aliases={"Sterling": "GBP", "Franc": "CHF"})
        mapping = filler.detect_columns(wb)
        assert mapping.currency_cols == {2: "GBP", 3: "CHF"}


# ---------------------------------------------------------------------------
# Class 2: TestSKUColumnDetection
# ---------------------------------------------------------------------------

class TestSKUColumnDetection:
    """Test detect_sku_column() for various header patterns."""

    def setup_method(self):
        self.filler = TemplateFiller()

    def test_sku_header(self):
        wb = _make_workbook(["SKU", "USD", "GBP"])
        mapping = self.filler.detect_columns(wb)
        assert mapping.sku_col == 0

    def test_item_header(self):
        wb = _make_workbook(["Item", "USD", "GBP"])
        mapping = self.filler.detect_columns(wb)
        assert mapping.sku_col == 0

    def test_item_id_header(self):
        wb = _make_workbook(["Item ID", "USD", "GBP"])
        mapping = self.filler.detect_columns(wb)
        assert mapping.sku_col == 0

    def test_external_id_header(self):
        wb = _make_workbook(["External ID", "USD", "GBP"])
        mapping = self.filler.detect_columns(wb)
        assert mapping.sku_col == 0

    def test_product_code_header(self):
        wb = _make_workbook(["Product Code", "USD", "GBP"])
        mapping = self.filler.detect_columns(wb)
        assert mapping.sku_col == 0

    def test_sku_not_first(self):
        wb = _make_workbook(["Category", "SKU", "USD"])
        mapping = self.filler.detect_columns(wb)
        assert mapping.sku_col == 1

    def test_no_sku_column(self):
        wb = _make_workbook(["Price", "GBP"])
        with pytest.raises(ValueError, match="(?i)sku"):
            self.filler.detect_columns(wb)


# ---------------------------------------------------------------------------
# Class 3: TestPriceColumnDetection
# ---------------------------------------------------------------------------

class TestPriceColumnDetection:
    """Test detect_price_column() for USD/base price detection."""

    def setup_method(self):
        self.filler = TemplateFiller()

    def test_usd_header(self):
        wb = _make_workbook(["SKU", "USD", "GBP"])
        mapping = self.filler.detect_columns(wb)
        assert mapping.price_col == 1

    def test_usd_price_header(self):
        wb = _make_workbook(["SKU", "USD Price", "GBP"])
        mapping = self.filler.detect_columns(wb)
        assert mapping.price_col == 1

    def test_base_price_header(self):
        wb = _make_workbook(["SKU", "Base Price", "GBP"])
        mapping = self.filler.detect_columns(wb)
        assert mapping.price_col == 1

    def test_price_usd_header(self):
        wb = _make_workbook(["SKU", "Price USD", "GBP"])
        mapping = self.filler.detect_columns(wb)
        assert mapping.price_col == 1

    def test_price_not_second(self):
        wb = _make_workbook(["SKU", "Name", "USD Price"])
        mapping = self.filler.detect_columns(wb)
        assert mapping.price_col == 2

    def test_no_price_column(self):
        wb = _make_workbook(["SKU", "GBP", "EUR"])
        with pytest.raises(ValueError, match="(?i)price|usd"):
            self.filler.detect_columns(wb)


# ---------------------------------------------------------------------------
# Class 4: TestTemplateFill
# ---------------------------------------------------------------------------

class TestTemplateFill:
    """Test fill() with actual Workbook objects."""

    def setup_method(self):
        self.filler = TemplateFiller()
        self.outputs = [
            _make_output("FRANVD0009", 1659, {"GBP": "1329", "EUR": "1499"}),
            _make_output("FRANKK00A1", 199, {"GBP": "159", "EUR": "179"}),
        ]

    def test_fill_basic(self):
        wb = _make_workbook(
            ["SKU", "USD", "GBP", "EUR"],
            [
                ["FRANVD0009", 1659, None, None],
                ["FRANKK00A1", 199, None, None],
            ],
        )
        mapping = self.filler.detect_columns(wb)
        self.filler.fill(wb, self.outputs, mapping)

        ws = wb.active
        assert ws.cell(row=2, column=3).value == Decimal("1329")
        assert ws.cell(row=2, column=4).value == Decimal("1499")
        assert ws.cell(row=3, column=3).value == Decimal("159")
        assert ws.cell(row=3, column=4).value == Decimal("179")

    def test_fill_preserves_existing_data(self):
        wb = _make_workbook(
            ["SKU", "USD", "GBP", "EUR", "Notes"],
            [
                ["FRANVD0009", 1659, None, None, "Important note"],
                ["FRANKK00A1", 199, None, None, "Another note"],
            ],
        )
        mapping = self.filler.detect_columns(wb)
        self.filler.fill(wb, self.outputs, mapping)

        ws = wb.active
        assert ws.cell(row=2, column=5).value == "Important note"
        assert ws.cell(row=3, column=5).value == "Another note"

    def test_fill_skips_unknown_currencies(self):
        wb = _make_workbook(
            ["SKU", "USD", "GBP", "XYZ"],
            [
                ["FRANVD0009", 1659, None, "existing"],
            ],
        )
        mapping = self.filler.detect_columns(wb)
        self.filler.fill(wb, self.outputs, mapping)

        ws = wb.active
        # GBP filled
        assert ws.cell(row=2, column=3).value == Decimal("1329")
        # XYZ not a known currency, col not in mapping — untouched
        assert ws.cell(row=2, column=4).value == "existing"

    def test_fill_partial_sku_match(self):
        wb = _make_workbook(
            ["SKU", "USD", "GBP", "EUR"],
            [
                ["FRANVD0009", 1659, None, None],
                ["FRANKK00A1", 199, None, None],
                ["UNKNOWN_SKU", 500, None, None],
            ],
        )
        mapping = self.filler.detect_columns(wb)
        self.filler.fill(wb, self.outputs, mapping)

        ws = wb.active
        # First two filled
        assert ws.cell(row=2, column=3).value == Decimal("1329")
        assert ws.cell(row=3, column=3).value == Decimal("159")
        # Third row — no match, remains empty
        assert ws.cell(row=4, column=3).value is None


# ---------------------------------------------------------------------------
# Class 5: TestDefaultOutput
# ---------------------------------------------------------------------------

class TestDefaultOutput:
    """Test generate_default_output() creates proper multi-sheet workbook."""

    def setup_method(self):
        self.filler = TemplateFiller()
        self.outputs = [
            _make_output("FRANVD0009", 1659, {"GBP": "1329", "EUR": "1499"}),
            _make_output("FRANKK00A1", 199, {"GBP": "159", "EUR": "179"}),
        ]

    def test_generate_default_prices_sheet(self):
        wb = self.filler.generate_default_output(self.outputs)

        assert "Prices" in wb.sheetnames
        ws = wb["Prices"]
        # Headers: SKU, Item Name, USD, then currency columns
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "SKU" in headers
        assert "Item Name" in headers
        assert "USD" in headers
        # Should have currency columns
        currency_headers = [h for h in headers if h in ("GBP", "EUR")]
        assert len(currency_headers) >= 2

        # Data rows
        assert ws.cell(row=2, column=1).value == "FRANVD0009"

    def test_generate_default_audit_sheet(self):
        wb = self.filler.generate_default_output(self.outputs)

        assert "Conversion Details" in wb.sheetnames
        ws = wb["Conversion Details"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "SKU" in headers
        assert "Currency" in headers
        # Should have per-SKU-per-currency rows (2 SKUs * 2 currencies = 4 rows)
        assert ws.max_row >= 5  # header + 4 data rows

    def test_generate_default_config_sheet(self):
        wb = self.filler.generate_default_output(self.outputs)

        assert "Config Snapshot" in wb.sheetnames
        ws = wb["Config Snapshot"]
        # Should have some content
        assert ws.max_row >= 1


# ---------------------------------------------------------------------------
# Class 6: TestNetSuiteCSV
# ---------------------------------------------------------------------------

class TestNetSuiteCSV:
    """Test generate_netsuite_csv() output format."""

    def setup_method(self):
        self.filler = TemplateFiller()
        self.outputs = [
            _make_output("FRANVD0009", 1659, {"GBP": "1329", "EUR": "1499"}),
        ]

    def test_netsuite_csv_format(self):
        csv_str = self.filler.generate_netsuite_csv(self.outputs)
        lines = csv_str.strip().split("\n")
        assert lines[0] == "External ID,Item,Price Level,Currency,Rate"

    def test_netsuite_csv_values(self):
        csv_str = self.filler.generate_netsuite_csv(self.outputs)
        lines = csv_str.strip().split("\n")
        # Should have header + 2 data rows (GBP + EUR)
        assert len(lines) == 3
        # Check one row has correct values
        data_lines = lines[1:]
        values_found = []
        for line in data_lines:
            parts = line.split(",")
            assert parts[0] == "FRANVD0009"  # External ID
            assert parts[1] == "FRANVD0009"  # Item
            values_found.append((parts[3], parts[4]))  # Currency, Rate

        currencies_in_csv = {v[0] for v in values_found}
        assert "GBP" in currencies_in_csv
        assert "EUR" in currencies_in_csv
