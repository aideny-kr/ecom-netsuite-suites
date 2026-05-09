"""Step 7 — pricing_to_sheets executor tests.

Read-only consumer: reads the latest pricing cache entry, re-parses the
saved Excel file via TaskFileService.get_file + openpyxl, and writes the
full row set to a new Google Sheet. Connector check FIRST so users get the
right setup-vs-session error.
"""

from __future__ import annotations

import asyncio
import io
import uuid
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from openpyxl import Workbook

from app.mcp.tools.pricing_tools import pricing_to_sheets_execute


def _build_excel_bytes(headers: list[str], rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _payload(excel_file_id: str | None = None, row_count: int = 2) -> dict:
    if excel_file_id is None:
        excel_file_id = str(uuid.uuid4())
    return {
        "seed_items": [],
        "effective_items": [
            {"sku": "SKU-1", "usd_price": "100", "item_name": None},
            {"sku": "SKU-2", "usd_price": "200", "item_name": None},
        ],
        "effective_currencies": ["GBP"],
        "effective_fx_overrides": {},
        "effective_vat_overrides": {},
        "effective_rounding_overrides": {},
        "effective_uplift_by_currency": {},
        "applied_overrides_log": [],
        "excel_file_id": excel_file_id,
        "netsuite_csv_file_id": "file-csv",
        "header_columns": ["SKU", "USD", "GBP"],
        "row_count": row_count,
    }


def _cached(payload: dict | None) -> MagicMock | None:
    if payload is None:
        return None
    cr = MagicMock()
    cr.result_type = "pricing"
    cr.payload = payload
    return cr


def _connector(metadata: dict | None = None) -> MagicMock:
    c = MagicMock()
    c.encrypted_credentials = b"x"
    c.metadata_json = metadata or {}
    return c


@pytest.fixture
def sheets_context():
    return {
        "db": AsyncMock(),
        "tenant_id": uuid.uuid4(),
        "actor_id": uuid.uuid4(),
        "conversation_id": "conv-abc",
    }


def _run(params, ctx, *, connector, payload, excel_bytes=None, file_lookup_raises=None, styling_side_effect=None):
    """Invoke pricing_to_sheets_execute with the standard mock harness.

    Returns (result, create_mock, write_mock, share_mock, styling_mock).
    `styling_side_effect`, if set, is attached to the styling mock so a test
    can simulate a Sheets API failure during styling.
    """
    if excel_bytes is None and payload is not None:
        excel_bytes = _build_excel_bytes(["SKU", "USD", "GBP"], [["SKU-1", 100, 99]])

    task_file = MagicMock()
    task_file.filename = "pricing-revised.xlsx"

    if file_lookup_raises:
        get_file_mock = AsyncMock(side_effect=file_lookup_raises)
    else:
        get_file_mock = AsyncMock(return_value=(task_file, excel_bytes))

    if styling_side_effect is not None:
        styling_mock = AsyncMock(side_effect=styling_side_effect)
    else:
        styling_mock = AsyncMock(return_value={"replies": []})

    with (
        patch(
            "app.mcp.tools.pricing_tools._get_sheets_connector",
            new_callable=AsyncMock,
            return_value=connector,
        ),
        patch(
            "app.mcp.tools.pricing_tools.get_latest_result_by_type",
            new_callable=AsyncMock,
            return_value=_cached(payload),
        ),
        patch(
            "app.mcp.tools.pricing_tools._file_svc.get_file",
            new=get_file_mock,
        ),
        patch(
            "app.mcp.tools.pricing_tools.create_spreadsheet",
            new_callable=AsyncMock,
            return_value={
                "spreadsheet_id": "ss-123",
                "url": "https://docs.google.com/spreadsheets/d/ss-123/edit",
            },
        ) as create_mock,
        patch(
            "app.mcp.tools.pricing_tools.write_range",
            new_callable=AsyncMock,
            return_value={"updated_rows": 3, "updated_columns": 3, "updated_range": "Sheet1!A1:C3"},
        ) as write_mock,
        patch(
            "app.mcp.tools.pricing_tools.share_spreadsheet",
            new_callable=AsyncMock,
            return_value={"permission_id": "p1"},
        ) as share_mock,
        patch(
            "app.mcp.tools.pricing_tools.apply_pricing_styling",
            new=styling_mock,
        ),
        patch(
            "app.mcp.tools.pricing_tools.decrypt_credentials",
            return_value={"service_account_json": {"foo": "bar"}},
        ),
        patch(
            "app.mcp.tools.pricing_tools._get_user_email",
            new_callable=AsyncMock,
            return_value="aiden@example.com",
        ),
    ):
        result = asyncio.run(pricing_to_sheets_execute(params, ctx))
    return result, create_mock, write_mock, share_mock, styling_mock


class TestConnectorCheckFirst:
    def test_no_connector_returns_actionable_error_first(self, sheets_context):
        # Even with no pricing run, the connector error MUST come first.
        result, create_mock, write_mock, _, _ = _run({}, sheets_context, connector=None, payload=None)
        assert result["error"] is True
        assert "google sheets" in result["message"].lower()
        assert "settings" in result["message"].lower()
        # Did NOT touch the Sheets API.
        create_mock.assert_not_awaited()
        write_mock.assert_not_awaited()


class TestCacheMiss:
    def test_cache_miss_returns_error_when_connector_present(self, sheets_context):
        result, create_mock, write_mock, _, _ = _run({}, sheets_context, connector=_connector(), payload=None)
        assert result["error"] is True
        assert "no pricing run" in result["message"].lower()
        create_mock.assert_not_awaited()
        write_mock.assert_not_awaited()


class TestExcelReparse:
    def test_writes_full_row_set_via_excel_reparse(self, sheets_context):
        # 50 SKUs in the Excel — Sheets should receive 50 + 1 header row.
        rows = [[f"SKU-{i}", float(100 + i), float(80 + i)] for i in range(50)]
        excel_bytes = _build_excel_bytes(["SKU", "USD", "GBP"], rows)
        payload = _payload(row_count=50)
        result, create_mock, write_mock, _, _ = _run(
            {"title": "My Pricing"},
            sheets_context,
            connector=_connector(),
            payload=payload,
            excel_bytes=excel_bytes,
        )
        assert result["success"] is True
        # write_range receives [headers, *50 data rows] = 51 rows.
        write_mock.assert_awaited_once()
        kwargs = write_mock.call_args.kwargs
        data = kwargs["data"]
        assert len(data) == 51
        assert data[0] == ["SKU", "USD", "GBP"]
        assert data[1][0] == "SKU-0"

    def test_uses_excel_file_not_cache_for_rows(self, sheets_context):
        """payload has excel_file_id + row_count but NO full_rows. Executor must
        still produce the correct row set by re-parsing the Excel file."""
        excel_bytes = _build_excel_bytes(
            ["SKU", "USD", "GBP"],
            [["A", 1.0, 1.0], ["B", 2.0, 2.0], ["C", 3.0, 3.0]],
        )
        payload = _payload(row_count=3)
        assert "full_rows" not in payload
        result, _, write_mock, _, _ = _run(
            {}, sheets_context, connector=_connector(), payload=payload, excel_bytes=excel_bytes
        )
        assert result["success"] is True
        data = write_mock.call_args.kwargs["data"]
        # 1 header + 3 data rows.
        assert len(data) == 4

    def test_excel_file_missing_returns_error(self, sheets_context):
        result, create_mock, write_mock, _, _ = _run(
            {},
            sheets_context,
            connector=_connector(),
            payload=_payload(),
            file_lookup_raises=ValueError("not found"),
        )
        assert result["error"] is True
        assert "excel" in result["message"].lower() or "file" in result["message"].lower()
        create_mock.assert_not_awaited()
        write_mock.assert_not_awaited()


class TestSheetsCreation:
    def test_default_title_when_none_given(self, sheets_context):
        result, create_mock, _, _, _ = _run({}, sheets_context, connector=_connector(), payload=_payload())
        assert result["success"] is True
        kwargs = create_mock.call_args.kwargs
        assert "Pricing Export" in kwargs["title"]

    def test_explicit_title_passed_through(self, sheets_context):
        result, create_mock, _, _, _ = _run(
            {"title": "Q1 Prices"}, sheets_context, connector=_connector(), payload=_payload()
        )
        kwargs = create_mock.call_args.kwargs
        assert kwargs["title"] == "Q1 Prices"

    def test_shared_drive_id_passed_through(self, sheets_context):
        connector = _connector(metadata={"shared_drive_id": "drive-456"})
        _, create_mock, _, _, _ = _run({}, sheets_context, connector=connector, payload=_payload())
        kwargs = create_mock.call_args.kwargs
        assert kwargs["shared_drive_id"] == "drive-456"

    def test_share_skipped_when_in_shared_drive(self, sheets_context):
        connector = _connector(metadata={"shared_drive_id": "drive-456"})
        _, _, _, share_mock, _ = _run({}, sheets_context, connector=connector, payload=_payload())
        share_mock.assert_not_awaited()

    def test_share_called_when_no_shared_drive(self, sheets_context):
        _, _, _, share_mock, _ = _run({}, sheets_context, connector=_connector(), payload=_payload())
        share_mock.assert_awaited_once()


class TestResultShape:
    def test_returns_url_and_sku_count(self, sheets_context):
        result, _, _, _, _ = _run({}, sheets_context, connector=_connector(), payload=_payload(row_count=2))
        assert result["success"] is True
        assert result["url"].startswith("https://docs.google.com/")
        assert result["sku_count"] == 2

    def test_does_not_emit_pricing_state(self, sheets_context):
        """pricing_to_sheets is read-only — must NOT return pricing_state in
        its result, otherwise the orchestrator would think it's a write tool."""
        result, _, _, _, _ = _run({}, sheets_context, connector=_connector(), payload=_payload())
        assert "pricing_state" not in result


class TestNoConversationId:
    def test_no_conversation_id_returns_error(self):
        ctx = {
            "db": AsyncMock(),
            "tenant_id": uuid.uuid4(),
            "actor_id": uuid.uuid4(),
        }
        result, _, _, _, _ = _run({}, ctx, connector=_connector(), payload=_payload())
        assert result["error"] is True


class TestApplyStyling:
    def test_styling_is_applied_after_write(self, sheets_context):
        result, _, write_mock, _, styling_mock = _run({}, sheets_context, connector=_connector(), payload=_payload())
        assert result["success"] is True
        write_mock.assert_awaited_once()
        styling_mock.assert_awaited_once()

    def test_styling_receives_headers_and_row_count(self, sheets_context):
        # 5 SKU rows → row_count passed to styling should be 5.
        rows = [[f"SKU-{i}", float(100 + i), float(80 + i)] for i in range(5)]
        excel_bytes = _build_excel_bytes(["SKU", "USD", "GBP"], rows)
        payload = _payload(row_count=5)
        _, _, _, _, styling_mock = _run(
            {}, sheets_context, connector=_connector(), payload=payload, excel_bytes=excel_bytes
        )
        styling_mock.assert_awaited_once()
        kwargs = styling_mock.call_args.kwargs
        assert kwargs["spreadsheet_id"] == "ss-123"
        assert kwargs["headers"] == ["SKU", "USD", "GBP"]
        assert kwargs["row_count"] == 5


class TestStylingFailureNonFatal:
    def test_export_succeeds_when_styling_raises(self, sheets_context):
        # If the Sheets API rejects a batchUpdate, the user still gets the URL.
        result, _, _, _, _ = _run(
            {},
            sheets_context,
            connector=_connector(),
            payload=_payload(),
            styling_side_effect=RuntimeError("styling boom"),
        )
        assert result["success"] is True
        assert "url" in result
